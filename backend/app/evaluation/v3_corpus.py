"""V3 检索语料构建器。

从黄金案例文件（PDF/Excel/Markdown）构建可复现的分块语料。
每个分块包含统一 schema：chunk_id, file_role, file_name, locator_type,
page_number, sheet_name, cell_range, text_chunk_index, section_title,
text, content_hash（完整 64 位 SHA-256）。

manifest 声明的文件缺失或 SHA256 不一致时必须失败。
"""

import hashlib
import json
from pathlib import Path
from typing import Any

import fitz
import openpyxl

# -- 分块配置版本常量 --
CHUNKING_VERSION = "1.0.0"

# PDF：每页最大字符数，超过则按段落进一步分割
PDF_PAGE_MAX_CHARS = 1200

# Excel：按连续数据行范围分块，每组最大行数
EXCEL_MAX_ROWS_PER_CHUNK = 20

# Markdown：按标题分段，每段最小字符数（短段合并到前一分块）
MD_SECTION_MIN_CHARS = 400


class CorpusBuildError(Exception):
    """语料构建失败。"""


def build_corpus(case_dir: Path) -> list[dict[str, Any]]:
    """从黄金案例目录构建检索语料。

    返回按 chunk_id 排序的分块列表。

    语料构建是可复现的：相同的 case_dir 总是产生相同的分块。
    manifest 文件缺失或 SHA256 不一致时立即失败。
    """
    if not case_dir.is_dir():
        raise CorpusBuildError(f"黄金案例目录不存在: {case_dir}")

    manifest_path = case_dir / "manifest.json"
    if not manifest_path.exists():
        raise CorpusBuildError(f"manifest.json 不存在: {manifest_path}")

    manifest_raw = manifest_path.read_text("utf-8")
    manifest = json.loads(manifest_raw)
    manifest["_sha256"] = hashlib.sha256(manifest_raw.encode("utf-8")).hexdigest()

    if not manifest.get("files"):
        raise CorpusBuildError("manifest.json 中 files 列表为空")

    chunks: list[dict[str, Any]] = []
    counter = [0]

    for file_info in manifest["files"]:
        filename = file_info["filename"]
        file_path = case_dir / filename

        if not file_path.exists():
            raise CorpusBuildError(
                f"manifest 声明的文件缺失: {filename}"
            )

        # 验证文件 SHA256
        actual_sha = _file_sha256(file_path)
        expected_sha = file_info.get("sha256", "")
        if expected_sha and actual_sha != expected_sha:
            raise CorpusBuildError(
                f"文件 SHA256 不一致: {filename}\n"
                f"  期望: {expected_sha}\n"
                f"  实际: {actual_sha}"
            )

        role = file_info.get("role", "unknown")

        if filename.endswith(".pdf"):
            chunks.extend(_parse_pdf(file_path, filename, role, file_info, counter))
        elif filename.endswith(".xlsx"):
            chunks.extend(_parse_excel(file_path, filename, role, file_info, counter))
        elif filename.endswith(".md"):
            chunks.extend(_parse_markdown(file_path, filename, role, file_info, counter))
        else:
            chunks.extend(_parse_text_file(file_path, filename, role, file_info, counter))

    chunks.sort(key=lambda c: c["chunk_id"])
    return chunks


# -- PDF 解析 --

def _parse_pdf(
    file_path: Path,
    filename: str,
    role: str,
    file_info: dict[str, Any],
    counter: list[int],
) -> list[dict[str, Any]]:
    chunks: list[dict[str, Any]] = []
    with fitz.open(file_path) as doc:
        for page_idx, page in enumerate(doc, start=1):
            text = page.get_text("text").strip()
            if not text:
                # 扫描页（无文本层）— 保留占位分块
                placeholder = f"[第{page_idx}页：扫描图片，无文本层，需要 OCR]"
                chunks.append(_make_chunk(
                    counter, filename, role,
                    locator_type="pdf_page",
                    page_number=page_idx,
                    text=placeholder,
                ))
                continue

            if len(text) <= PDF_PAGE_MAX_CHARS:
                chunks.append(_make_chunk(
                    counter, filename, role,
                    locator_type="pdf_page",
                    page_number=page_idx,
                    text=text,
                ))
            else:
                paragraphs = _split_paragraphs(text, PDF_PAGE_MAX_CHARS)
                for para_text in paragraphs:
                    chunks.append(_make_chunk(
                        counter, filename, role,
                        locator_type="pdf_page",
                        page_number=page_idx,
                        text=para_text,
                    ))
    return chunks


# -- Excel 解析 --

def _parse_excel(
    file_path: Path,
    filename: str,
    role: str,
    file_info: dict[str, Any],
    counter: list[int],
) -> list[dict[str, Any]]:
    chunks: list[dict[str, Any]] = []
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 1:
            wb.close()
            continue

        # 收集有数据的行
        data_rows: list[tuple[int, list[str]]] = []
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1
        ):
            cells = [str(cell) if cell is not None else "" for cell in row]
            has_data = any(c.strip() for c in cells)
            if has_data:
                data_rows.append((row_idx, cells))

        if not data_rows:
            wb.close()
            continue

        # 按连续行范围分块
        group_start = data_rows[0][0]
        group_rows: list[tuple[int, list[str]]] = []

        for row_idx, cells in data_rows:
            if group_rows and (row_idx - group_rows[-1][0] > 2
                               or len(group_rows) >= EXCEL_MAX_ROWS_PER_CHUNK):
                chunks.append(_emit_excel_chunk(
                    counter, filename, role, sheet_name, group_rows
                ))
                group_start = row_idx
                group_rows = [(row_idx, cells)]
            else:
                group_rows.append((row_idx, cells))

        if group_rows:
            chunks.append(_emit_excel_chunk(
                counter, filename, role, sheet_name, group_rows
            ))

    wb.close()
    return chunks


def _emit_excel_chunk(
    counter: list[int],
    filename: str,
    role: str,
    sheet_name: str,
    rows: list[tuple[int, list[str]]],
) -> dict[str, Any]:
    """生成一个 Excel 分块，带真实 cell_range。"""
    max_col = max(len(cells) for _, cells in rows)
    col_letter = _col_to_letter(max_col)
    start_row = rows[0][0]
    end_row = rows[-1][0]
    cell_range = f"A{start_row}:{col_letter}{end_row}"

    # 重建文本
    text_lines: list[str] = []
    for row_idx, cells in rows:
        text_lines.append("|".join(cells))
    text = "\n".join(text_lines)

    return _make_chunk(
        counter, filename, role,
        locator_type="spreadsheet_cell",
        page_number=None,
        sheet_name=sheet_name,
        cell_range=cell_range,
        text=text,
    )


# -- Markdown 解析 --

def _parse_markdown(
    file_path: Path,
    filename: str,
    role: str,
    file_info: dict[str, Any],
    counter: list[int],
) -> list[dict[str, Any]]:
    text = file_path.read_text("utf-8")
    return _split_markdown(text, filename, role, counter)


def _split_markdown(
    text: str,
    filename: str,
    role: str,
    counter: list[int],
) -> list[dict[str, Any]]:
    """按 Markdown 标题分割，短节合并。使用稳定的 text_chunk_index。"""
    lines = text.split("\n")

    sections: list[tuple[str | None, list[str]]] = []
    current_title: str | None = None
    current_lines: list[str] = []

    for line in lines:
        stripped = line.strip()
        if stripped.startswith("#"):
            if current_lines:
                sections.append((current_title, current_lines))
            current_title = stripped.lstrip("#").strip()
            current_lines = [line]
        else:
            current_lines.append(line)
    if current_lines:
        sections.append((current_title, current_lines))

    # 合并过短的节
    merged: list[tuple[str | None, str]] = []
    buf_lines: list[str] = []
    buf_title: str | None = None

    for title, sec_lines in sections:
        buf_lines.extend(sec_lines)
        if buf_title is None and title is not None:
            buf_title = title
        combined = "\n".join(buf_lines).strip()
        if len(combined) >= MD_SECTION_MIN_CHARS:
            merged.append((buf_title, combined))
            buf_lines = []
            buf_title = None

    if buf_lines:
        combined = "\n".join(buf_lines).strip()
        if combined:
            merged.append((buf_title, combined))

    chunks: list[dict[str, Any]] = []
    for idx, (title, chunk_text) in enumerate(merged):
        chunks.append(_make_chunk(
            counter, filename, role,
            locator_type="text_chunk",
            page_number=None,
            section_title=title,
            text_chunk_index=idx,
            text=chunk_text,
        ))
    return chunks


# -- 纯文本文件 --

def _parse_text_file(
    file_path: Path,
    filename: str,
    role: str,
    file_info: dict[str, Any],
    counter: list[int],
) -> list[dict[str, Any]]:
    try:
        text = file_path.read_text("utf-8")
    except UnicodeDecodeError:
        text = file_path.read_text("latin-1")
    return [_make_chunk(
        counter, filename, role,
        locator_type="text_chunk",
        page_number=None,
        text_chunk_index=0,
        text=text,
    )]


# -- 内部工具 --

def _make_chunk(
    counter: list[int],
    filename: str,
    role: str,
    *,
    locator_type: str,
    page_number: int | None = None,
    sheet_name: str | None = None,
    cell_range: str | None = None,
    text_chunk_index: int | None = None,
    section_title: str | None = None,
    text: str,
) -> dict[str, Any]:
    counter[0] += 1
    content_hash = hashlib.sha256(text.encode("utf-8")).hexdigest()
    return {
        "chunk_id": f"C{counter[0]:04d}",
        "file_role": role,
        "file_name": filename,
        "locator_type": locator_type,
        "page_number": page_number,
        "sheet_name": sheet_name,
        "cell_range": cell_range,
        "text_chunk_index": text_chunk_index,
        "section_title": section_title,
        "text": text,
        "content_hash": content_hash,
    }


def _file_sha256(file_path: Path) -> str:
    """计算文件的完整 SHA-256。"""
    hasher = hashlib.sha256()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _split_paragraphs(text: str, max_chars: int) -> list[str]:
    """按双换行分割，短段合并。"""
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    result: list[str] = []
    buf: list[str] = []
    buf_len = 0
    for para in paragraphs:
        if buf and buf_len + len(para) > max_chars:
            result.append("\n\n".join(buf))
            buf = []
            buf_len = 0
        buf.append(para)
        buf_len += len(para)
    if buf:
        result.append("\n\n".join(buf))
    return result


def _col_to_letter(col: int) -> str:
    """列号转 Excel 字母 (1->A, 27->AA)。"""
    result = ""
    while col > 0:
        col -= 1
        result = chr(col % 26 + ord("A")) + result
        col //= 26
    return result
