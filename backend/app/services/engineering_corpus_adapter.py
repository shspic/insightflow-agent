"""真实 Corpus 适配器。

从实际上传文件读取内容并生成确定性 CorpusChunk，不依赖预先存在的 FileChunk。
支持 PDF（文本+OCR）、Excel、Markdown 三种文件类型。

locator_type 统一为三类：
    - pdf_page: PDF 文本或 OCR（含真实 page_number）
    - spreadsheet_cell: Excel 工作表单元格范围（含 sheet_name/cell_range）
    - text_chunk: Markdown 文本块（含 section_path）
"""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from app.retrieval.schemas import (
    ENGINEERING_ROLES,
    CHUNK_ID_FORMAT,
    CorpusChunk,
)

# 分块参数
DEFAULT_CHUNK_SIZE = 800
DEFAULT_CHUNK_OVERLAP = 100


# ── PDF 适配 ─────────────────────────────────────────────────────────


def _read_pdf_chunks(
    file_path: Path,
    workspace_id: int,
    owner_user_id: int,
    file_id: int,
    file_name: str,
    file_role: str,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    chunk_overlap: int = DEFAULT_CHUNK_OVERLAP,
) -> list[CorpusChunk]:
    """使用 PyMuPDF 读取真实 PDF，按页解析并按固定窗口分块。"""
    import fitz  # PyMuPDF

    chunks: list[CorpusChunk] = []
    chunk_index = 0

    doc = fitz.open(str(file_path))
    try:
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text()
            if not text or not text.strip():
                continue

            # 每页内按滑动窗口分块
            page_chunks = _sliding_window_chunks(text, chunk_size, chunk_overlap)
            for pc_text in page_chunks:
                cid = CorpusChunk.make_chunk_id(workspace_id, file_id, chunk_index)
                content_hash = CorpusChunk.compute_content_hash(pc_text)
                chunks.append(
                    CorpusChunk(
                        chunk_id=cid,
                        workspace_id=workspace_id,
                        owner_user_id=owner_user_id,
                        file_id=file_id,
                        file_name=file_name,
                        file_role=file_role,
                        locator_type="pdf_page",
                        text=pc_text,
                        content_hash=content_hash,
                        page_number=page_num + 1,
                        text_chunk_index=chunk_index,
                        parser_name="pymupdf",
                        parser_version="1.0.0",
                    )
                )
                chunk_index += 1
    finally:
        doc.close()

    return chunks


def _read_pdf_ocr_chunks(
    file_path: Path,
    workspace_id: int,
    owner_user_id: int,
    file_id: int,
    file_name: str,
    file_role: str,
    existing_ocr_chunks: list[dict[str, Any]],
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    chunk_overlap: int = DEFAULT_CHUNK_OVERLAP,
) -> list[CorpusChunk]:
    """从已有 OCR FileChunk 中读取文本并构建 CorpusChunk。

    existing_ocr_chunks: [{"page_number": int, "chunk_text": str, "parser_version": str}, ...]
    """
    chunks: list[CorpusChunk] = []
    for fc in existing_ocr_chunks:
        text = fc.get("chunk_text", "") or ""
        if not text.strip():
            continue
        idx = fc.get("chunk_index", len(chunks))
        page_num = fc.get("page_number")
        pv = fc.get("parser_version", "1.0.0-ocr")
        cid = CorpusChunk.make_chunk_id(workspace_id, file_id, idx)
        content_hash = CorpusChunk.compute_content_hash(text)
        chunks.append(
            CorpusChunk(
                chunk_id=cid,
                workspace_id=workspace_id,
                owner_user_id=owner_user_id,
                file_id=file_id,
                file_name=file_name,
                file_role=file_role,
                locator_type="pdf_page",
                text=text,
                content_hash=content_hash,
                page_number=page_num,
                text_chunk_index=idx,
                parser_name="pymupdf+ocr",
                parser_version=pv,
            )
        )
    return chunks


# ── Excel 适配 ────────────────────────────────────────────────────────


def _read_excel_chunks(
    file_path: Path,
    workspace_id: int,
    owner_user_id: int,
    file_id: int,
    file_name: str,
    file_role: str,
    rows_per_chunk: int = 50,
) -> list[CorpusChunk]:
    """使用 openpyxl 读取真实 Excel，按工作表+行区间确定性分块。"""
    from openpyxl import load_workbook

    chunks: list[CorpusChunk] = []
    chunk_index = 0

    wb = load_workbook(str(file_path), data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if all(v is None for v in (row or [])):
                    continue
                rows.append([str(v) if v is not None else "" for v in row])

            if not rows:
                continue

            # 按固定行区间分块
            for start in range(0, len(rows), rows_per_chunk):
                end = min(start + rows_per_chunk, len(rows))
                slice_rows = rows[start:end]
                lines = []
                for ri, row in enumerate(slice_rows):
                    lines.append("\t".join(row))
                text = "\n".join(lines)

                ncols = max(len(r) for r in slice_rows) if slice_rows else 1
                cell_range = f"A{start + 1}:{_col_letter(ncols)}{end}"

                cid = CorpusChunk.make_chunk_id(workspace_id, file_id, chunk_index)
                content_hash = CorpusChunk.compute_content_hash(text)
                chunks.append(
                    CorpusChunk(
                        chunk_id=cid,
                        workspace_id=workspace_id,
                        owner_user_id=owner_user_id,
                        file_id=file_id,
                        file_name=file_name,
                        file_role=file_role,
                        locator_type="spreadsheet_cell",
                        text=text,
                        content_hash=content_hash,
                        sheet_name=sheet_name,
                        cell_range=cell_range,
                        text_chunk_index=chunk_index,
                        parser_name="openpyxl",
                        parser_version="1.0.0",
                    )
                )
                chunk_index += 1
    finally:
        wb.close()

    return chunks


def _col_letter(n: int) -> str:
    """列号→列字母（1→A, 27→AA）。"""
    result = ""
    while n > 0:
        n -= 1
        result = chr(ord("A") + n % 26) + result
        n //= 26
    return result


# ── Markdown 适配 ─────────────────────────────────────────────────────


def _read_markdown_chunks(
    file_path: Path,
    workspace_id: int,
    owner_user_id: int,
    file_id: int,
    file_name: str,
    file_role: str,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    chunk_overlap: int = DEFAULT_CHUNK_OVERLAP,
) -> list[CorpusChunk]:
    """读取真实 Markdown 文件，按标题感知+固定窗口分块。"""
    try:
        text = file_path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        text = file_path.read_text(encoding="gb18030")

    if not text or not text.strip():
        return []

    chunks: list[CorpusChunk] = []
    sections = _split_by_headings(text)

    chunk_index = 0
    for section_path, section_text in sections:
        section_chunks = _sliding_window_chunks(section_text, chunk_size, chunk_overlap)
        for sc_text in section_chunks:
            cid = CorpusChunk.make_chunk_id(workspace_id, file_id, chunk_index)
            content_hash = CorpusChunk.compute_content_hash(sc_text)
            chunks.append(
                CorpusChunk(
                    chunk_id=cid,
                    workspace_id=workspace_id,
                    owner_user_id=owner_user_id,
                    file_id=file_id,
                    file_name=file_name,
                    file_role=file_role,
                    locator_type="text_chunk",
                    text=sc_text,
                    content_hash=content_hash,
                    section_path=section_path or None,
                    text_chunk_index=chunk_index,
                    parser_name="markdown_parser",
                    parser_version="1.0.0",
                )
            )
            chunk_index += 1

    return chunks


def _split_by_headings(text: str) -> list[tuple[str, str]]:
    """按 Markdown 标题分节。返回 [(section_path, section_text), ...]。"""
    import re

    heading_re = re.compile(r"^(#{1,6})\s+(.+)$", re.MULTILINE)
    matches = list(heading_re.finditer(text))

    if not matches:
        return [("", text)]

    sections: list[tuple[str, str]] = []
    for i, m in enumerate(matches):
        level = len(m.group(1))
        title = m.group(2).strip()
        start = m.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        section_text = text[start:end].strip()
        if section_text:
            sections.append((f"H{level} > {title}", section_text))

    # also include text before first heading
    if matches and matches[0].start() > 0:
        pre_text = text[:matches[0].start()].strip()
        if pre_text:
            sections.insert(0, ("", pre_text))

    return sections


# ── 滑动窗口分块 ──────────────────────────────────────────────────────


def _sliding_window_chunks(
    text: str, chunk_size: int, chunk_overlap: int
) -> list[str]:
    """字符级滑动窗口分块。"""
    if len(text) <= chunk_size:
        return [text]

    chunks: list[str] = []
    start = 0
    while start < len(text):
        end = min(start + chunk_size, len(text))
        chunks.append(text[start:end])
        start += chunk_size - chunk_overlap

    return chunks


# ── 主入口 ────────────────────────────────────────────────────────────


def build_corpus_from_files(
    *,
    workspace_id: int,
    owner_user_id: int,
    files_info: list[dict[str, Any]],
) -> tuple[list[CorpusChunk], list[str]]:
    """从真实文件构建 Corpus。

    files_info 每条必须包含：
        - file_id: int
        - file_name: str
        - file_type: str (pdf/xlsx/csv/md/markdown)
        - file_path: str (磁盘绝对/相对路径)
        - confirmed_role: str (必须是 ENGINEERING_ROLES 之一)
        - ocr_chunks: list[dict] | None (扫描 PDF 的 OCR 数据)

    返回 (chunks, warnings)。
    """
    warnings: list[str] = []
    all_chunks: list[CorpusChunk] = []

    for finfo in files_info:
        file_id = finfo["file_id"]
        file_name = finfo["file_name"]
        file_type = (finfo.get("file_type") or "").lower()
        file_path_str = finfo.get("file_path", "")
        confirmed_role = finfo.get("confirmed_role", "")
        ocr_chunks = finfo.get("ocr_chunks")

        if confirmed_role not in ENGINEERING_ROLES:
            warnings.append(
                f"文件 {file_id} ({file_name}) 角色 '{confirmed_role}' 不在工程检索白名单，跳过"
            )
            continue

        file_path = _resolve_path(file_path_str)
        if file_path is None:
            warnings.append(
                f"file:{file_id}: 文件路径不可安全读取，已跳过"
            )
            continue

        try:
            if file_type == "pdf":
                # 逐页合并文本层和 OCR 层
                # Step 1: 提取文本层
                text_chunks = _read_pdf_chunks(
                    file_path=file_path,
                    workspace_id=workspace_id,
                    owner_user_id=owner_user_id,
                    file_id=file_id,
                    file_name=file_name,
                    file_role=confirmed_role,
                )

                # 记录有文本的页面
                pages_with_text: set[int] = set()
                for tc in text_chunks:
                    if tc.page_number is not None:
                        pages_with_text.add(tc.page_number)

                # Step 2: 提取 OCR 层（仅当存在 OCR 数据时）
                ocr_corpus_chunks: list[CorpusChunk] = []
                if ocr_chunks:
                    ocr_corpus_chunks = _read_pdf_ocr_chunks(
                        file_path=file_path,
                        workspace_id=workspace_id,
                        owner_user_id=owner_user_id,
                        file_id=file_id,
                        file_name=file_name,
                        file_role=confirmed_role,
                        existing_ocr_chunks=ocr_chunks,
                    )

                # Step 3: 逐页合并 — 文本优先，OCR 回退
                merged: list[CorpusChunk] = list(text_chunks)

                for oc in ocr_corpus_chunks:
                    if oc.page_number is None or oc.page_number not in pages_with_text:
                        merged.append(oc)
                    # 该页已有文本，跳过 OCR → 不重复

                # Step 4: 重新分配稳定 chunk index
                if merged:
                    merged.sort(key=lambda c: (c.page_number or 0, c.chunk_id))
                    for i, chunk in enumerate(merged):
                        chunk.text_chunk_index = i
                        chunk.chunk_id = CorpusChunk.make_chunk_id(
                            workspace_id, file_id, i
                        )
                    all_chunks.extend(merged)
                else:
                    warnings.append(
                        f"file:{file_id}: PDF 无可提取文本内容，已跳过"
                    )
                continue  # PDF 已在分支内完成处理
            elif file_type in ("xlsx", "csv"):
                if file_type == "csv":
                    chunks = _read_csv_chunks(
                        file_path=file_path,
                        workspace_id=workspace_id,
                        owner_user_id=owner_user_id,
                        file_id=file_id,
                        file_name=file_name,
                        file_role=confirmed_role,
                    )
                else:
                    chunks = _read_excel_chunks(
                        file_path=file_path,
                        workspace_id=workspace_id,
                        owner_user_id=owner_user_id,
                        file_id=file_id,
                        file_name=file_name,
                        file_role=confirmed_role,
                    )
            elif file_type in ("md", "markdown"):
                chunks = _read_markdown_chunks(
                    file_path=file_path,
                    workspace_id=workspace_id,
                    owner_user_id=owner_user_id,
                    file_id=file_id,
                    file_name=file_name,
                    file_role=confirmed_role,
                )
            else:
                warnings.append(
                    f"文件 {file_id} ({file_name}) 类型 '{file_type}' 不支持检索分块，跳过"
                )
                continue

            if not chunks:
                warnings.append(
                    f"文件 {file_id} ({file_name}) 无可提取文本内容，跳过"
                )
                continue

            all_chunks.extend(chunks)

        except Exception as e:
            warnings.append(
                f"文件 {file_id} ({file_name}) 分块失败: 内部错误，跳过"
            )
            continue

    # 按 chunk_id 排序
    all_chunks.sort(key=lambda c: c.chunk_id)

    if not all_chunks:
        return [], warnings + ["没有可用于检索的语料内容"]

    return all_chunks, warnings


def _read_csv_chunks(
    file_path: Path,
    workspace_id: int,
    owner_user_id: int,
    file_id: int,
    file_name: str,
    file_role: str,
    rows_per_chunk: int = 50,
) -> list[CorpusChunk]:
    """读取 CSV 文件，按行区间分块。"""
    import csv

    rows: list[list[str]] = []
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with open(file_path, "r", encoding=encoding, newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    rows.append([v or "" for v in row])
            break
        except (UnicodeDecodeError, Exception):
            continue

    if not rows:
        return []

    chunks: list[CorpusChunk] = []
    chunk_index = 0
    for start in range(0, len(rows), rows_per_chunk):
        end = min(start + rows_per_chunk, len(rows))
        slice_rows = rows[start:end]
        lines = ["\t".join(row) for row in slice_rows]
        text = "\n".join(lines)

        ncols = max(len(r) for r in slice_rows) if slice_rows else 1
        cell_range = f"A{start + 1}:{_col_letter(ncols)}{end}"

        cid = CorpusChunk.make_chunk_id(workspace_id, file_id, chunk_index)
        content_hash = CorpusChunk.compute_content_hash(text)
        chunks.append(
            CorpusChunk(
                chunk_id=cid,
                workspace_id=workspace_id,
                owner_user_id=owner_user_id,
                file_id=file_id,
                file_name=file_name,
                file_role=file_role,
                locator_type="spreadsheet_cell",
                text=text,
                content_hash=content_hash,
                sheet_name=file_name,
                cell_range=cell_range,
                text_chunk_index=chunk_index,
                parser_name="csv_parser",
                parser_version="1.0.0",
            )
        )
        chunk_index += 1
    return chunks


def _resolve_path(path_str: str) -> Path | None:
    """安全解析文件路径：必须在 upload root 内，拒绝越狱。"""
    if not path_str:
        return None

    from app.core.config import settings

    backend = Path(__file__).resolve().parents[2]
    upload_root = (backend / settings.upload_dir).resolve()

    p = Path(path_str)
    if not p.is_absolute():
        resolved = (backend / p).resolve()
    else:
        resolved = p.resolve()

    # 安全检查：必须在 upload_root 内
    try:
        resolved.relative_to(upload_root)
    except ValueError:
        return None

    # 不能是 upload_root 本身
    if resolved == upload_root:
        return None

    # 必须是普通文件
    if not resolved.is_file():
        return None

    # 不能是符号链接
    if resolved.is_symlink():
        return None

    return resolved
