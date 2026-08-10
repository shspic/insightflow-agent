"""Evidence 来源完整性（阶段 6A 补修）。

证据的「来源」与「记录本身」是两个独立概念：
- Evidence.content_hash = 证据记录规范哈希（元数据 JSON SHA-256，见 review_engine_service）；
- 本模块提供来源文件字节哈希与真实 locator 校验，由服务端对安全解析后的
  当前文件/Corpus 计算，禁止信任 API 客户端传入的哈希。

安全边界：
- 文件路径必须落在 upload root 内，拒绝符号链接与越狱路径；
- 失败时返回 None（由调用方决定 stale / 阻断），错误信息不含磁盘路径。
"""

from __future__ import annotations

import hashlib
import re
from pathlib import Path

# 来源类型
FIELD_LOCATOR = "field_locator"
CORPUS_CHUNK = "corpus_chunk"
PROVENANCE_TYPES = frozenset({FIELD_LOCATOR, CORPUS_CHUNK})


def resolve_upload_file(file_path: str) -> Path | None:
    """安全解析上传文件：必须在 upload root 内、非符号链接、存在且为普通文件。"""
    if not file_path:
        return None
    from app.core.config import settings

    backend = Path(__file__).resolve().parents[2]
    upload_root = (backend / settings.upload_dir).resolve()

    p = Path(file_path)
    if not p.is_absolute():
        resolved = (backend / p).resolve()
    else:
        resolved = p.resolve()

    try:
        resolved.relative_to(upload_root)
    except ValueError:
        return None
    if resolved == upload_root:
        return None
    if not resolved.is_file():
        return None
    if resolved.is_symlink():
        return None
    return resolved


def compute_file_sha256_safe(file_path: str) -> str | None:
    """来源文件字节 SHA-256；无法安全读取时返回 None。"""
    resolved = resolve_upload_file(file_path)
    if resolved is None:
        return None
    try:
        return hashlib.sha256(resolved.read_bytes()).hexdigest()
    except OSError:
        return None


def validate_pdf_page(file_path: str, page_number: int | None) -> bool:
    """PDF 校验 page 存在（1 ≤ page ≤ 当前文件页数）。"""
    if not isinstance(page_number, int) or page_number < 1:
        return False
    resolved = resolve_upload_file(file_path)
    if resolved is None:
        return False
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(str(resolved))
        try:
            return page_number <= doc.page_count
        finally:
            doc.close()
    except Exception:
        return False


def validate_spreadsheet_cell(file_path: str, sheet_name: str | None, cell_range: str | None) -> bool:
    """Excel 校验 sheet/cell 存在（锚点单元格在当前工作表行列范围内）。"""
    if not sheet_name or not cell_range:
        return False
    resolved = resolve_upload_file(file_path)
    if resolved is None:
        return False
    anchor = _cell_range_anchor(cell_range)
    if anchor is None:
        return False
    row, col = anchor
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(resolved), data_only=True, read_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]
            return row <= (ws.max_row or 0) and col <= (ws.max_column or 0)
        finally:
            wb.close()
    except Exception:
        return False


def _cell_range_anchor(cell_range: str) -> tuple[int, int] | None:
    """从 cell_range（"B3" 或 "A1:B8"）解析锚点单元格 (row, col)；失败返回 None。"""
    m = re.fullmatch(r"([A-Za-z]+)(\d+)(?::[A-Za-z]+\d+)?", cell_range.strip())
    if m is None:
        return None
    col = 0
    for ch in m.group(1).upper():
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return int(m.group(2)), col


def validate_text_chunk_index(file_path: str, chunk_id: int | None) -> bool:
    """text_chunk 校验块编号存在（0-based：0 ≤ chunk_id < 当前文件块数）。"""
    if not isinstance(chunk_id, int) or isinstance(chunk_id, bool) or chunk_id < 0:
        return False
    resolved = resolve_upload_file(file_path)
    if resolved is None:
        return False
    from app.services.engineering_corpus_adapter import (
        DEFAULT_CHUNK_SIZE,
        DEFAULT_CHUNK_OVERLAP,
        _read_markdown_chunks,
    )

    try:
        chunks = _read_markdown_chunks(
            file_path=resolved,
            workspace_id=0,
            owner_user_id=0,
            file_id=0,
            file_name="",
            file_role="",
            chunk_size=DEFAULT_CHUNK_SIZE,
            chunk_overlap=DEFAULT_CHUNK_OVERLAP,
        )
    except Exception:
        return False
    return chunk_id < len(chunks)
