import csv
import hashlib
import json
import re
import time
from datetime import date, datetime
from app.core.timeutils import utcnow
from pathlib import Path
from typing import Any, Literal

import fitz
import pandas as pd
from pydantic import BaseModel, Field, ValidationError
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.file import File
from app.models.file_chunk import FileChunk
from app.models.file_processing_run import FileProcessingRun
from app.models.file_profile import FileProfile
from app.models.workspace_file import WorkspaceFile
from app.schemas.file_understanding import FileProfileResponse, FileUnderstandOptions
from app.services.audit_service import add_audit_log
from app.services.llm_service import call_llm, safe_json_dumps
from app.services.ocr_service import (
    FileOcrError,
    extract_scanned_pdf_pages,
    extract_text_from_image,
)
from app.services.rag_service import RagServiceError, index_pdf_file
from app.services.workspace_service import get_owned_workspace_file, safe_public_text


PROFILE_VERSION = "2.0"
PARSER_VERSION = "v2.03.1"
PROMPT_VERSION = "file-understanding-v1"
SUPPORTED_TYPES = {"csv", "xlsx", "pdf", "png", "jpg", "jpeg", "webp", "md", "markdown"}
BUILT_IN_ROLES = {
    "primary_dataset",
    "supplementary_dataset",
    "rule_document",
    "reference_document",
    "resume",
    "job_description",
    "research_material",
    "image_evidence",
    "report_template",
    "supporting_material",
    "unknown",
    "custom",
}
ROLE_TEXT_PATTERN = re.compile(r"^[\w\u4e00-\u9fff .·()（）/_-]{1,60}$")
CONTROL_CHAR_PATTERN = re.compile(r"[\x00-\x1f\x7f]")
MARKDOWN_LINK_PATTERN = re.compile(r"!?\[[^\]]*]\(([^)]+)\)")
MARKDOWN_HEADING_PATTERN = re.compile(r"^(#{1,6})\s+(.+?)\s*$")
CODE_FENCE_PATTERN = re.compile(r"```.*?```|~~~.*?~~~", re.DOTALL)


class FileUnderstandingError(Exception):
    def __init__(self, message: str, code: str = "FILE_UNDERSTANDING_FAILED") -> None:
        self.message = message
        self.code = code
        super().__init__(message)


class SemanticEnhancement(BaseModel):
    title: str | None = Field(default=None, max_length=500)
    summary: str = Field(min_length=1, max_length=2000)
    suggested_role: Literal[
        "primary_dataset",
        "supplementary_dataset",
        "rule_document",
        "reference_document",
        "resume",
        "job_description",
        "research_material",
        "image_evidence",
        "report_template",
        "supporting_material",
        "unknown",
    ]
    tags: list[str] = Field(default_factory=list, max_length=12)
    confidence: float = Field(ge=0, le=1)


def understand_file(
    db: Session,
    *,
    file_id: int,
    workspace_id: int,
    owner_user_id: int,
    options: FileUnderstandOptions | None = None,
) -> FileProfile:
    options = options or FileUnderstandOptions()
    file_record = get_owned_workspace_file(
        db,
        workspace_id=workspace_id,
        file_id=file_id,
        owner_user_id=owner_user_id,
    )
    if file_record is None:
        raise FileUnderstandingError("文件不存在或无权访问", "FILE_NOT_FOUND")
    association = _get_workspace_file(db, workspace_id, file_id)
    next_version = (
        db.scalar(
            select(func.max(FileProfile.profile_version)).where(
                FileProfile.workspace_id == workspace_id,
                FileProfile.file_id == file_id,
            )
        )
        or 0
    ) + 1
    profile = FileProfile(
        file_id=file_id,
        workspace_id=workspace_id,
        owner_user_id=owner_user_id,
        profile_version=next_version,
        status="validating",
        confirmed_role=association.user_confirmed_role,
        fallback_used=False,
    )
    db.add(profile)
    db.flush()
    run = FileProcessingRun(
        file_id=file_id,
        workspace_id=workspace_id,
        owner_user_id=owner_user_id,
        profile_id=profile.id,
        stage="validating",
        status="running",
        processor="unified_file_understanding",
        retry_count=max(0, next_version - 1),
    )
    db.add(run)
    file_record.status = "validating"
    db.commit()
    started = time.perf_counter()

    try:
        file_type = (file_record.file_type or "").lower()
        if file_type not in SUPPORTED_TYPES:
            raise FileUnderstandingError("当前文件类型不支持统一理解", "UNSUPPORTED_FILE_TYPE")
        file_path = Path(file_record.file_path)
        if not file_path.exists() or not file_path.is_file():
            raise FileUnderstandingError("文件内容不存在", "FILE_CONTENT_MISSING")

        _set_stage(db, file_record, profile, run, "parsing")
        parsed = _parse_deterministically(db, file_record, options)
        _set_stage(db, file_record, profile, run, "profiling")

        semantic = None
        model_latency_ms = None
        model_message = None
        if options.use_deepseek:
            semantic, model_latency_ms, model_message = _semantic_enhancement(parsed)

        summary = parsed["summary"]
        title = parsed.get("title")
        suggested_role = parsed["suggested_role"]
        system_tags = parsed["system_tags"]
        confidence = parsed["confidence"]
        fallback_used = bool(parsed.get("fallback_used"))
        if semantic is not None:
            summary = semantic.summary
            title = semantic.title or title
            suggested_role = semantic.suggested_role
            system_tags = _normalize_tags([*system_tags, *semantic.tags], max_count=12)
            confidence = semantic.confidence
        elif options.use_deepseek:
            fallback_used = True
            if model_message:
                parsed["quality_issues"].append(
                    _quality_issue(
                        "MODEL_FALLBACK",
                        "info",
                        "语义增强不可用，已使用确定性结果。",
                    )
                )

        profile.status = "ready"
        profile.file_category = parsed["file_category"]
        profile.detected_mime_type = file_record.mime_type
        profile.language = parsed.get("language")
        profile.title = title
        profile.summary = summary
        profile.structure_json = _json_dumps(parsed["structure"])
        profile.statistics_json = _json_dumps(parsed["statistics"])
        profile.quality_issues_json = _json_dumps(parsed["quality_issues"])
        profile.suggested_role = suggested_role
        profile.confirmed_role = association.user_confirmed_role
        profile.tags_json = _json_dumps(system_tags)
        profile.confidence = confidence
        profile.parser_name = parsed["parser_name"]
        profile.parser_version = PARSER_VERSION
        profile.model_provider = settings.llm_provider if options.use_deepseek else None
        profile.model_name = settings.llm_model if options.use_deepseek else None
        profile.prompt_version = PROMPT_VERSION if options.use_deepseek else None
        profile.model_latency_ms = model_latency_ms
        profile.fallback_used = fallback_used
        profile.completed_at = utcnow()
        association.file_role = suggested_role
        file_record.status = "ready"
        file_record.summary = summary
        run.stage = "profiling"
        run.status = "success"
        run.duration_ms = int((time.perf_counter() - started) * 1000)
        run.used_model = semantic is not None
        run.fallback_used = fallback_used
        run.completed_at = utcnow()
        db.commit()
        db.refresh(profile)
        return profile
    except FileUnderstandingError as exc:
        is_unsupported = exc.code == "UNSUPPORTED_FILE_TYPE"
        _finish_failed_profile(
            db,
            file_record=file_record,
            profile=profile,
            run=run,
            error_code=exc.code,
            error_message=exc.message,
            started=started,
            unsupported=is_unsupported,
        )
        return profile
    except Exception as exc:
        _finish_failed_profile(
            db,
            file_record=file_record,
            profile=profile,
            run=run,
            error_code="FILE_UNDERSTANDING_FAILED",
            error_message=str(exc),
            started=started,
            unsupported=False,
        )
        return profile


def get_latest_profile(
    db: Session,
    *,
    workspace_id: int,
    file_id: int,
    owner_user_id: int,
    ready_only: bool = False,
) -> FileProfile | None:
    filters = [
        FileProfile.workspace_id == workspace_id,
        FileProfile.file_id == file_id,
        FileProfile.owner_user_id == owner_user_id,
    ]
    if ready_only:
        filters.append(FileProfile.status == "ready")
    return db.scalar(
        select(FileProfile)
        .where(*filters)
        .order_by(FileProfile.profile_version.desc())
        .limit(1)
    )


def list_profile_versions(
    db: Session,
    *,
    workspace_id: int,
    file_id: int,
    owner_user_id: int,
) -> list[FileProfile]:
    return list(
        db.scalars(
            select(FileProfile)
            .where(
                FileProfile.workspace_id == workspace_id,
                FileProfile.file_id == file_id,
                FileProfile.owner_user_id == owner_user_id,
            )
            .order_by(FileProfile.profile_version.desc())
        ).all()
    )


def update_profile_confirmation(
    db: Session,
    *,
    workspace_id: int,
    file_id: int,
    owner_user_id: int,
    confirmed_role: str | None,
    custom_role: str | None,
    user_tags: list[str] | None,
) -> FileProfile:
    file_record = get_owned_workspace_file(
        db,
        workspace_id=workspace_id,
        file_id=file_id,
        owner_user_id=owner_user_id,
    )
    if file_record is None:
        raise FileUnderstandingError("文件不存在或无权访问", "FILE_NOT_FOUND")
    profile = get_latest_profile(
        db,
        workspace_id=workspace_id,
        file_id=file_id,
        owner_user_id=owner_user_id,
    )
    if profile is None:
        raise FileUnderstandingError("文件尚无理解结果", "PROFILE_NOT_FOUND")
    association = _get_workspace_file(db, workspace_id, file_id)
    if confirmed_role is not None:
        normalized_role = _normalize_confirmed_role(confirmed_role, custom_role)
        association.user_confirmed_role = normalized_role
        profile.confirmed_role = normalized_role
    if user_tags is not None:
        association.tags_json = _json_dumps(_normalize_tags(user_tags, max_count=20))
    add_audit_log(
        db,
        user_id=owner_user_id,
        action="file.profile.confirmation.update",
        resource_type="file",
        resource_id=file_id,
        status="success",
        details={
            "workspace_id": workspace_id,
            "confirmed_role": association.user_confirmed_role,
            "user_tag_count": len(_json_list(association.tags_json)),
        },
    )
    db.commit()
    db.refresh(profile)
    return profile


def profile_response(
    profile: FileProfile,
    association: WorkspaceFile,
) -> FileProfileResponse:
    confirmed_role = association.user_confirmed_role or profile.confirmed_role
    return FileProfileResponse(
        id=profile.id,
        file_id=profile.file_id,
        workspace_id=profile.workspace_id,
        profile_version=profile.profile_version,
        status=profile.status,
        file_category=profile.file_category,
        detected_mime_type=profile.detected_mime_type,
        language=profile.language,
        title=profile.title,
        summary=safe_public_text(profile.summary),
        structure=_json_dict(profile.structure_json),
        statistics=_json_dict(profile.statistics_json),
        quality_issues=_json_object_list(profile.quality_issues_json),
        suggested_role=profile.suggested_role,
        confirmed_role=confirmed_role,
        effective_role=confirmed_role or profile.suggested_role,
        system_tags=_json_list(profile.tags_json),
        user_tags=_json_list(association.tags_json),
        confidence=profile.confidence,
        parser_name=profile.parser_name,
        parser_version=profile.parser_version,
        model_provider=profile.model_provider,
        model_name=profile.model_name,
        prompt_version=profile.prompt_version,
        model_latency_ms=profile.model_latency_ms,
        fallback_used=profile.fallback_used,
        error_code=profile.error_code,
        error_message=safe_public_text(profile.error_message),
        created_at=profile.created_at,
        updated_at=profile.updated_at,
        completed_at=profile.completed_at,
    )


def get_workspace_file_association(
    db: Session,
    *,
    workspace_id: int,
    file_id: int,
) -> WorkspaceFile:
    return _get_workspace_file(db, workspace_id, file_id)


def _parse_deterministically(
    db: Session,
    file_record: File,
    options: FileUnderstandOptions,
) -> dict[str, Any]:
    file_type = (file_record.file_type or "").lower()
    path = Path(file_record.file_path)
    if file_type == "csv":
        return _parse_csv(path, file_record.filename)
    if file_type == "xlsx":
        return _parse_xlsx(path, file_record.filename)
    if file_type == "pdf":
        return _parse_pdf(db, file_record, path, run_ocr=options.run_ocr)
    if file_type in {"png", "jpg", "jpeg", "webp"}:
        return _parse_image(file_record, path, options.run_ocr)
    return _parse_markdown(db, file_record, path)


def _parse_csv(path: Path, filename: str) -> dict[str, Any]:
    encoding, delimiter, encoding_issue = _detect_csv_format(path)
    try:
        sample = pd.read_csv(
            path,
            encoding=encoding,
            sep=delimiter,
            nrows=max(1, settings.profile_sample_rows),
            on_bad_lines="warn",
        )
    except Exception as exc:
        raise FileUnderstandingError(f"CSV 读取失败：{exc}", "CSV_PARSE_FAILED") from exc

    row_count = 0
    missing_counts = {str(column): 0 for column in sample.columns}
    duplicate_rows = 0
    try:
        for chunk in pd.read_csv(
            path,
            encoding=encoding,
            sep=delimiter,
            chunksize=max(1000, settings.profile_sample_rows),
            on_bad_lines="warn",
        ):
            row_count += int(len(chunk))
            for column in chunk.columns:
                missing_counts[str(column)] = missing_counts.get(str(column), 0) + int(
                    chunk[column].isna().sum()
                )
            duplicate_rows += int(chunk.duplicated().sum())
    except Exception as exc:
        raise FileUnderstandingError(f"CSV 分块读取失败：{exc}", "CSV_PARSE_FAILED") from exc

    table = _build_table_profile(
        sample,
        table_name=Path(filename).stem,
        row_count=row_count,
        missing_counts=missing_counts,
        duplicate_rows=duplicate_rows,
        statistics_scope="sampled",
    )
    quality = list(table.pop("quality_issues"))
    if encoding_issue:
        quality.append(_quality_issue("CSV_ENCODING_FALLBACK", "warning", encoding_issue))
    structure = {
        "profile_schema_version": PROFILE_VERSION,
        "table_count": 1,
        "tables": [table],
        "encoding": encoding,
        "delimiter": delimiter,
    }
    return _base_result(
        filename=filename,
        file_category="table",
        parser_name="pandas_csv",
        structure=structure,
        statistics={"row_count": row_count, "column_count": int(sample.shape[1])},
        quality_issues=quality,
        summary=f"CSV「{filename}」包含 {row_count} 行、{sample.shape[1]} 列，已提取字段结构和质量概况。",
        content_text=" ".join(str(column) for column in sample.columns),
    )


def _parse_xlsx(path: Path, filename: str) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook

        workbook_meta = load_workbook(path, read_only=True, data_only=False, keep_links=False)
        sheet_meta = {
            worksheet.title: {
                "row_count": max(int(worksheet.max_row or 0) - 1, 0),
                "column_count": int(worksheet.max_column or 0),
                "visibility": worksheet.sheet_state,
            }
            for worksheet in workbook_meta.worksheets
        }
        workbook_meta.close()
        excel_file = pd.ExcelFile(path)
    except Exception as exc:
        raise FileUnderstandingError(f"Excel 读取失败：{exc}", "XLSX_PARSE_FAILED") from exc

    tables = []
    quality: list[dict[str, Any]] = []
    total_rows = 0
    try:
        for sheet_name in excel_file.sheet_names:
            metadata = sheet_meta.get(sheet_name, {})
            row_count = int(metadata.get("row_count", 0))
            total_rows += row_count
            try:
                dataframe = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    nrows=max(1, settings.profile_sample_rows),
                )
            except Exception as exc:
                quality.append(
                    _quality_issue(
                        "XLSX_SHEET_PARSE_FAILED",
                        "warning",
                        f"工作表「{sheet_name}」读取失败：{_safe_error(str(exc))}",
                    )
                )
                tables.append(
                    {
                        "sheet_name": sheet_name,
                        "row_count": row_count,
                        "column_count": int(metadata.get("column_count", 0)),
                        "visibility": metadata.get("visibility", "visible"),
                        "columns": [],
                        "parse_status": "failed",
                    }
                )
                continue
            missing_counts = {
                str(column): int(dataframe[column].isna().sum())
                for column in dataframe.columns
            }
            table = _build_table_profile(
                dataframe,
                table_name=sheet_name,
                row_count=row_count,
                missing_counts=missing_counts,
                duplicate_rows=int(dataframe.duplicated().sum()),
                statistics_scope="sampled",
            )
            quality.extend(table.pop("quality_issues"))
            table["sheet_name"] = sheet_name
            table["visibility"] = metadata.get("visibility", "visible")
            table["parse_status"] = "ready"
            tables.append(table)
    finally:
        excel_file.close()

    structure = {
        "profile_schema_version": PROFILE_VERSION,
        "workbook_name": filename,
        "sheet_count": len(tables),
        "sheet_names": [table["sheet_name"] for table in tables],
        "tables": tables,
    }
    columns = [
        column["name"]
        for table in tables
        for column in table.get("columns", [])
    ]
    return _base_result(
        filename=filename,
        file_category="table",
        parser_name="pandas_openpyxl",
        structure=structure,
        statistics={"sheet_count": len(tables), "total_rows": total_rows},
        quality_issues=quality,
        summary=f"Excel「{filename}」包含 {len(tables)} 个工作表、约 {total_rows} 行，已逐表提取结构和质量概况。",
        content_text=" ".join(columns),
    )


def _parse_pdf(
    db: Session,
    file_record: File,
    path: Path,
    *,
    run_ocr: bool = True,
) -> dict[str, Any]:
    pages: list[dict[str, Any]] = []
    text_parts: list[str] = []
    headings: list[str] = []
    metadata: dict[str, Any] = {}
    try:
        with fitz.open(path) as document:
            page_count = len(document)
            if page_count > max(1, settings.pdf_max_pages):
                raise FileUnderstandingError(
                    f"PDF 页数超过安全上限 {settings.pdf_max_pages}",
                    "PDF_PAGE_LIMIT_EXCEEDED",
                )
            metadata = {
                key: value
                for key, value in (document.metadata or {}).items()
                if key in {"title", "author", "subject"} and value
            }
            for index, page in enumerate(document, start=1):
                text = page.get_text("text").strip()
                text_parts.append(text)
                page_lines = [line.strip() for line in text.splitlines() if line.strip()]
                headings.extend(_heading_candidates(page_lines))
                pages.append(
                    {
                        "page_number": index,
                        "text_length": len(text),
                        "has_extractable_text": bool(text),
                    }
                )
    except FileUnderstandingError:
        raise
    except Exception as exc:
        raise FileUnderstandingError(f"PDF 读取失败：{exc}", "PDF_PARSE_FAILED") from exc

    full_text = "\n".join(text_parts).strip()
    scanned_page_numbers = [
        page["page_number"]
        for page in pages
        if page["text_length"] < max(1, settings.pdf_ocr_min_text_chars)
    ]
    scanned_pages = len(scanned_page_numbers)
    scan_ratio = scanned_pages / len(pages) if pages else 0
    quality: list[dict[str, Any]] = []
    ocr_results: list[dict[str, Any]] = []
    if run_ocr and scanned_page_numbers:
        try:
            ocr_results = extract_scanned_pdf_pages(path, scanned_page_numbers)
            _replace_pdf_ocr_chunks(db, file_record.id, ocr_results)
            successful = [item for item in ocr_results if item.get("status") == "success"]
            for item in successful:
                text = str(item.get("text") or "").strip()
                if text:
                    text_parts.append(text)
                    for page in pages:
                        if page["page_number"] == item["page_number"]:
                            page["ocr_status"] = "success"
                            page["ocr_text_length"] = len(text)
                            page["ocr_confidence"] = item.get("confidence")
                            break
            failed_count = len(ocr_results) - len(successful)
            if successful:
                quality.append(
                    _quality_issue(
                        "PDF_OCR_UNVERIFIED",
                        "info",
                        f"已对 {len(successful)} 个缺少文本的页面执行 OCR；重要内容需要人工核对。",
                    )
                )
            if failed_count:
                quality.append(
                    _quality_issue(
                        "PDF_OCR_PARTIAL",
                        "warning",
                        f"{failed_count} 个页面 OCR 未成功，已保留其余页面和原始 Profile 信息。",
                    )
                )
        except FileOcrError as exc:
            quality.append(
                _quality_issue("PDF_OCR_UNAVAILABLE", "warning", _safe_error(exc.message))
            )
    full_text = "\n".join(text_parts).strip()
    if not full_text:
        quality.append(
            _quality_issue(
                "PDF_OCR_REQUIRED",
                "warning",
                "PDF 未提取到文本，疑似扫描件，需要 OCR 后才能进行内容检索。",
            )
        )
    elif scan_ratio >= 0.8:
        quality.append(
            _quality_issue(
                "PDF_MOSTLY_SCANNED",
                "warning",
                "大部分页面文本过少，可能是扫描 PDF。",
            )
        )
    chunk_count = db.scalar(
        select(func.count()).select_from(FileChunk).where(FileChunk.file_id == file_record.id)
    ) or 0
    if full_text and chunk_count == 0:
        try:
            index_result = index_pdf_file(db, file_record)
            chunk_count = int(index_result["chunk_count"])
        except RagServiceError as exc:
            quality.append(
                _quality_issue(
                    "PDF_INDEX_DEGRADED",
                    "warning",
                    f"PDF 分块索引未完成：{_safe_error(exc.message)}",
                )
            )

    title = str(metadata.get("title") or "").strip() or (headings[0] if headings else Path(file_record.filename).stem)
    content_summary = _compact_text(full_text)[:1200]
    summary = (
        f"PDF「{title}」共 {len(pages)} 页，提取文本 {len(full_text)} 个字符。"
        + (f" 内容开头：{content_summary}" if content_summary else " 当前没有可用文本内容。")
    )
    structure = {
        "profile_schema_version": PROFILE_VERSION,
        "page_count": len(pages),
        "title_candidate": title,
        "metadata": metadata,
        "heading_candidates": _deduplicate_strings(headings, 30),
        "page_text_overview": pages,
        "suspected_scanned": scan_ratio >= 0.8,
        "scanned_page_ratio": round(scan_ratio, 4),
        "ocr": {
            "requested": bool(run_ocr and scanned_page_numbers),
            "source_type": "scanned_pdf_ocr",
            "candidate_pages": scanned_page_numbers,
            "successful_pages": [
                item["page_number"]
                for item in ocr_results
                if item.get("status") == "success" and item.get("text")
            ],
            "partial": any(item.get("status") != "success" for item in ocr_results),
        },
        "chunk_count": int(chunk_count),
        "citation_capability": {
            "page_numbers": True,
            "chunk_ids": bool(chunk_count),
        },
    }
    return _base_result(
        filename=file_record.filename,
        file_category="document",
        parser_name="pymupdf_existing_chunks",
        structure=structure,
        statistics={"page_count": len(pages), "text_length": len(full_text)},
        quality_issues=quality,
        summary=summary,
        content_text=full_text,
        title=title,
        fallback_used=not bool(full_text),
    )


def _parse_image(file_record: File, path: Path, run_ocr: bool) -> dict[str, Any]:
    try:
        from PIL import Image

        with Image.open(path) as image:
            width, height = image.size
            image_format = image.format or (file_record.file_type or "").upper()
            color_mode = image.mode
            image.verify()
    except Exception as exc:
        raise FileUnderstandingError(f"图片读取失败：{exc}", "IMAGE_PARSE_FAILED") from exc

    pixels = width * height
    quality: list[dict[str, Any]] = []
    if pixels > max(1, settings.image_max_pixels):
        quality.append(
            _quality_issue(
                "IMAGE_PIXEL_LIMIT",
                "warning",
                f"图片像素数超过建议上限 {settings.image_max_pixels}，OCR 可能较慢。",
            )
        )
    ocr_status = "skipped"
    ocr_text = ""
    ocr_engine = None
    fallback_used = False
    if run_ocr:
        try:
            ocr_result = extract_text_from_image(file_record)
            ocr_status = "success"
            ocr_text = str(ocr_result.get("text") or "")
            ocr_engine = ocr_result.get("engine")
            quality.append(
                _quality_issue(
                    "OCR_UNVERIFIED",
                    "info",
                    "OCR 文本可能存在识别误差，重要内容需要人工核对。",
                )
            )
        except FileOcrError as exc:
            ocr_status = "unavailable"
            fallback_used = True
            quality.append(
                _quality_issue(
                    "OCR_UNAVAILABLE",
                    "warning",
                    _safe_error(exc.message),
                )
            )
    image_kind = _classify_image(width, height, ocr_text)
    language = _detect_language(ocr_text)
    excerpt = _compact_text(ocr_text)[:1000]
    summary = (
        f"图片「{file_record.filename}」尺寸为 {width}×{height}，格式 {image_format}，"
        f"判断为{_image_kind_label(image_kind)}。"
    )
    if excerpt:
        summary += f" OCR 文本摘要：{excerpt}"
    elif ocr_status == "unavailable":
        summary += " OCR 当前不可用，已保留图片基础信息。"
    structure = {
        "profile_schema_version": PROFILE_VERSION,
        "width": width,
        "height": height,
        "pixel_count": pixels,
        "format": image_format,
        "color_mode": color_mode,
        "ocr_status": ocr_status,
        "ocr_engine": ocr_engine,
        "ocr_text_length": len(ocr_text),
        "ocr_text_excerpt": excerpt,
        "language_candidate": language,
        "image_kind": image_kind,
    }
    result = _base_result(
        filename=file_record.filename,
        file_category="image",
        parser_name="pillow_tesseract",
        structure=structure,
        statistics={"width": width, "height": height, "ocr_text_length": len(ocr_text)},
        quality_issues=quality,
        summary=summary,
        content_text=ocr_text,
        fallback_used=fallback_used,
    )
    result["language"] = language
    return result


def _parse_markdown(db: Session, file_record: File, path: Path) -> dict[str, Any]:
    text, encoding, encoding_issue = _read_text_file(path)
    headings: list[dict[str, Any]] = []
    heading_stack: list[str] = []
    for line_number, line in enumerate(text.splitlines(), start=1):
        match = MARKDOWN_HEADING_PATTERN.match(line)
        if not match:
            continue
        level = len(match.group(1))
        title = _clean_text(match.group(2), 300)
        heading_stack = heading_stack[: level - 1] + [title]
        headings.append(
            {
                "level": level,
                "title": title,
                "line_number": line_number,
                "path": " > ".join(heading_stack),
            }
        )
    code_blocks = CODE_FENCE_PATTERN.findall(text)
    links = MARKDOWN_LINK_PATTERN.findall(text)
    external_links = [
        link for link in links if re.match(r"^https?://", link.strip(), flags=re.IGNORECASE)
    ]
    table_count = _count_markdown_tables(text)
    plain_text = _markdown_plain_text(text)
    language = _detect_language(plain_text)
    title = headings[0]["title"] if headings else Path(file_record.filename).stem
    quality: list[dict[str, Any]] = []
    if encoding_issue:
        quality.append(_quality_issue("MARKDOWN_ENCODING_FALLBACK", "warning", encoding_issue))
    if external_links:
        quality.append(
            _quality_issue(
                "MARKDOWN_EXTERNAL_LINKS_NOT_FETCHED",
                "info",
                f"检测到 {len(external_links)} 个外部链接，解析过程中未访问。",
            )
        )
    _replace_markdown_chunks(db, file_record.id, text)
    chunk_count = db.scalar(
        select(func.count()).select_from(FileChunk).where(
            FileChunk.file_id == file_record.id,
            FileChunk.source_type == "markdown",
        )
    ) or 0
    excerpt = _compact_text(plain_text)[:1200]
    summary = (
        f"Markdown「{title}」包含 {len(headings)} 个标题、{len(code_blocks)} 个代码块、"
        f"{table_count} 个表格和 {len(links)} 个链接。"
        + (f" 内容开头：{excerpt}" if excerpt else "")
    )
    structure = {
        "profile_schema_version": PROFILE_VERSION,
        "encoding": encoding,
        "title": title,
        "headings": headings[:100],
        "heading_count": len(headings),
        "code_block_count": len(code_blocks),
        "table_count": table_count,
        "link_count": len(links),
        "external_link_count": len(external_links),
        "text_length": len(plain_text),
        "chunk_count": int(chunk_count),
        "security": {
            "html_executed": False,
            "code_executed": False,
            "external_links_fetched": False,
            "local_references_followed": False,
        },
    }
    result = _base_result(
        filename=file_record.filename,
        file_category="document",
        parser_name="safe_markdown_parser",
        structure=structure,
        statistics={
            "text_length": len(plain_text),
            "heading_count": len(headings),
            "chunk_count": int(chunk_count),
        },
        quality_issues=quality,
        summary=summary,
        content_text=plain_text,
        title=title,
    )
    result["language"] = language
    return result


def _build_table_profile(
    dataframe: pd.DataFrame,
    *,
    table_name: str,
    row_count: int,
    missing_counts: dict[str, int],
    duplicate_rows: int,
    statistics_scope: str,
) -> dict[str, Any]:
    columns: list[dict[str, Any]] = []
    quality: list[dict[str, Any]] = []
    date_ranges: dict[str, dict[str, str | None]] = {}
    numeric_statistics: dict[str, dict[str, Any]] = {}
    primary_key_candidates: list[str] = []
    normalized_names: set[str] = set()

    for raw_column in dataframe.columns:
        column_name = _clean_text(str(raw_column), 200)
        normalized = column_name.casefold()
        if normalized in normalized_names:
            quality.append(
                _quality_issue(
                    "DUPLICATE_COLUMN_NAME",
                    "warning",
                    f"表「{table_name}」存在重复字段名「{column_name}」。",
                )
            )
        normalized_names.add(normalized)
        series = dataframe[raw_column]
        inferred_type = _infer_series_type(series)
        non_null = series.dropna()
        unique_count = int(non_null.nunique(dropna=True))
        missing_count = int(missing_counts.get(str(raw_column), int(series.isna().sum())))
        missing_ratio = missing_count / row_count if row_count else 0
        sample_values = [
            _safe_sample_value(value)
            for value in non_null.drop_duplicates().head(
                max(1, settings.profile_sample_values_per_column)
            )
        ]
        columns.append(
            {
                "name": column_name,
                "inferred_type": inferred_type,
                "sample_values": sample_values,
                "missing_count": missing_count,
                "missing_ratio": round(missing_ratio, 6),
                "sample_unique_count": unique_count,
            }
        )
        if missing_ratio >= 0.5:
            quality.append(
                _quality_issue(
                    "HIGH_MISSING_RATIO",
                    "warning",
                    f"字段「{column_name}」缺失比例为 {missing_ratio:.1%}。",
                )
            )
        if inferred_type == "numeric" and not non_null.empty:
            numeric = pd.to_numeric(non_null, errors="coerce").dropna()
            if not numeric.empty:
                numeric_statistics[column_name] = {
                    "count": int(numeric.count()),
                    "mean": _finite_number(numeric.mean()),
                    "min": _finite_number(numeric.min()),
                    "max": _finite_number(numeric.max()),
                    "median": _finite_number(numeric.median()),
                }
        if inferred_type == "date" and not non_null.empty:
            parsed = pd.to_datetime(non_null, errors="coerce", format="mixed").dropna()
            if not parsed.empty:
                date_ranges[column_name] = {
                    "min": parsed.min().isoformat(),
                    "max": parsed.max().isoformat(),
                }
        if (
            row_count > 0
            and missing_count == 0
            and len(dataframe) > 0
            and unique_count == len(non_null)
            and (
                "id" in normalized
                or "编号" in column_name
                or "代码" in column_name
                or "key" in normalized
            )
        ):
            primary_key_candidates.append(column_name)

    if duplicate_rows:
        quality.append(
            _quality_issue(
                "DUPLICATE_ROWS",
                "warning",
                f"在当前读取范围内发现 {duplicate_rows} 行重复记录。",
            )
        )
    if row_count == 0:
        quality.append(_quality_issue("EMPTY_TABLE", "warning", f"表「{table_name}」为空。"))

    return {
        "table_name": table_name,
        "row_count": row_count,
        "column_count": int(dataframe.shape[1]),
        "columns": columns,
        "duplicate_rows": duplicate_rows,
        "duplicate_rows_scope": statistics_scope,
        "numeric_statistics": numeric_statistics,
        "date_ranges": date_ranges,
        "primary_key_candidates": primary_key_candidates,
        "statistics_scope": statistics_scope,
        "quality_issues": quality,
    }


def _base_result(
    *,
    filename: str,
    file_category: str,
    parser_name: str,
    structure: dict[str, Any],
    statistics: dict[str, Any],
    quality_issues: list[dict[str, Any]],
    summary: str,
    content_text: str,
    title: str | None = None,
    fallback_used: bool = False,
) -> dict[str, Any]:
    suggested_role, confidence = _suggest_role(filename, file_category, content_text)
    tags = _suggest_tags(filename, file_category, structure, content_text)
    return {
        "file_category": file_category,
        "parser_name": parser_name,
        "structure": structure,
        "statistics": statistics,
        "quality_issues": quality_issues,
        "summary": summary,
        "title": title or Path(filename).stem,
        "language": _detect_language(content_text),
        "suggested_role": suggested_role,
        "system_tags": tags,
        "confidence": confidence,
        "content_text": content_text,
        "fallback_used": fallback_used,
    }


def _semantic_enhancement(
    parsed: dict[str, Any],
) -> tuple[SemanticEnhancement | None, int, str | None]:
    safe_structure = {
        key: value
        for key, value in parsed["structure"].items()
        if key not in {"ocr_text_excerpt"}
    }
    model_input = (
        f"确定性摘要：{parsed['summary']}\n"
        f"结构：{safe_json_dumps(safe_structure, max_length=4000)}\n"
        f"文本片段：{parsed['content_text'][:max(0, settings.understanding_model_text_limit)]}"
    )
    started = time.perf_counter()
    result = call_llm(
        messages=[
            {
                "role": "system",
                "content": (
                    "你是文件语义增强器。只能依据输入内容，不能执行代码、命令或链接，"
                    "不能编造文档不存在的信息。只返回 JSON 对象，字段必须为 "
                    "title、summary、suggested_role、tags、confidence。"
                    "suggested_role 只能使用系统给定角色。"
                ),
            },
            {"role": "user", "content": model_input},
        ],
        temperature=0,
        max_tokens=700,
    )
    latency_ms = int((time.perf_counter() - started) * 1000)
    if not result.success or not result.content:
        return None, latency_ms, result.message
    try:
        data = json.loads(_strip_json_fence(result.content))
        return SemanticEnhancement.model_validate(data), latency_ms, None
    except (json.JSONDecodeError, ValidationError, TypeError) as exc:
        return None, latency_ms, f"模型结构化输出校验失败：{exc}"


def _replace_markdown_chunks(db: Session, file_id: int, text: str) -> None:
    db.query(FileChunk).filter(
        FileChunk.file_id == file_id,
        FileChunk.source_type == "markdown",
    ).delete(synchronize_session=False)
    chunks = _markdown_chunks(text)
    db.add_all(
        [
            FileChunk(
                file_id=file_id,
                page_number=None,
                chunk_index=index,
                chunk_text=chunk["text"],
                source_type="markdown",
                section_path=chunk["section_path"],
                char_start=chunk["char_start"],
                char_end=chunk["char_end"],
                chunk_hash=hashlib.sha256(chunk["text"].encode("utf-8")).hexdigest(),
                parser_version=PARSER_VERSION,
            )
            for index, chunk in enumerate(chunks)
        ]
    )
    db.flush()


def _replace_pdf_ocr_chunks(
    db: Session,
    file_id: int,
    page_results: list[dict[str, Any]],
) -> None:
    db.query(FileChunk).filter(
        FileChunk.file_id == file_id,
        FileChunk.source_type == "scanned_pdf_ocr",
    ).delete(synchronize_session=False)
    chunk_size = max(200, settings.rag_chunk_size)
    records: list[FileChunk] = []
    chunk_index = 0
    for result in page_results:
        if result.get("status") != "success":
            continue
        text = str(result.get("text") or "").strip()
        if not text:
            continue
        for offset in range(0, len(text), chunk_size):
            chunk_text = text[offset : offset + chunk_size]
            records.append(
                FileChunk(
                    file_id=file_id,
                    page_number=int(result["page_number"]),
                    chunk_index=chunk_index,
                    chunk_text=chunk_text,
                    source_type="scanned_pdf_ocr",
                    char_start=offset,
                    char_end=offset + len(chunk_text),
                    chunk_hash=hashlib.sha256(chunk_text.encode("utf-8")).hexdigest(),
                    parser_version=f"{PARSER_VERSION}-ocr",
                )
            )
            chunk_index += 1
    if records:
        db.add_all(records)
    db.flush()


def _markdown_chunks(text: str) -> list[dict[str, Any]]:
    chunk_size = max(200, settings.rag_chunk_size)
    heading_stack: list[str] = []
    chunks: list[dict[str, Any]] = []
    position = 0
    buffer: list[str] = []
    buffer_start = 0

    def flush() -> None:
        nonlocal buffer, buffer_start
        content = "\n".join(buffer).strip()
        if not content:
            buffer = []
            return
        offset = 0
        while offset < len(content):
            part = content[offset : offset + chunk_size]
            chunks.append(
                {
                    "text": part,
                    "section_path": " > ".join(heading_stack) or None,
                    "char_start": buffer_start + offset,
                    "char_end": buffer_start + offset + len(part),
                }
            )
            offset += chunk_size
        buffer = []

    for line in text.splitlines(keepends=True):
        heading = MARKDOWN_HEADING_PATTERN.match(line.rstrip("\r\n"))
        if heading:
            flush()
            level = len(heading.group(1))
            heading_stack = heading_stack[: level - 1] + [_clean_text(heading.group(2), 300)]
            buffer_start = position
        elif not buffer:
            buffer_start = position
        buffer.append(line.rstrip("\r\n"))
        if sum(len(item) + 1 for item in buffer) >= chunk_size:
            flush()
            buffer_start = position + len(line)
        position += len(line)
    flush()
    return chunks


def _set_stage(
    db: Session,
    file_record: File,
    profile: FileProfile,
    run: FileProcessingRun,
    stage: str,
) -> None:
    file_record.status = stage
    profile.status = stage
    run.stage = stage
    db.commit()


def _finish_failed_profile(
    db: Session,
    *,
    file_record: File,
    profile: FileProfile,
    run: FileProcessingRun,
    error_code: str,
    error_message: str,
    started: float,
    unsupported: bool,
) -> None:
    safe_message = _safe_error(error_message)
    final_status = "unsupported" if unsupported else "failed"
    profile.status = final_status
    profile.error_code = error_code
    profile.error_message = safe_message
    profile.completed_at = utcnow()
    file_record.status = final_status
    run.status = final_status
    run.error_code = error_code
    run.error_message = safe_message
    run.duration_ms = int((time.perf_counter() - started) * 1000)
    run.completed_at = utcnow()
    db.commit()
    db.refresh(profile)


def _get_workspace_file(db: Session, workspace_id: int, file_id: int) -> WorkspaceFile:
    association = db.scalar(
        select(WorkspaceFile).where(
            WorkspaceFile.workspace_id == workspace_id,
            WorkspaceFile.file_id == file_id,
        )
    )
    if association is None:
        raise FileUnderstandingError("文件不属于当前工作区", "FILE_NOT_IN_WORKSPACE")
    return association


def _detect_csv_format(path: Path) -> tuple[str, str, str | None]:
    raw = path.read_bytes()[:65536]
    issue = None
    encoding = "utf-8-sig"
    decoded = None
    for candidate in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            decoded = raw.decode(candidate)
            encoding = candidate
            if candidate == "gb18030":
                issue = "CSV 使用 GB18030 兼容编码读取。"
            break
        except UnicodeDecodeError:
            continue
    if decoded is None:
        decoded = raw.decode("utf-8", errors="replace")
        encoding = "utf-8"
        issue = "CSV 包含无法解码字符，已使用替换字符读取。"
    try:
        delimiter = csv.Sniffer().sniff(decoded[:8192], delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ","
    return encoding, delimiter, issue


def _read_text_file(path: Path) -> tuple[str, str, str | None]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            issue = "Markdown 使用 GB18030 兼容编码读取。" if encoding == "gb18030" else None
            return raw.decode(encoding), encoding, issue
        except UnicodeDecodeError:
            continue
    return (
        raw.decode("utf-8", errors="replace"),
        "utf-8",
        "Markdown 包含无法解码字符，已使用替换字符读取。",
    )


def _infer_series_type(series: pd.Series) -> str:
    if pd.api.types.is_bool_dtype(series):
        return "boolean"
    if pd.api.types.is_numeric_dtype(series):
        return "numeric"
    if pd.api.types.is_datetime64_any_dtype(series):
        return "date"
    non_empty = series.dropna()
    if not non_empty.empty:
        with pd.option_context("mode.chained_assignment", None):
            parsed = pd.to_datetime(non_empty, errors="coerce", format="mixed")
        if parsed.notna().sum() / len(non_empty) >= 0.8:
            return "date"
    return "text"


def _safe_sample_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, str):
        return _clean_text(value, 80)
    if isinstance(value, (int, float, bool)):
        return value
    return _clean_text(str(value), 80)


def _finite_number(value: Any) -> float | int | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if pd.isna(number) or number in {float("inf"), float("-inf")}:
        return None
    return int(number) if number.is_integer() else round(number, 6)


def _heading_candidates(lines: list[str]) -> list[str]:
    candidates = []
    for line in lines[:80]:
        if len(line) > 100:
            continue
        if re.match(r"^(\d+(?:\.\d+)*)[、.\s]+", line) or re.match(
            r"^(第[一二三四五六七八九十\d]+[章节部分])",
            line,
        ):
            candidates.append(_clean_text(line, 100))
    return candidates[:20]


def _count_markdown_tables(text: str) -> int:
    lines = text.splitlines()
    count = 0
    for index in range(1, len(lines)):
        if "|" not in lines[index - 1]:
            continue
        separator = lines[index].strip()
        if re.match(r"^\|?\s*:?-{3,}", separator) and "|" in separator:
            count += 1
    return count


def _markdown_plain_text(text: str) -> str:
    without_code = CODE_FENCE_PATTERN.sub(" ", text)
    without_html = re.sub(r"<[^>]+>", " ", without_code)
    without_links = MARKDOWN_LINK_PATTERN.sub(" ", without_html)
    without_marks = re.sub(r"(^|\s)[#>*_~`-]+", " ", without_links)
    return _compact_text(without_marks)


def _suggest_role(filename: str, category: str, content: str) -> tuple[str, float]:
    haystack = f"{filename} {content[:2000]}".casefold()
    rules = [
        (("简历", "resume", "curriculum vitae", "cv"), "resume", 0.9),
        (("岗位", "招聘", "job description", " jd "), "job_description", 0.88),
        (("规则", "要求", "评分", "标准", "requirement", "rule"), "rule_document", 0.86),
        (("模板", "template"), "report_template", 0.84),
        (("论文", "研究", "paper", "research"), "research_material", 0.8),
        (("补充", "附录", "supplement", "appendix"), "supplementary_dataset", 0.78),
    ]
    for keywords, role, confidence in rules:
        if any(keyword in haystack for keyword in keywords):
            return role, confidence
    if category == "table":
        return "primary_dataset", 0.72
    if category == "image":
        return "image_evidence", 0.75
    if category == "document":
        return "reference_document", 0.65
    return "unknown", 0.4


def _suggest_tags(
    filename: str,
    category: str,
    structure: dict[str, Any],
    content: str,
) -> list[str]:
    tags = [category]
    suffix = Path(filename).suffix.lower().removeprefix(".")
    if suffix:
        tags.append(suffix)
    if structure.get("sheet_count", 0) > 1:
        tags.append("多工作表")
    if structure.get("suspected_scanned"):
        tags.append("疑似扫描件")
    if structure.get("table_count", 0):
        tags.append("含表格")
    if structure.get("code_block_count", 0):
        tags.append("含代码块")
    text = f"{filename} {content[:1500]}".casefold()
    for keyword, tag in (
        ("求职", "求职"),
        ("岗位", "岗位"),
        ("课程", "课程"),
        ("实验", "实验"),
        ("研究", "研究"),
        ("简历", "简历"),
    ):
        if keyword in text:
            tags.append(tag)
    return _normalize_tags(tags, max_count=12)


def _normalize_confirmed_role(role: str, custom_role: str | None) -> str:
    normalized = role.strip()
    if normalized not in BUILT_IN_ROLES:
        raise FileUnderstandingError("文件角色不在允许范围内", "INVALID_FILE_ROLE")
    if normalized != "custom":
        return normalized
    custom = (custom_role or "").strip()
    if not ROLE_TEXT_PATTERN.fullmatch(custom) or CONTROL_CHAR_PATTERN.search(custom):
        raise FileUnderstandingError(
            "自定义角色只能包含中英文、数字、空格和常用分隔符，长度 1-60",
            "INVALID_CUSTOM_ROLE",
        )
    return f"custom:{custom}"


def _normalize_tags(tags: list[str], max_count: int) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for raw_tag in tags:
        tag = _clean_text(str(raw_tag).strip(), 30)
        if not tag or CONTROL_CHAR_PATTERN.search(tag) or "<" in tag or ">" in tag:
            continue
        key = tag.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(tag)
        if len(result) >= max_count:
            break
    return result


def _classify_image(width: int, height: int, ocr_text: str) -> str:
    compact = _compact_text(ocr_text)
    if compact.count("|") >= 4 or len(re.findall(r"\s{2,}", ocr_text)) >= 5:
        return "table_image"
    if len(compact) >= 150 and height >= width:
        return "scanned_document"
    if len(compact) >= 30 and width >= height:
        return "screenshot"
    return "ordinary_image"


def _image_kind_label(value: str) -> str:
    return {
        "table_image": "表格图",
        "scanned_document": "扫描件",
        "screenshot": "截图",
        "ordinary_image": "普通图片",
    }[value]


def _detect_language(text: str) -> str | None:
    if not text.strip():
        return None
    chinese = len(re.findall(r"[\u4e00-\u9fff]", text))
    english = len(re.findall(r"[A-Za-z]", text))
    if chinese and english:
        return "zh-en" if chinese / max(english, 1) >= 0.2 else "en-zh"
    if chinese:
        return "zh"
    if english:
        return "en"
    return "unknown"


def _quality_issue(code: str, severity: str, message: str) -> dict[str, Any]:
    return {"code": code, "severity": severity, "message": _clean_text(message, 500)}


def _safe_error(message: str) -> str:
    return _clean_text(safe_public_text(message) or "处理失败", 500)


def _clean_text(value: str, max_length: int) -> str:
    return CONTROL_CHAR_PATTERN.sub(" ", value).strip()[:max_length]


def _compact_text(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip()


def _deduplicate_strings(values: list[str], max_count: int) -> list[str]:
    result = []
    seen = set()
    for value in values:
        key = value.casefold()
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
        if len(result) >= max_count:
            break
    return result


def _strip_json_fence(content: str) -> str:
    text = content.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)
    return text


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, default=str)


def _json_dict(value: str | None) -> dict[str, Any]:
    if not value:
        return {}
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _json_list(value: str | None) -> list[str]:
    if not value:
        return []
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return []
    return [str(item) for item in parsed] if isinstance(parsed, list) else []


def _json_object_list(value: str | None) -> list[dict[str, Any]]:
    if not value:
        return []
    try:
        parsed = json.loads(value)
    except json.JSONDecodeError:
        return []
    return [item for item in parsed if isinstance(item, dict)] if isinstance(parsed, list) else []
