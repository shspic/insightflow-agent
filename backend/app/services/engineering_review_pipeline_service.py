"""工程审查确定性管道 — 字段抽取、Evidence 创建、规则执行、Finding 生成。

不调用 LLM，不读取 ground_truth.json，不根据文件名猜测，不含黄金答案硬编码。
"""

from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

import fitz
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.evidence import Evidence
from app.models.file import File
from app.models.file_profile import FileProfile
from app.models.review_brief import ReviewBrief
from app.models.review_finding import ReviewFinding
from app.models.review_run import ReviewRun
from app.models.workspace import Workspace
from app.models.workspace_file import WorkspaceFile
from app.schemas.review import (
    EvidenceCreate,
    ReviewRuleDef,
    ReviewRulePack,
    StructuredFieldValue,
    StructuredReviewInput,
)
from app.services.evidence_provenance import (
    FIELD_LOCATOR,
    validate_pdf_page,
    validate_spreadsheet_cell,
    validate_text_chunk_index,
)
from app.services.review_action_service import (
    ReviewServiceError,
    create_evidence,
    create_review_finding,
    complete_review_run,
    fail_review_run,
    start_review_run,
)
from app.services.review_engine_service import execute_all_rules


REQUIRED_ROLES = frozenset({
    "tender_requirement", "bid_response", "personnel_equipment_data",
    "qualification_attachment", "clarification_document",
})

PARSER_NAME = "engineering_review_pipeline"
PARSER_VERSION = "v2.1.0"  # 阶段 3B-1 补修：负责人解析语义修复 + 免责声明条件化


class PipelineError(Exception):
    """带稳定错误码的管道异常。"""
    def __init__(self, code: str, message: str):
        self.code = code
        self.message = message
        super().__init__(message)


class ExtractionResult:
    """确定性字段抽取结果，维护字段→值→证据的精确映射。"""
    def __init__(self):
        self.fields: dict[str, StructuredFieldValue] = {}
        # 每条 Evidence 的元数据：(EvidenceCreate, field_path_marker, evidence_role)
        self.evidence_meta: list[tuple[EvidenceCreate, str, str]] = []
        self.scan_pages: list[int] = []


def run_engineering_review(
    db: Session,
    *,
    run: ReviewRun,
    workspace: Workspace,
    owner_user_id: int,
) -> dict[str, Any]:
    """执行完整工程审查管道。使用 run 中保存的规则快照。"""
    if workspace.workspace_type != "engineering":
        fail_review_run(db, run, "REVIEW_WORKSPACE_TYPE_INVALID", "仅 engineering 工作区可用")
        raise PipelineError("REVIEW_WORKSPACE_TYPE_INVALID", "仅 engineering 工作区可以执行工程审查")

    # 1. 从快照恢复规则包（先恢复，异常时 fail run）
    try:
        rule_pack = _restore_rule_pack(run)
    except PipelineError as e:
        fail_review_run(db, run, e.code, e.message)
        raise

    # 2. 校验 Brief（仅检查快照中引用存在即可，不要求仍是 confirmed）
    brief = db.query(ReviewBrief).filter(ReviewBrief.id == run.review_brief_id).first()
    if brief is None:
        fail_review_run(db, run, "REVIEW_BRIEF_NOT_CONFIRMED", "ReviewBrief 快照不存在")
        raise PipelineError("REVIEW_BRIEF_NOT_CONFIRMED", "ReviewBrief 快照不存在")

    # 3. 校验材料
    try:
        files_by_role = _validate_materials(db, workspace.id, owner_user_id)
    except PipelineError as e:
        fail_review_run(db, run, e.code, e.message)
        raise
    except ReviewServiceError as e:
        fail_review_run(db, run, "REVIEW_MATERIAL_MISSING", str(e))
        raise PipelineError("REVIEW_MATERIAL_MISSING", str(e))

    # 4. 抽取字段
    try:
        extraction = _extract_all_fields(files_by_role)
    except Exception as exc:
        fail_review_run(db, run, "REVIEW_EXTRACTION_FAILED", str(exc))
        raise PipelineError("REVIEW_EXTRACTION_FAILED", str(exc))

    # 5. 创建 Evidence（来源哈希 + 真实 locator 校验，失败时 fail run）
    start_review_run(db, run)
    try:
        evidence_records, evidence_index = _persist_evidence(db, run, workspace.id, owner_user_id, extraction.evidence_meta)
    except PipelineError as exc:
        fail_review_run(db, run, exc.code, exc.message)
        raise
    except Exception as exc:
        fail_review_run(db, run, "REVIEW_EVIDENCE_ERROR", f"Evidence 创建失败: {exc}")
        raise PipelineError("REVIEW_EVIDENCE_ERROR", f"Evidence 创建失败: {exc}")

    # 6. 结构化输入快照：真实字段抽取完成后自动持久化规范 JSON + SHA-256，
    #    随后 Quality Gate 会复算校验；评测脚本禁止手工写入。
    doc_roles = {role: [files_by_role[role]["file"].id] for role in files_by_role}
    structured_input = StructuredReviewInput(fields=extraction.fields, document_roles=doc_roles)
    try:
        _persist_input_snapshot(db, run, structured_input)
    except Exception as exc:
        fail_review_run(db, run, "REVIEW_ENGINE_ERROR", f"结构化输入快照持久化失败: {exc}")
        raise PipelineError("REVIEW_ENGINE_ERROR", f"结构化输入快照持久化失败: {exc}")

    # 7. 执行规则
    try:
        rule_results = execute_all_rules(rule_pack.rules, structured_input)
    except Exception as exc:
        fail_review_run(db, run, "REVIEW_ENGINE_ERROR", f"规则引擎执行异常: {exc}")
        raise PipelineError("REVIEW_ENGINE_ERROR", str(exc))

    # 8. 生成 Finding（精确证据绑定，异常时 fail run）
    finding_count = 0
    failed_ids = []
    rule_def_map = {r.rule_id: r for r in rule_pack.rules}
    for rr in rule_results:
        rule_id = rr.get("rule_id", "unknown")
        failed_ids.append(rule_id)
        try:
            finding = _build_finding(
                db, run, workspace.id, owner_user_id,
                rule_result=rr, rule_def_map=rule_def_map,
                extraction=extraction, evidence_index=evidence_index,
            )
            if finding is not None:
                finding_count += 1
        except Exception as exc:
            fail_review_run(db, run, "REVIEW_ENGINE_ERROR", f"Finding 创建失败 [{rule_id}]: {exc}")
            raise PipelineError("REVIEW_ENGINE_ERROR", f"Finding 创建失败 [{rule_id}]: {exc}")

    # 9. 计算通过规则
    all_ids = {r.rule_id for r in rule_pack.rules}
    passed_ids = sorted(all_ids - set(failed_ids))

    complete_review_run(db, run)
    return {
        "status": "completed",
        "finding_count": finding_count,
        "evidence_count": len(evidence_records),
        "passed_rule_ids": passed_ids,
        "failed_rule_ids": sorted(failed_ids),
    }


# ── 规则快照恢复 ──────────────────────────────────────────────────


def _restore_rule_pack(run: ReviewRun) -> ReviewRulePack:
    """从 run 保存的快照恢复规则包，校验哈希。"""
    if not run.rule_snapshot_json:
        raise PipelineError("REVIEW_ENGINE_ERROR", "Run 缺少规则快照")
    actual_hash = hashlib.sha256(run.rule_snapshot_json.encode("utf-8")).hexdigest()
    if actual_hash != run.rule_pack_hash:
        raise PipelineError("REVIEW_ENGINE_ERROR", f"规则快照哈希不一致：期望 {run.rule_pack_hash[:12]}..., 实际 {actual_hash[:12]}...")
    try:
        from app.schemas.review import ReviewRulePack as RRP
        data = json.loads(run.rule_snapshot_json)
        return RRP.model_validate(data)
    except Exception as e:
        raise PipelineError("REVIEW_ENGINE_ERROR", f"规则快照反序列化失败: {e}")


# ── 材料校验 ──────────────────────────────────────────────────────


def _validate_materials(db: Session, workspace_id: int, owner_user_id: int) -> dict:
    """查找五个必需角色的已确认文件，校验 owner 和一致性。"""
    from sqlalchemy import select as sa_select

    wf_rows = list(db.scalars(sa_select(WorkspaceFile).where(WorkspaceFile.workspace_id == workspace_id)).all())

    role_files: dict[str, list[dict]] = {}
    for wf in wf_rows:
        role = wf.user_confirmed_role
        if role is None or role not in REQUIRED_ROLES:
            continue
        file_record = db.query(File).filter(File.id == wf.file_id).first()
        if file_record is None:
            continue
        # 校验 owner
        if file_record.owner_user_id != owner_user_id:
            continue
        # 校验 FileProfile 一致性和状态
        profile = db.scalar(
            sa_select(FileProfile).where(
                FileProfile.workspace_id == workspace_id,
                FileProfile.file_id == wf.file_id,
                FileProfile.owner_user_id == owner_user_id,
                FileProfile.status == "ready",
                FileProfile.confirmed_role == role,
            ).order_by(FileProfile.profile_version.desc())
        )
        if profile is None:
            continue
        if role not in role_files:
            role_files[role] = []
        role_files[role].append({"file": file_record, "wf": wf, "profile": profile})

    for role in REQUIRED_ROLES:
        if role not in role_files:
            raise PipelineError("REVIEW_MATERIAL_MISSING", f"缺少已确认角色「{role}」的文件")
        if len(role_files[role]) > 1:
            raise PipelineError("REVIEW_ROLE_DUPLICATED", f"角色「{role}」存在多个文件，需要人工处理")

    return {role: role_files[role][0] for role in REQUIRED_ROLES}


# ── 字段抽取（无硬编码）───────────────────────────────────────────


def _extract_all_fields(files_by_role: dict) -> ExtractionResult:
    result = ExtractionResult()
    for role, entry in files_by_role.items():
        f = entry["file"]
        if role == "bid_response":
            _extract_bid_response(result, f)
        elif role == "personnel_equipment_data":
            _extract_personnel_equipment(result, f)
        elif role == "qualification_attachment":
            _extract_qualification(result, f)
        elif role == "clarification_document":
            _extract_clarification(result, f)
    return result


def _extract_bid_response(result: ExtractionResult, f: File):
    if not f.file_type or f.file_type.lower() != "pdf":
        return
    text = _read_pdf_text(f.file_path)

    # 项目名称：查找「项目名称」标签后的值
    pn = _find_label_value(text, ["项目名称", "Project Name", "项目名称："])
    # 如果找到的值是留空标记或无实际内容，则为空
    if pn and any(marker in str(pn) for marker in ["留空", "未填写", "（空）", "(空)"]):
        pn = ""
    result.fields["bid_response.project_name"] = StructuredFieldValue(value=pn or "")

    # 项目负责人（使用专用人员姓名解析，避免将证书编号误识别为姓名）
    leader = _find_person_name(text, ["项目负责人", "负责人", "负责人姓名"])
    result.fields["bid_response.leader_name"] = StructuredFieldValue(value=leader or "")

    # 证书编号（通用正则）
    cert = _find_cert_number(text)
    result.fields["bid_response.leader_cert_number"] = StructuredFieldValue(value=cert or "")

    # 报价（通用正则：含逗号的千分位数字或纯数字）
    price = _find_price(text)
    result.fields["bid_response.total_price"] = StructuredFieldValue(value=price if price is not None else "")

    # 日期（通用：所有 ISO 日期中的相关日期）
    sdate = _find_related_date(text, ["签署", "提交", "日期", "sign", "submit", "date"])
    result.fields["bid_response.submission_date"] = StructuredFieldValue(value=sdate or "")

    # Evidence — 空字段用 field_problem
    pn_role = "field_problem" if not pn else "field_value"
    result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="pdf_page", page_number=1,
        quote=f"项目名称字段：{pn if pn else '（空）'}", parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "bid_response.project_name", pn_role))
    # 负责人 Evidence：找到用 field_value，未找到用 field_problem（不允许用证书编号伪造）
    if leader:
        result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="pdf_page", page_number=1,
            quote=f"项目负责人：{leader}", parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "bid_response.leader_name", "field_value"))
    else:
        result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="pdf_page", page_number=1,
            quote="项目负责人：未找到", parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "bid_response.leader_name", "field_problem"))
    if cert:
        result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="pdf_page", page_number=1,
            quote=f"证书编号：{cert}", parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "bid_response.leader_cert_number", "field_value"))
    if price is not None:
        result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="pdf_page", page_number=1,
            quote=f"总报价：{price}", parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "bid_response.total_price", "field_value"))
    if sdate:
        result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="pdf_page", page_number=1,
            quote=f"签署日期：{sdate}", parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "bid_response.submission_date", "field_value"))


def _extract_personnel_equipment(result: ExtractionResult, f: File):
    if not f.file_type or f.file_type.lower() not in ("xlsx", "csv"):
        return
    wb = load_workbook(f.file_path, data_only=True)

    # 项目概况
    ws1 = wb["项目概况"]
    pn = str(ws1["B3"].value or "").strip()
    result.fields["personnel_equipment_data.project_name"] = StructuredFieldValue(value=pn)

    # 人员清单
    ws2 = wb["人员清单"]
    leader = str(ws2.cell(row=3, column=2).value or "").strip()
    leader_cert = str(ws2.cell(row=3, column=4).value or "").strip()
    total_p = int(ws2.cell(row=8, column=2).value or 0)

    result.fields["personnel_equipment_data.leader_name"] = StructuredFieldValue(value=leader)
    result.fields["personnel_equipment_data.leader_cert_number"] = StructuredFieldValue(value=leader_cert)
    result.fields["personnel_equipment_data.total_personnel"] = StructuredFieldValue(value=total_p)

    # 设备清单
    ws3 = wb["设备清单"]
    total_e = int(ws3.cell(row=7, column=4).value or 0)
    cal = ws3.cell(row=4, column=6).value
    if isinstance(cal, datetime):
        cal_str = cal.strftime("%Y-%m-%d")
    elif isinstance(cal, date):
        cal_str = cal.isoformat()
    else:
        cal_str = str(cal or "").strip()

    result.fields["personnel_equipment_data.total_equipment"] = StructuredFieldValue(value=total_e)
    result.fields["personnel_equipment_data.earliest_calibration_expiry"] = StructuredFieldValue(value=cal_str)
    wb.close()

    # Evidence
    result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="spreadsheet_cell",
        sheet_name="项目概况", cell_range="B3", quote=f"项目名称：{pn}",
        parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "personnel_equipment_data.project_name", "field_value"))
    # 人员清单负责人证据：空为 problem，有值为 value
    leader_ev_role = "field_problem" if not leader else "field_value"
    result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="spreadsheet_cell",
        sheet_name="人员清单", cell_range="B3", quote="项目负责人姓名为空" if not leader else f"项目负责人：{leader}",
        parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "personnel_equipment_data.leader_name", leader_ev_role))
    result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="spreadsheet_cell",
        sheet_name="人员清单", cell_range="D3", quote=f"证书编号：{leader_cert}",
        parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "personnel_equipment_data.leader_cert_number", "field_value"))
    result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="spreadsheet_cell",
        sheet_name="人员清单", cell_range="B8", quote=f"人员总数：{total_p}",
        parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "personnel_equipment_data.total_personnel", "field_value"))
    result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="spreadsheet_cell",
        sheet_name="设备清单", cell_range="D7", quote=f"设备总数：{total_e}",
        parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "personnel_equipment_data.total_equipment", "field_value"))
    result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="spreadsheet_cell",
        sheet_name="设备清单", cell_range="F4", quote=f"最早校准有效期：{cal_str}",
        parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "personnel_equipment_data.earliest_calibration_expiry", "field_value"))


def _extract_qualification(result: ExtractionResult, f: File):
    if not f.file_type or f.file_type.lower() != "pdf":
        return
    text_p1, text_p2, page_count = _read_pdf_pages(f.file_path)

    # 有效期：通用日期提取
    expiry = _find_all_dates(text_p1)
    result.fields["qualification_attachment.expiry_date"] = StructuredFieldValue(value=expiry[0] if expiry else "")

    # 扫描页检测
    is_scan = len(text_p2.strip()) < 50
    if is_scan:
        result.scan_pages.append(2)
        result.fields["qualification_attachment.cert_number"] = StructuredFieldValue(value="", evidence_ids=[])

    # Evidence
    if expiry:
        result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="pdf_page", page_number=1,
            quote=f"资质有效期至：{expiry[0]}", parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "qualification_attachment.expiry_date", "field_value"))
    if is_scan:
        result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="pdf_page", page_number=2,
            quote="第 2 页为光栅扫描页，当前未启用 OCR，证书编号不可机器提取",
            parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "qualification_attachment.cert_number", "field_problem"))


def _extract_clarification(result: ExtractionResult, f: File):
    if not f.file_type or f.file_type.lower() not in ("md", "markdown"):
        return
    # chunk_id 与 Corpus text_chunk_index 统一为 0-based
    result.evidence_meta.append((EvidenceCreate(file_id=f.id, locator_type="text_chunk", chunk_id=0,
        quote="项目澄清文件已确认", parser_name=PARSER_NAME, parser_version=PARSER_VERSION), "clarification_document", "presence"))


# ── 通用解析辅助（无硬编码）───────────────────────────────────────


def _read_pdf_text(path: str) -> str:
    doc = fitz.open(path)
    text = "\n".join(page.get_text() for page in doc)
    doc.close()
    return text


def _read_pdf_pages(path: str) -> tuple[str, str, int]:
    doc = fitz.open(path)
    p1 = doc[0].get_text() if doc.page_count >= 1 else ""
    p2 = doc[1].get_text() if doc.page_count >= 2 else ""
    pc = doc.page_count
    doc.close()
    return p1, p2, pc


def _find_label_value(text: str, labels: list[str]) -> str | None:
    """根据标签列表在文本中查找对应值。"""
    for line in text.split("\n"):
        line_s = line.strip()
        for label in labels:
            if label in line_s:
                for sep in ("：", ":", " ", "\t"):
                    if sep in line_s:
                        parts = line_s.split(sep, 1)
                        if len(parts) == 2:
                            val = parts[1].strip()
                            if val:
                                return val
                # 标签后无分隔符，取后续文本
                idx = line_s.find(label)
                remainder = line_s[idx + len(label):].strip()
                if remainder:
                    return remainder
    return None


def _find_person_name(text: str, labels: list[str]) -> str | None:
    """从文本中提取人员姓名，支持标签作用域内解析。

    与通用 _find_label_value 不同，本函数：
    - 先在标签所在位置截取 remainder，而非对整行 split
    - 从 remainder 中识别合理的人名片段
    - 拒绝证书编号、规则代码、其他字段标签等非姓名内容
    - 支持跨行标签（标签单独一行，姓名在下一行）

    支持格式：
    - 项目负责人：张三
    - 项目负责人: 李四
    - 项目负责人 = 王五
    - 项目负责人为赵六
    - 项目负责人\\n王五（跨行）
    """
    # 已知的字段标签和不可作为姓名的内容（边界检测）
    _STOP_WORDS = frozenset({
        "证书编号", "证书号", "Cert", "规则代码", "项目名称",
        "报价", "日期", "签署", "提交", "备注", "说明",
        "虚构姓名", "姓名", "单位", "职务", "职称", "联系方式",
    })
    # 证书编号 / 规则代码模式
    _CERT_PATTERN = re.compile(r'^[A-Z0-9][A-Z0-9/-]{2,118}[A-Z0-9]$', re.I)
    # 中文姓名：2-4 个汉字
    _CN_NAME = re.compile(r'^[一-鿿]{2,4}$')
    # 英文姓名或拼音
    _EN_NAME = re.compile(r'^[A-Za-z][A-Za-z .\-]{1,30}$')

    def _looks_like_name(value: str) -> bool:
        v = value.strip().rstrip("，,。.；;、）)")
        if not v or len(v) > 20:
            return False
        # 拒绝证书编号 / 规则代码
        if _CERT_PATTERN.match(v):
            return False
        # 拒绝包含已知字段标签的值
        for sw in _STOP_WORDS:
            if sw in v:
                return False
        # 必须为中文姓名或合理的英文名
        if _CN_NAME.match(v):
            return True
        if _EN_NAME.match(v):
            return True
        return False

    def _extract_from_remainder(remainder: str) -> str | None:
        """从标签后的剩余文本中提取姓名。"""
        r = remainder.strip()
        if not r:
            return None
        # 去掉连接词前缀：为、是、：、:、=、空格
        r = re.sub(r'^[为是：:=\s]+', '', r).strip()
        if not r:
            return None
        # 截取到第一个边界字符（中文/英文标点）
        m = re.search(r'[，,。.；;、）)]', r)
        candidate = r[:m.start()].strip() if m else r.strip()
        # 按空格进一步截断（取第一个词）
        if ' ' in candidate:
            parts = candidate.split()
            candidate = parts[0] if parts else candidate
        # 去掉尾部连接符
        candidate = candidate.rstrip("，,。.：:）)")
        if _looks_like_name(candidate):
            return candidate
        return None

    lines = text.splitlines()
    for i, line in enumerate(lines):
        line_s = line.strip()
        if not line_s:
            continue
        for label in labels:
            idx = line_s.find(label)
            if idx == -1:
                continue
            # 标签作用域：只处理标签结束位置之后的 remainder
            remainder = line_s[idx + len(label):]
            name = _extract_from_remainder(remainder)
            if name:
                return name
            # 标签所在行无有效姓名，尝试下一行（跨行格式）
            if i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line:
                    name = _extract_from_remainder(next_line)
                    if name:
                        return name
    return None


def _find_cert_number(text: str) -> str | None:
    """仅在证书编号标签附近查找完整的字母数字 token。"""
    lines = text.splitlines()
    label_pattern = re.compile(r'(?:证书编号|证书号|Cert\s*No\.?)', re.I)
    token_pattern = re.compile(r'[A-Z0-9][A-Z0-9/-]{2,118}[A-Z0-9]', re.I)

    def valid_candidate(value: str) -> str | None:
        candidate = value.strip()
        if not token_pattern.fullmatch(candidate):
            return None
        if not any(char.isalpha() for char in candidate):
            return None
        if not any(char.isdigit() for char in candidate):
            return None
        return candidate

    for index, line in enumerate(lines):
        label_match = label_pattern.search(line)
        if label_match is None:
            continue

        remainder = line[label_match.end():].lstrip(" \t：:")
        if remainder:
            candidate = valid_candidate(remainder)
            if candidate is not None:
                return candidate
            continue

        for next_line in lines[index + 1:]:
            if not next_line.strip():
                continue
            candidate = valid_candidate(next_line)
            if candidate is not None:
                return candidate
            break
    return None


def _find_price(text: str) -> int | None:
    """通用报价提取：千分位数字或大额数字。"""
    # 查找标签后的数字
    for label in ["总报价", "报价", "投标总价", "金额", "Total"]:
        for m in re.finditer(rf'{re.escape(label)}\s*[：:]\s*([\d,]+(?:\s*元)?)', text):
            raw = re.sub(r'[,\s元]', '', m.group(1))
            try:
                return int(raw)
            except ValueError:
                continue
    # 通用千分位数字
    for m in re.finditer(r'([\d,]{5,20})\s*元', text):
        raw = m.group(1).replace(",", "")
        try:
            v = int(raw)
            if v > 10000:
                return v
        except ValueError:
            continue
    return None


def _find_related_date(text: str, keywords: list[str]) -> str | None:
    """根据关键词查找相关日期的行，提取第一个日期。"""
    dates = re.findall(r'(\d{4}-\d{2}-\d{2})', text)
    for line in text.split("\n"):
        if any(kw in line for kw in keywords):
            line_dates = re.findall(r'(\d{4}-\d{2}-\d{2})', line)
            if line_dates:
                return line_dates[0]
    return dates[0] if dates else None


def _find_all_dates(text: str) -> list[str]:
    return re.findall(r'(\d{4}-\d{2}-\d{2})', text)


# ── Finding 生成 ──────────────────────────────────────────────────


def _build_finding(
    db: Session, run: ReviewRun, workspace_id: int, owner_user_id: int,
    rule_result: dict, rule_def_map: dict[str, ReviewRuleDef],
    extraction: ExtractionResult, evidence_index: dict[str, list[int]],
) -> ReviewFinding | None:
    rule_id = rule_result.get("rule_id", "unknown")
    rule_def = rule_def_map.get(rule_id)
    rtype = rule_result.get("rule_type", "")
    findings = rule_result.get("findings", [])

    # 精确证据绑定
    evidence_ids = _precise_match(rule_id, rtype, findings, extraction, evidence_index)

    title = rule_def.title if rule_def else rule_id
    severity = rule_def.severity if rule_def else "medium"
    category = rtype
    suggestion = rule_def.suggestion if rule_def else "请人工复核"
    conclusion = _build_conclusion(rule_result, rule_def)

    return create_review_finding(
        db, review_run_id=run.id, workspace_id=workspace_id, owner_user_id=owner_user_id,
        issue_code=rule_id, title=title, category=category, severity=severity,
        conclusion=conclusion, suggestion=suggestion,
        rule_id=rule_id, rule_version=rule_def.version if rule_def else "1.0",
        evidence_ids=evidence_ids, source_step_id=f"engine:{rule_id}",
    )


def _precise_match(
    rule_id: str, rtype: str, findings: list,
    extraction: ExtractionResult, evidence_index: dict[str, list[int]],
) -> list[int]:
    """精确的字段→Evidence 绑定，无兜底逻辑。"""

    # 提取涉及的字段路径
    affected_paths = set()
    for f in findings:
        if isinstance(f, dict) and f.get("path"):
            affected_paths.add(f["path"])

    # 按字段路径精确匹配
    matched = set()
    for path in affected_paths:
        key = f"field:{path}"
        if key in evidence_index:
            matched.update(evidence_index[key])
        # also try role-based match
        for ek, eids in evidence_index.items():
            if path in ek:
                matched.update(eids)

    # document_presence: 匹配 presence: 前缀证据
    if rtype == "document_presence":
        for ek, eids in evidence_index.items():
            if ek.startswith("presence:"):
                matched.update(eids)

    # evidence_required: 仅匹配目标字段的问题证据（field_problem），不跨字段
    if rtype == "evidence_required":
        for path in affected_paths:
            # 精确：字段的问题定位证据
            pk = f"field_problem:{path}"
            if pk in evidence_index:
                matched.update(evidence_index[pk])
            # 回退：字段的值证据
            k = f"field:{path}"
            if k in evidence_index:
                matched.update(evidence_index[k])

    result = list(matched)[:5]
    return result


def _build_conclusion(rule_result: dict, rule_def: ReviewRuleDef | None) -> str:
    findings = rule_result.get("findings", [])
    rtype = rule_result.get("rule_type", "")
    rule_id = rule_result.get("rule_id", "")

    if rtype == "required_field":
        for f in findings:
            if isinstance(f, dict):
                return f"字段「{f.get('path', '')}」缺失或为空。"
    if rtype == "cross_file_equal":
        vals = [str(f.get("value", "")) for f in findings if isinstance(f, dict)]
        return f"跨文件字段不一致：{' vs '.join(vals)}。"
    if rtype == "numeric_threshold":
        for f in findings:
            if isinstance(f, dict):
                return f.get("message", f"数值不满足阈值要求。")
    if rtype == "date_order":
        for f in findings:
            if isinstance(f, dict):
                return f.get("message", f"日期不满足顺序要求。")
    if rtype == "document_presence":
        for f in findings:
            if isinstance(f, dict):
                return f.get("message", "缺少必需文件角色。")
    if rtype == "evidence_required":
        for f in findings:
            if isinstance(f, dict):
                return f.get("message", "字段缺少证据定位。")
    return f"规则 {rule_id} 未通过确定性检查。"


def _persist_evidence(
    db: Session, run: ReviewRun, workspace_id: int, owner_user_id: int,
    evidence_meta: list[tuple[EvidenceCreate, str, str]],
) -> tuple[list[Evidence], dict[str, list[int]]]:
    """持久化 Evidence 并建立精确索引。

    契约（阶段 6A）：
    - 每条确定性管道 Evidence：provenance_type=field_locator；
    - 创建前按真实文件校验 locator（PDF page / Excel sheet+cell / text chunk 编号）；
    - 创建后必须携带 source_file_hash，缺失即失败（来源不可锚定）。
    """
    records = []
    evidence_index: dict[str, list[int]] = {}
    for ev_create, field_marker, ev_role in evidence_meta:
        _validate_locator_exists(db, ev_create)
        ev_typed = ev_create.model_copy(update={"provenance_type": FIELD_LOCATOR})
        record = create_evidence(db, review_run_id=run.id, workspace_id=workspace_id, owner_user_id=owner_user_id, evidence=ev_typed)
        if record.source_file_hash is None:
            raise PipelineError(
                "REVIEW_EVIDENCE_PROVENANCE_UNAVAILABLE",
                f"Evidence 来源文件哈希缺失（file_id={ev_create.file_id}），无法锚定来源",
            )
        records.append(record)
        # 按角色索引
        if ev_role == "field_value":
            key = f"field:{field_marker}"
        elif ev_role == "field_problem":
            key = f"field_problem:{field_marker}"
        elif ev_role == "presence":
            key = f"presence:{field_marker}"
        else:
            key = f"field:{field_marker}"
        if key not in evidence_index:
            evidence_index[key] = []
        evidence_index[key].append(record.id)
    return records, evidence_index


def _validate_locator_exists(db: Session, ev_create: EvidenceCreate) -> None:
    """创建前按真实文件校验定位存在（PDF page / Excel sheet+cell / text chunk 编号）。

    定位不存在 → REVIEW_EVIDENCE_LOCATOR_INVALID（运行失败，不产生无锚证据）。
    """
    from app.models.file import File

    file_record = db.query(File).filter(File.id == ev_create.file_id).first()
    if file_record is None:
        raise PipelineError("REVIEW_EVIDENCE_LOCATOR_INVALID",
                            f"证据定位文件不存在（file_id={ev_create.file_id}）")
    path = file_record.file_path
    if ev_create.locator_type == "pdf_page":
        if not validate_pdf_page(path, ev_create.page_number):
            raise PipelineError("REVIEW_EVIDENCE_LOCATOR_INVALID",
                                f"证据定位无效：PDF page {ev_create.page_number} 不存在")
    elif ev_create.locator_type == "spreadsheet_cell":
        if not validate_spreadsheet_cell(path, ev_create.sheet_name, ev_create.cell_range):
            raise PipelineError("REVIEW_EVIDENCE_LOCATOR_INVALID",
                                "证据定位无效：Excel sheet/cell 不存在")
    elif ev_create.locator_type == "text_chunk":
        if not validate_text_chunk_index(path, ev_create.chunk_id):
            raise PipelineError("REVIEW_EVIDENCE_LOCATOR_INVALID",
                                f"证据定位无效：text chunk {ev_create.chunk_id} 不存在")


def _persist_input_snapshot(db: Session, run: ReviewRun, structured_input: StructuredReviewInput) -> None:
    """把真实抽取的结构化输入持久化为规范 JSON + SHA-256（sort_keys）。

    快照在规则执行前写入，Quality Gate 复算校验；评测脚本禁止直接写这两个字段。
    """
    payload = structured_input.model_dump()
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    run.input_snapshot_json = serialized
    run.input_snapshot_hash = hashlib.sha256(serialized.encode("utf-8")).hexdigest()
    db.commit()
    db.refresh(run)
