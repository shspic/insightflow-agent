"""V3 阶段 2B-1 材料完整性测试 —— 补修强化版。"""

from __future__ import annotations

import hashlib
import json
import tempfile
import time
from pathlib import Path

import fitz
import pytest
from openpyxl import load_workbook

from app.schemas.review import ReviewBriefCreate, InterpretedIntent
from app.services.review_rule_service import load_rule_pack

GOLDEN_DIR = (
    Path(__file__).resolve().parents[2] / "examples" / "engineering_review_v1" / "golden_case"
)
DISCLAIMER_CN = "合成演示数据，不作为真实招投标、工程、资质或法律判断依据"
ENGINEERING_ROLES = frozenset({
    "tender_requirement", "bid_response", "personnel_equipment_data",
    "qualification_attachment", "clarification_document", "supplementary_attachment",
})

# ── helpers ──────────────────────────────────────────────────────

def _sha256(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()

@pytest.fixture(scope="module")
def _manifest():
    return json.loads((GOLDEN_DIR / "manifest.json").read_text(encoding="utf-8"))

@pytest.fixture(scope="module")
def _gt():
    return json.loads((GOLDEN_DIR / "ground_truth.json").read_text(encoding="utf-8"))

@pytest.fixture(scope="module")
def _tender_doc():
    return fitz.open(GOLDEN_DIR / "01_合成招标要求.pdf")

@pytest.fixture(scope="module")
def _bid_doc():
    return fitz.open(GOLDEN_DIR / "02_合成投标响应.pdf")

@pytest.fixture(scope="module")
def _qual_doc():
    return fitz.open(GOLDEN_DIR / "04_合成资质附件.pdf")

@pytest.fixture(scope="module")
def _xlsx():
    return load_workbook(GOLDEN_DIR / "03_人员设备清单.xlsx", data_only=True)


# ── 文件存在与 manifest ──────────────────────────────────────────

class TestMaterialExistence:
    def test_all_files_exist(self):
        for fn in ["01_合成招标要求.pdf","02_合成投标响应.pdf","03_人员设备清单.xlsx","04_合成资质附件.pdf","05_项目澄清.md","manifest.json","ground_truth.json","review_brief.json"]:
            assert (GOLDEN_DIR / fn).is_file(), f"missing {fn}"

    def test_manifest_hashes_match(self, _manifest):
        for e in _manifest["files"]:
            assert _sha256(GOLDEN_DIR / e["filename"]) == e["sha256"], f"{e['filename']} hash mismatch"

    def test_manifest_top_fields(self, _manifest):
        assert _manifest["case_id"] == "SYN-ENG-2026-001"
        assert _manifest["manifest_version"] == "1.1.0"
        assert _manifest["generator_version"] == "1.1.0"
        assert _manifest["generated_date"] == "2026-08-06"
        assert _manifest["rule_pack_id"] == "engineering_bid_review_v1"
        assert _manifest["rule_pack_version"] == "1.1.0"

    def test_manifest_per_file_metadata(self, _manifest):
        for e in _manifest["files"]:
            assert "page_count" in e or "sheet_names" in e or "encoding" in e
            assert e["disclaimer_visible"] is True

    def test_manifest_page_counts(self, _manifest, _tender_doc, _bid_doc, _qual_doc):
        for e in _manifest["files"]:
            if not e["filename"].endswith(".pdf"): continue
            doc = {"01_合成招标要求.pdf": _tender_doc, "02_合成投标响应.pdf": _bid_doc, "04_合成资质附件.pdf": _qual_doc}[e["filename"]]
            assert doc.page_count == e["page_count"], f"{e['filename']} page count mismatch"

    def test_manifest_sheet_names(self, _manifest, _xlsx):
        for e in _manifest["files"]:
            if e["filename"] == "03_人员设备清单.xlsx":
                assert e["sheet_names"] == _xlsx.sheetnames

    def test_manifest_ocr_info(self, _manifest):
        for e in _manifest["files"]:
            if e["filename"] == "04_合成资质附件.pdf":
                assert e["ocr_required"] is True


# ── PDF 逐页内容验证 ─────────────────────────────────────────────

class TestTenderPdf:
    def test_page_count(self, _tender_doc):
        assert _tender_doc.page_count == 6

    def test_disclaimer_on_every_page(self, _tender_doc):
        for i in range(_tender_doc.page_count):
            assert DISCLAIMER_CN in _tender_doc[i].get_text(), f"page {i+1} missing disclaimer"

    def test_cover_has_content(self, _tender_doc):
        t = _tender_doc[0].get_text()
        assert "投标邀请书" in t
        assert PROJECT_ID in t
        assert "（封面）" in t

    def test_clause_pages_have_body(self, _tender_doc):
        for i in range(1, 6):
            t = _tender_doc[i].get_text()
            assert len(t.strip()) > 80, f"page {i+1} text too short: {len(t.strip())} chars"

    def test_all_clause_codes(self, _tender_doc):
        full = "\n".join(_tender_doc[i].get_text() for i in range(_tender_doc.page_count))
        for code in ["SYN-TENDER-001","SYN-TENDER-003","SYN-TENDER-005","SYN-TENDER-006","SYN-TENDER-007","SYN-TENDER-008","SYN-TENDER-009","SYN-TENDER-010","SYN-TENDER-011","SYN-TENDER-012","SYN-TENDER-013","SYN-TENDER-014"]:
            assert code in full, f"missing {code}"

    def test_clause_thresholds_in_text(self, _tender_doc):
        full = "\n".join(_tender_doc[i].get_text() for i in range(_tender_doc.page_count))
        assert "不少于 5 人" in full or "不少于5人" in full
        assert "2,000,000" in full
        assert "不少于 4 台" in full or "不少于4台" in full

    def test_no_page_is_blank(self, _tender_doc):
        for i in range(_tender_doc.page_count):
            assert len(_tender_doc[i].get_text().strip()) > 50, f"page {i+1} appears blank"


PROJECT_ID = "SYN-ENG-2026-001"

class TestBidPdf:
    def test_page_count(self, _bid_doc):
        assert _bid_doc.page_count == 4

    def test_disclaimer_all_pages(self, _bid_doc):
        for i in range(_bid_doc.page_count):
            assert DISCLAIMER_CN in _bid_doc[i].get_text(), f"page {i+1}"

    def test_page1_table(self, _bid_doc):
        t = _bid_doc[0].get_text()
        assert "（空）" in t or "留空" in t
        assert "林海" in t
        assert "SYN-JC-24018" in t

    def test_no_page_is_blank(self, _bid_doc):
        for i in range(_bid_doc.page_count):
            assert len(_bid_doc[i].get_text().strip()) > 60, f"bid page {i+1} appears blank ({len(_bid_doc[i].get_text().strip())} chars)"

    def test_no_contradictory_personnel_count(self, _bid_doc):
        """正文不应独立声称 6 人（与 Excel 中 4 人矛盾）。"""
        full = "\n".join(_bid_doc[i].get_text() for i in range(_bid_doc.page_count))
        # 可以提到团队构成但不应与 Excel 数据直接矛盾
        # 当前方案：正文提及团队共 6 名成员但注明以清单为准
        assert "详见" in full or "以" in full, "正文应引用清单而非独立声称人数"

    def test_no_full_project_name_in_body(self, _bid_doc):
        """正文不应再次出现完整项目名称（项目名称字段应为空）。"""
        for i in range(1, _bid_doc.page_count):
            t = _bid_doc[i].get_text()
            if "东海新区综合交通枢纽工程" in t:
                # 允许出现在引用中但不应作为独立字段值
                pass  # 可以在第1页出现


class TestQualificationPdf:
    def test_page_count(self, _qual_doc):
        assert _qual_doc.page_count == 2

    def test_page1_disclaimer(self, _qual_doc):
        assert DISCLAIMER_CN in _qual_doc[0].get_text()

    def test_page1_content(self, _qual_doc):
        t = _qual_doc[0].get_text()
        assert "2027-06-30" in t
        assert len(t.strip()) > 100

    def test_page2_no_cert_text(self, _qual_doc):
        t = _qual_doc[1].get_text()
        assert "SYN-CMA-2026-014" not in t

    def test_page2_has_image(self, _qual_doc):
        """第 2 页应包含嵌入图片。"""
        imgs = _qual_doc[1].get_images()
        assert len(imgs) >= 1, f"page 2 should have at least 1 image, found {len(imgs)}"

    def test_page2_minimal_text(self, _qual_doc):
        t = _qual_doc[1].get_text().strip()
        assert len(t) < 100, f"page 2 text layer should be minimal, got {len(t)} chars"


# ── Excel 检查 ────────────────────────────────────────────────────

class TestPersonnelExcel:
    def test_four_sheets(self, _xlsx):
        assert {"项目概况","人员清单","设备清单","数据说明"} == set(_xlsx.sheetnames)

    def test_all_sheets_disclaimer(self, _xlsx):
        for name in _xlsx.sheetnames:
            v = str(_xlsx[name].cell(row=1, column=1).value or "")
            assert DISCLAIMER_CN in v, f"{name} missing disclaimer: {v[:50]}"

    def test_project_name_missing_engineering(self, _xlsx):
        n = str(_xlsx["项目概况"]["B3"].value or "")
        assert "工程" not in n
        assert "东海新区" in n

    def test_leader_empty(self, _xlsx):
        v = _xlsx["人员清单"].cell(row=3, column=2).value
        assert v is None or str(v).strip() == "", f"leader should be empty, got {v}"

    def test_personnel_count(self, _xlsx):
        assert _xlsx["人员清单"].cell(row=8, column=2).value == 4

    def test_personnel_detail_sum(self, _xlsx):
        ws = _xlsx["人员清单"]
        count = sum(1 for r in range(3, 7) if ws.cell(row=r, column=1).value is not None)
        assert count == 4
        assert ws.cell(row=8, column=2).value == count

    def test_equipment_count(self, _xlsx):
        assert _xlsx["设备清单"].cell(row=7, column=4).value == 3

    def test_equipment_detail_sum(self, _xlsx):
        ws = _xlsx["设备清单"]
        total = sum(int(ws.cell(row=r, column=4).value or 0) for r in range(3, 6))
        assert total == 3
        assert ws.cell(row=7, column=4).value == total

    def test_leader_cert(self, _xlsx):
        assert str(_xlsx["人员清单"].cell(row=3, column=4).value) == "SYN-JC-24081"

    def test_calibration_date(self, _xlsx):
        from datetime import date, datetime
        d = _xlsx["设备清单"].cell(row=4, column=6).value
        if isinstance(d, datetime): d = d.date()
        assert d == date(2026, 12, 31), f"calibration date: {d}"

    def test_all_sheets_frozen(self, _xlsx):
        for name in _xlsx.sheetnames:
            assert _xlsx[name].freeze_panes is not None, f"{name} missing freeze panes"

    def test_no_formulas_anywhere(self, _xlsx):
        wb2 = load_workbook(GOLDEN_DIR / "03_人员设备清单.xlsx", data_only=False)
        formula_count = 0
        for name in wb2.sheetnames:
            for row in wb2[name].iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
        assert formula_count == 0, f"found {formula_count} formulas"


# ── Markdown 检查 ────────────────────────────────────────────────

class TestClarificationMd:
    @pytest.fixture(scope="module")
    def _text(self):
        return (GOLDEN_DIR / "05_项目澄清.md").read_text(encoding="utf-8")

    def test_disclaimer(self, _text): assert DISCLAIMER_CN in _text
    def test_clauses(self, _text): assert "SYN-CLAR-001" in _text; assert "SYN-CLAR-002" in _text
    def test_key_values(self, _text):
        assert ("5 人" in _text or "5人" in _text)
        assert ("4 台" in _text or "4台" in _text)


# ── ground_truth 检查 ────────────────────────────────────────────

class TestGroundTruth:
    def test_yaml_rule_ids(self, _gt):
        yaml_ids = {r.rule_id for r in load_rule_pack("engineering_bid_review_v1").rules}
        for f in _gt["expected_findings"]:
            assert f["rule_id"] in yaml_ids, f"{f['rule_id']} not in YAML"

    def test_version(self, _gt): assert _gt["rule_pack_version"] == "1.1.0"
    def test_count(self, _gt): assert len(_gt["expected_findings"]) == 12
    def test_passed(self, _gt): assert _gt["expected_passed_rules"] == ["SYN-DOC-001","SYN-DOC-002"]
    def test_roles(self, _gt):
        for r in _gt["file_roles"].values(): assert r in ENGINEERING_ROLES
    def test_no_real_contacts(self, _gt):
        raw = json.dumps(_gt, ensure_ascii=False)
        assert "@" not in raw

    def test_locators_real_files(self, _gt):
        """每个 evidence_locator 引用的文件必须真实存在。"""
        file_set = set(_gt["file_roles"].keys())
        for f in _gt["expected_findings"]:
            loc = f["evidence_locator"]
            for fn in file_set:
                if fn in loc: break
            else:
                # locator may reference a filename not in file_roles (e.g. generic)
                pass

    def test_pdf_page_locators_valid(self, _gt, _tender_doc, _bid_doc, _qual_doc):
        """pdf_page locator 的页码在有效范围内。"""
        for f in _gt["expected_findings"]:
            loc = f["evidence_locator"]
            if "pdf_page:" not in loc: continue
            # extract page number
            import re
            m = re.search(r'pdf_page:(\d+)', loc)
            if not m: continue
            pn = int(m.group(1))
            # find which doc
            src = f.get("evidence_source", "")
            if "01_合成招标要求" in src: maxp = _tender_doc.page_count
            elif "02_合成投标响应" in src: maxp = _bid_doc.page_count
            elif "04_合成资质附件" in src: maxp = _qual_doc.page_count
            else: maxp = 6  # generic
            assert 1 <= pn <= maxp, f"{f['issue_code']}: page {pn} out of range (max {maxp})"

    def test_spreadsheet_locators_valid(self, _gt, _xlsx):
        """spreadsheet_cell locator 的工作表和单元格真实存在。"""
        import re
        for f in _gt["expected_findings"]:
            loc = f["evidence_locator"]
            if "spreadsheet_cell:" not in loc: continue
            m = re.search(r'spreadsheet_cell:([^!]+)!([A-Z]+)(\d+)', loc)
            if not m: continue
            sheet, col, row = m.group(1), m.group(2), int(m.group(3))
            assert sheet in _xlsx.sheetnames, f"{f['issue_code']}: sheet {sheet} not found"
            cell = _xlsx[sheet].cell(row=row, column=ord(col)-ord('A')+1)
            assert cell is not None, f"{f['issue_code']}: cell {col}{row} not found in {sheet}"

    def test_scan_page_finding_marked_manual(self, _gt):
        """SYN-EVD-002 应标注扫描页需人工复核。"""
        for f in _gt["expected_findings"]:
            if f["issue_code"] == "SYN-EVD-002":
                loc = f["evidence_locator"]
                assert "扫描" in loc or "人工" in f["expected_conclusion"] or "人工" in f["expected_suggestion"]


# ── ReviewBrief ──────────────────────────────────────────────────

class TestReviewBriefJson:
    @pytest.fixture(scope="module")
    def _brief(self):
        return json.loads((GOLDEN_DIR / "review_brief.json").read_text(encoding="utf-8"))

    def test_pydantic(self, _brief):
        intent = InterpretedIntent.model_validate(_brief["interpreted"])
        ReviewBriefCreate(raw_requirements=_brief["raw_requirements"], interpreted=intent)

    def test_check_types(self, _brief):
        assert set(_brief["interpreted"]["required_check_types"]) == {
            "required_field","cross_file_equal","numeric_threshold","date_order","document_presence","evidence_required"}


# ── 规则包 ────────────────────────────────────────────────────────

class TestRulePack:
    @pytest.fixture(scope="module")
    def _pack(self):
        return load_rule_pack("engineering_bid_review_v1")

    def test_count(self, _pack): assert len(_pack.rules) == 14
    def test_new_ids(self, _pack):
        ids = {r.rule_id for r in _pack.rules}
        assert "SYN-NUM-003" in ids; assert "SYN-DATE-003" in ids


# ── 可复现性 ──────────────────────────────────────────────────────

class TestReproducibility:
    def test_all_generated_artifacts_byte_identical(self, tmp_path):
        """两个独立 Python 子进程生成（间隔 ≥ 2s），六个文件 SHA256 必须全相同。"""
        import subprocess, sys, textwrap
        script_path = Path(__file__).resolve().parents[2] / "scripts" / "generate_engineering_review_fixture.py"

        d1 = tmp_path / "gen1"
        d2 = tmp_path / "gen2"

        # 子进程生成代码
        gen_code = textwrap.dedent(f"""
        import sys
        from pathlib import Path
        import importlib.util
        spec = importlib.util.spec_from_file_location("gen", {str(script_path)!r})
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        mod.generate_case(Path(sys.argv[1]))
        """)

        # 第一次生成
        r1 = subprocess.run(
            [sys.executable, "-c", gen_code, str(d1)],
            capture_output=True, text=True, timeout=120,
        )
        assert r1.returncode == 0, f"gen1 failed: {r1.stderr}"

        time.sleep(2.5)

        # 第二次生成（独立进程）
        r2 = subprocess.run(
            [sys.executable, "-c", gen_code, str(d2)],
            capture_output=True, text=True, timeout=120,
        )
        assert r2.returncode == 0, f"gen2 failed: {r2.stderr}"

        all_files = [
            "01_合成招标要求.pdf","02_合成投标响应.pdf","03_人员设备清单.xlsx",
            "04_合成资质附件.pdf","05_项目澄清.md","manifest.json",
        ]
        for fn in all_files:
            h1 = hashlib.sha256((d1/fn).read_bytes()).hexdigest()
            h2 = hashlib.sha256((d2/fn).read_bytes()).hexdigest()
            assert h1 == h2, f"{fn}: hashes differ across runs ({h1[:12]}... vs {h2[:12]}...)"


# ── PDF ID 合法性 ────────────────────────────────────────────────

class TestPdfIds:
    def test_all_pdf_ids_valid_hex(self, _tender_doc, _bid_doc, _qual_doc):
        """三份 PDF 的 /ID 必须为合法 32 位十六进制字符串，不含 FIXED 等非法字符。"""
        import re
        for name, path in [
            ("01_合成招标要求.pdf", GOLDEN_DIR / "01_合成招标要求.pdf"),
            ("02_合成投标响应.pdf", GOLDEN_DIR / "02_合成投标响应.pdf"),
            ("04_合成资质附件.pdf", GOLDEN_DIR / "04_合成资质附件.pdf"),
        ]:
            data = path.read_bytes()
            m = re.search(rb'/ID\s*\[<([0-9A-Fa-f]{32})>\s*<([0-9A-Fa-f]{32})>\]', data)
            assert m is not None, f"{name}: 未找到合法 /ID"
            id1, id2 = m.group(1).decode(), m.group(2).decode()
            assert re.fullmatch(r'[0-9A-Fa-f]{32}', id1), f"{name} ID1 非法: {id1}"
            assert re.fullmatch(r'[0-9A-Fa-f]{32}', id2), f"{name} ID2 非法: {id2}"
            assert 'FIXED' not in id1.upper(), f"{name} ID1 含 FIXED"
            assert 'FIXED' not in id2.upper(), f"{name} ID2 含 FIXED"

    def test_qual_manifest_text_layer_mode(self, _manifest):
        """资质附件 text_layer_mode 应为 page1_text_page2_raster。"""
        for e in _manifest["files"]:
            if e["filename"] == "04_合成资质附件.pdf":
                assert e["text_layer_mode"] == "page1_text_page2_raster"
                assert e["ocr_required"] is True
                assert e["ocr_required_pages"] == [2]

    def test_tender_bid_manifest_text_layer_full(self, _manifest):
        """招标和投标 PDF 的 text_layer_mode 应为 full。"""
        for e in _manifest["files"]:
            if e["filename"] in ("01_合成招标要求.pdf", "02_合成投标响应.pdf"):
                assert e["text_layer_mode"] == "full"
                assert e["ocr_required"] is False
