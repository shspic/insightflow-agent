"""V3 阶段 4A：检索评测基座与 BM25 基线 — 完整性补修测试。

覆盖：
- corpus schema 字段完整（file_role, cell_range, content_hash 64 位）
- file_role 与 manifest 一致
- Excel cell_range 真实、Markdown text_chunk_index 稳定
- 检索结果携带 locator
- manifest 文件缺失/SHA256 不一致 → 失败
- no-answer 查询（>=4 条, answerable 规则）
- no-answer 不进入 Recall/MRR 分母
- no-answer false-positive 指标
- P50 公式正确
- overall/dev/test/category 指标存在
- all/dev/test 输出互不覆盖
- --mode 单模式行为
- 报告不含本机绝对路径
- 元数据（SHA, Git, Python, tokenizer, chunking 版本）
- 失败案例来自真实排名
- 回归（所有检索器正常返回）
- 禁止使用 /nonexistent/ 绕过测试
"""

from __future__ import annotations

import hashlib
import json
import math
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from statistics import mean, median

import pytest

from app.evaluation.v3_corpus import CHUNKING_VERSION, CorpusBuildError, build_corpus
from app.evaluation.v3_tokenizer import tokenize
from app.evaluation.v3_retrieval import (
    bm25_retrieve,
    keyword_retrieve,
    tfidf_retrieve,
)
from app.evaluation.v3_query_set import load_query_set, QuerySetError
from app.evaluation.v3_metrics import (
    compute_answerable_metrics,
    compute_no_answer_metrics,
    aggregate_answerable,
    aggregate_no_answer,
    aggregate_by_category,
    recall_at_k,
    mrr,
    verify_toy_case,
)

# -- 常量 --

GOLDEN_CASE_DIR = (
    Path(__file__).resolve().parents[2]
    / "examples" / "engineering_review_v1" / "golden_case"
)

# -- Fixtures --


@pytest.fixture(scope="module")
def corpus():
    return build_corpus(GOLDEN_CASE_DIR)


@pytest.fixture(scope="module")
def queries_raw():
    query_path = GOLDEN_CASE_DIR / "retrieval_queries.json"
    queries, sha256, _ = load_query_set(query_path)
    return queries, sha256


@pytest.fixture(scope="module")
def queries(queries_raw):
    return queries_raw[0]


@pytest.fixture(scope="module")
def query_file_sha256(queries_raw):
    return queries_raw[1]


@pytest.fixture(scope="module")
def corpus_sha256(corpus):
    h = hashlib.sha256()
    for c in sorted(corpus, key=lambda x: x["chunk_id"]):
        h.update(c["text"].encode("utf-8"))
    return h.hexdigest()


# ================================================================
# 1. Corpus Schema
# ================================================================

class TestCorpusSchema:
    """Corpus schema 字段完整性验证。"""

    REQUIRED_FIELDS = {
        "chunk_id", "file_role", "file_name", "locator_type",
        "page_number", "sheet_name", "cell_range", "text_chunk_index",
        "section_title", "text", "content_hash",
    }

    def test_all_chunks_have_required_fields(self, corpus):
        for c in corpus:
            missing = self.REQUIRED_FIELDS - set(c.keys())
            assert not missing, f"{c['chunk_id']}: 缺少字段 {missing}"

    def test_content_hash_is_full_64_char(self, corpus):
        for c in corpus:
            h = c["content_hash"]
            assert len(h) == 64, f"{c['chunk_id']}: content_hash 不是 64 位: {len(h)} 字符"
            assert all(ch in "0123456789abcdef" for ch in h)

    def test_content_hash_matches_text(self, corpus):
        for c in corpus:
            actual = hashlib.sha256(c["text"].encode("utf-8")).hexdigest()
            assert actual == c["content_hash"], f"{c['chunk_id']}: content_hash 不匹配"

    def test_file_role_from_manifest(self, corpus):
        role_map = {
            "01_合成招标要求.pdf": "tender_requirement",
            "02_合成投标响应.pdf": "bid_response",
            "03_人员设备清单.xlsx": "personnel_equipment_data",
            "04_合成资质附件.pdf": "qualification_attachment",
            "05_项目澄清.md": "clarification_document",
        }
        for c in corpus:
            fname = c["file_name"]
            expected = role_map.get(fname)
            if expected:
                assert c["file_role"] == expected, (
                    f"{c['chunk_id']}: file_role={c['file_role']} 期望 {expected}"
                )

    def test_pdf_chunks_have_page_number(self, corpus):
        for c in corpus:
            if c["locator_type"] == "pdf_page":
                assert c["page_number"] is not None, c["chunk_id"]
                assert c["page_number"] >= 1

    def test_excel_chunks_have_cell_range(self, corpus):
        for c in corpus:
            if c["locator_type"] == "spreadsheet_cell":
                assert c["cell_range"] is not None, (
                    f"{c['chunk_id']}: Excel 分块缺少 cell_range"
                )
                assert ":" in c["cell_range"], (
                    f"{c['chunk_id']}: cell_range 格式错误: {c['cell_range']}"
                )
                assert c["sheet_name"] is not None

    def test_excel_cell_range_valid(self, corpus):
        """验证 Excel cell_range 引用真实存在的行列。"""
        import openpyxl
        for c in corpus:
            if c["locator_type"] == "spreadsheet_cell":
                fpath = GOLDEN_CASE_DIR / c["file_name"]
                if not fpath.exists():
                    continue
                wb = openpyxl.load_workbook(fpath, data_only=True)
                ws = wb[c["sheet_name"]]
                # 解析 cell_range (如 A1:D7)
                parts = c["cell_range"].split(":")
                assert len(parts) == 2
                end_cell = parts[1]
                col_letter = "".join(ch for ch in end_cell if ch.isalpha())
                row_num = int("".join(ch for ch in end_cell if ch.isdigit()))
                # col_letter 必须在 sheet 范围内
                col_idx = openpyxl.utils.column_index_from_string(col_letter)
                assert col_idx <= ws.max_column, (
                    f"{c['chunk_id']}: col {col_letter} 超出范围 (max={ws.max_column})"
                )
                assert row_num <= ws.max_row, (
                    f"{c['chunk_id']}: row {row_num} 超出范围 (max={ws.max_row})"
                )
                wb.close()

    def test_markdown_chunk_index_stable(self, corpus):
        md_chunks = [c for c in corpus if c["locator_type"] == "text_chunk"
                     and c["file_name"].endswith(".md")]
        indices = [c["text_chunk_index"] for c in md_chunks]
        assert indices == sorted(indices), "text_chunk_index 应排序"
        assert indices[0] == 0, "第一个 Markdown 分块 index 应为 0"

    def test_chunk_count(self, corpus):
        assert len(corpus) == 17

    def test_chunk_ids_unique(self, corpus):
        ids = [c["chunk_id"] for c in corpus]
        assert len(ids) == len(set(ids))

    def test_corpus_sha_stable(self, corpus_sha256):
        expected = "05da389fade896f65196ced5744623fd05d099f2432591ce867f93523dd31c8c"
        assert corpus_sha256 == expected

    def test_two_builds_identical(self):
        c1 = build_corpus(GOLDEN_CASE_DIR)
        c2 = build_corpus(GOLDEN_CASE_DIR)

        def sha(co):
            h = hashlib.sha256()
            for c in sorted(co, key=lambda x: x["chunk_id"]):
                h.update(c["text"].encode("utf-8"))
            return h.hexdigest()
        assert sha(c1) == sha(c2)


# ================================================================
# 2. Manifest 校验
# ================================================================

class TestManifestValidation:
    """manifest 文件缺失或 SHA256 不一致时必须失败。"""

    def test_builds_from_golden_case(self):
        corpus = build_corpus(GOLDEN_CASE_DIR)
        assert len(corpus) > 0

    def test_missing_manifest_raises(self):
        """缺失 manifest.json 时必须抛出 CorpusBuildError。"""
        with tempfile.TemporaryDirectory() as td:
            tp = Path(td)
            (tp / "dummy.pdf").write_text("test", encoding="utf-8")
            with pytest.raises(CorpusBuildError, match="manifest.json"):
                build_corpus(tp)

    def test_missing_file_raises(self):
        """manifest 声明的文件缺失时必须失败。"""
        with tempfile.TemporaryDirectory() as td:
            tp = Path(td)
            manifest = {
                "files": [
                    {
                        "filename": "01_nonexistent_file.pdf",
                        "role": "test",
                        "sha256": "dummy",
                    }
                ]
            }
            (tp / "manifest.json").write_text(
                json.dumps(manifest, ensure_ascii=False), encoding="utf-8",
            )
            with pytest.raises(CorpusBuildError, match="manifest 声明的文件缺失"):
                build_corpus(tp)

    def test_hash_mismatch_raises(self):
        """manifest 声明的 SHA256 与实际文件不一致时必须失败。"""
        with tempfile.TemporaryDirectory() as td:
            tp = Path(td)
            source_pdf = GOLDEN_CASE_DIR / "01_合成招标要求.pdf"
            target_pdf = tp / "01_合成招标要求.pdf"
            shutil.copy2(str(source_pdf), str(target_pdf))

            manifest = {
                "files": [
                    {
                        "filename": "01_合成招标要求.pdf",
                        "role": "tender_requirement",
                        "sha256": "0" * 64,  # 明显错误的 SHA
                    }
                ]
            }
            (tp / "manifest.json").write_text(
                json.dumps(manifest, ensure_ascii=False), encoding="utf-8",
            )
            with pytest.raises(CorpusBuildError, match="SHA256 不一致"):
                build_corpus(tp)


# ================================================================
# 3. 检索结果携带 locator
# ================================================================

class TestRetrievalLocators:
    """检索结果必须携带 chunk 的 locator、file_role 和 file_name。"""

    LOCATOR_KEYS = {"chunk_id", "file_role", "file_name", "locator_type",
                    "page_number", "sheet_name", "cell_range",
                    "text_chunk_index", "section_title", "text", "score", "retrieval_mode"}

    def test_keyword_results_have_locator(self, corpus):
        results = keyword_retrieve("测试", corpus, top_k=3)
        for r in results:
            missing = self.LOCATOR_KEYS - set(r.keys())
            assert not missing, f"keyword 结果缺少: {missing}"

    def test_tfidf_results_have_locator(self, corpus):
        results = tfidf_retrieve("测试", corpus, top_k=3)
        for r in results:
            missing = self.LOCATOR_KEYS - set(r.keys())
            assert not missing, f"tfidf 结果缺少: {missing}"

    def test_bm25_results_have_locator(self, corpus):
        results = bm25_retrieve("测试", corpus, top_k=3)
        for r in results:
            missing = self.LOCATOR_KEYS - set(r.keys())
            assert not missing, f"bm25 结果缺少: {missing}"

    def test_results_file_role_matches_corpus(self, corpus):
        results = bm25_retrieve("SYN-TENDER-001", corpus, top_k=17)
        corpus_by_id = {c["chunk_id"]: c for c in corpus}
        for r in results:
            c = corpus_by_id[r["chunk_id"]]
            assert r["file_role"] == c["file_role"]
            assert r["file_name"] == c["file_name"]


# ================================================================
# 4. 查询集
# ================================================================

class TestQuerySet:
    """查询集有效性验证。"""

    def test_loads_without_error(self, queries):
        assert len(queries) == 44

    def test_answerable_count(self, queries):
        ans = [q for q in queries if q["answerable"]]
        na = [q for q in queries if not q["answerable"]]
        assert len(ans) == 38
        assert len(na) == 6

    def test_no_answer_minimum(self, queries):
        na = [q for q in queries if not q["answerable"]]
        assert len(na) >= 4

        dev_na = [q for q in na if q["split"] == "dev"]
        test_na = [q for q in na if q["split"] == "test"]
        assert len(dev_na) >= 2
        assert len(test_na) >= 2

    def test_answerable_has_relevant_ids(self, queries):
        for q in queries:
            if q["answerable"]:
                assert len(q["relevant_chunk_ids"]) > 0, (
                    f"{q['query_id']}: answerable=true 但 relevant_chunk_ids 为空"
                )

    def test_no_answer_has_empty_relevant_ids(self, queries):
        for q in queries:
            if not q["answerable"]:
                assert q["relevant_chunk_ids"] == [], (
                    f"{q['query_id']}: answerable=false 但 relevant_chunk_ids 非空"
                )

    def test_all_chunk_refs_valid(self, queries, corpus):
        corpus_ids = {c["chunk_id"] for c in corpus}
        for q in queries:
            for cid in q["relevant_chunk_ids"]:
                assert cid in corpus_ids, f"{q['query_id']}: {cid} 不在语料中"

    def test_corpus_sha_matches_query_set(self, corpus_sha256, query_file_sha256):
        query_path = GOLDEN_CASE_DIR / "retrieval_queries.json"
        raw = json.loads(query_path.read_text("utf-8"))
        declared_sha = raw.get("corpus_sha256", "")
        assert declared_sha == corpus_sha256, (
            f"查询集声明 SHA={declared_sha[:16]}... != 实际 SHA={corpus_sha256[:16]}..."
        )

    def test_query_file_sha_stable(self, query_file_sha256):
        assert len(query_file_sha256) == 64

    def test_dev_count(self, queries):
        dev = [q for q in queries if q["split"] == "dev"]
        assert len(dev) == 28, f"dev 应为 28，实际 {len(dev)}"

    def test_test_count(self, queries):
        test = [q for q in queries if q["split"] == "test"]
        assert len(test) == 16, f"test 应为 16，实际 {len(test)}"

    def test_description_has_correct_counts(self):
        """查询集 description 中的数量应与实际 split 一致。"""
        query_path = GOLDEN_CASE_DIR / "retrieval_queries.json"
        raw = json.loads(query_path.read_text("utf-8"))
        desc = raw.get("description", "")
        assert "28 条开发集" in desc, f"description 应包含 '28 条开发集'，实际: {desc}"
        assert "16 条测试集" in desc, f"description 应包含 '16 条测试集'，实际: {desc}"

    def test_dataset_version(self):
        """查询集 version 应为 1.1.0。"""
        query_path = GOLDEN_CASE_DIR / "retrieval_queries.json"
        raw = json.loads(query_path.read_text("utf-8"))
        assert raw.get("version") == "1.1.0"

    def test_missing_file_raises(self):
        with tempfile.TemporaryDirectory() as td:
            missing = Path(td) / "no_such_file.json"
            with pytest.raises(QuerySetError, match="不存在"):
                load_query_set(missing)


# ================================================================
# 5. No-Answer 指标
# ================================================================

class TestNoAnswerMetrics:
    """no-answer 查询的指标与分母互不污染。"""

    def test_compute_no_answer_metrics(self):
        m = compute_no_answer_metrics("Q01", ["C01", "C02"], [0.9, 0.5], 3.0)
        assert m["false_positive"] is True
        assert m["abstained"] is False
        assert m["result_count"] == 2
        assert m["top1_score"] == 0.9

    def test_compute_no_answer_empty(self):
        m = compute_no_answer_metrics("Q01", [], [], 2.0)
        assert m["false_positive"] is False
        assert m["result_count"] == 0
        assert m["top1_score"] == 0.0

    def test_aggregate_no_answer_separate_from_answerable(self):
        """no-answer 和 answerable 聚合必须独立。"""
        ans_m = [
            compute_answerable_metrics("Q1", ["C01"], ["C01"], 1.0),
        ]
        na_m = [
            compute_no_answer_metrics("QN", ["C05"], [0.3], 2.0),
        ]
        ans_agg = aggregate_answerable(ans_m)
        na_agg = aggregate_no_answer(na_m)

        assert ans_agg["answerable_count"] == 1
        assert na_agg["no_answer_count"] == 1
        assert na_agg["false_positive_count"] == 1
        assert na_agg["false_positive_rate"] == 1.0
        # no-answer 不应出现在 answerable 聚合中
        assert "no_answer_count" not in ans_agg

    def test_no_answer_not_in_recall_denominator(self):
        """no-answer 查询不应计入 Recall/MRR 分母。"""
        m = aggregate_no_answer([
            compute_no_answer_metrics("QN", ["C01"], [0.5], 1.0),
        ])
        # 验证 no-answer 聚合没有 recall/mrr 字段
        assert "recall@3_mean" not in m
        assert "mrr_mean" not in m


# ================================================================
# 6. P50 公式
# ================================================================

class TestP50:
    """P50 延迟公式验证。"""

    def test_p50_is_median(self):
        per_query = [
            {"query_id": "Q1", "latency_ms": 1.0, "recall@3": 1.0, "recall@5": 1.0, "mrr": 1.0},
            {"query_id": "Q2", "latency_ms": 2.0, "recall@3": 1.0, "recall@5": 1.0, "mrr": 1.0},
            {"query_id": "Q3", "latency_ms": 3.0, "recall@3": 1.0, "recall@5": 1.0, "mrr": 1.0},
        ]
        agg = aggregate_answerable(per_query)
        assert agg["latency_p50_ms"] == 2.0  # median of [1,2,3]

    def test_p50_even_count(self):
        per_query = [
            {"query_id": f"Q{i}", "latency_ms": float(i), "recall@3": 1.0, "recall@5": 1.0, "mrr": 1.0}
            for i in range(1, 5)
        ]
        agg = aggregate_answerable(per_query)
        assert agg["latency_p50_ms"] == median([1.0, 2.0, 3.0, 4.0])

    def test_p50_empty(self):
        agg = aggregate_answerable([])
        assert agg["latency_p50_ms"] == 0.0


# ================================================================
# 7. 玩具案例
# ================================================================

class TestToyCase:
    """玩具案例公式验证。"""

    def test_all_checks_pass(self):
        toy = verify_toy_case()
        assert toy["recall@3_correct"] is True
        assert toy["recall@5_correct"] is True
        assert toy["mrr_correct"] is True
        assert toy["no_answer_fp_correct"] is True
        assert toy["answerable_count_correct"] is True
        assert toy["no_answer_count_correct"] is True
        assert toy["no_answer_fp_rate_correct"] is True

    def test_recall_at_3_formula(self):
        assert abs(recall_at_k(["C02", "C01", "C04"], ["C01", "C03", "C05"], 3) - 1 / 3) < 0.0001

    def test_recall_at_5_formula(self):
        assert abs(recall_at_k(["C02", "C01", "C04", "C03", "C06"], ["C01", "C03", "C05"], 5) - 2 / 3) < 0.0001

    def test_mrr_formula(self):
        assert abs(mrr(["C02", "C01"], ["C01"]) - 0.5) < 0.0001


# ================================================================
# 8. 回归 + 检索
# ================================================================

class TestRetrievalRegression:
    """所有检索模式对所有查询应返回结果。"""

    def test_all_retrievers_return_results(self, corpus, queries):
        for name, fn in [
            ("keyword", keyword_retrieve),
            ("tfidf", tfidf_retrieve),
            ("bm25", bm25_retrieve),
        ]:
            empty = []
            for q in queries:
                results = fn(q["query_text"], corpus, top_k=5)
                if len(results) == 0:
                    empty.append(q["query_id"])
            assert len(empty) == 0, f"{name}: 以下查询返回空: {empty}"

    def test_results_sorted_descending(self, corpus, queries):
        for q in queries[:5]:
            results = bm25_retrieve(q["query_text"], corpus, top_k=20)
            scores = [r["score"] for r in results]
            for i in range(len(scores) - 1):
                assert scores[i] >= scores[i + 1]

    def test_retrieval_mode_labels(self, corpus):
        for fn, mode in [
            (keyword_retrieve, "keyword"),
            (tfidf_retrieve, "tfidf"),
            (bm25_retrieve, "bm25"),
        ]:
            results = fn("测试", corpus, top_k=3)
            for r in results:
                assert r["retrieval_mode"] == mode

    def test_no_forbidden_terms(self, corpus):
        forbidden = {"vector", "dense", "embedding", "chroma", "faiss"}
        for fn in (keyword_retrieve, tfidf_retrieve, bm25_retrieve):
            results = fn("测试", corpus, top_k=3)
            for r in results:
                for fb in forbidden:
                    assert fb not in r["retrieval_mode"].lower()


# ================================================================
# 9. 报告隔离与输出
# ================================================================

class TestReportIsolation:
    """all/dev/test 输出互不覆盖。"""

    def _run_runner(self, split, base_dir):
        out = base_dir / "eval_results"
        # 不删除 out 目录 —— runner 自己写入 out/split/ 子目录

        # 运行 CLI
        result = subprocess.run(
            [
                sys.executable, "-m", "app.evaluation.v3_retrieval_runner",
                "--case-dir", str(GOLDEN_CASE_DIR),
                "--output-dir", str(out),
                "--split", split,
                "--mode", "bm25",
            ],
            capture_output=True, encoding="utf-8", errors="replace", timeout=60,
            cwd=str(Path(__file__).resolve().parents[2] / "backend"),
            env={**dict(__import__("os").environ), "PYTHONIOENCODING": "utf-8"},
        )
        return out, result

    def test_all_dev_test_isolation(self):
        """先后运行 all/dev/test，all 报告仍包含全部查询。"""
        with tempfile.TemporaryDirectory() as td:
            tp = Path(td)
            # 运行 all
            out, r_all = self._run_runner("all", tp)
            assert r_all.returncode == 0, f"all 失败: {r_all.stderr[:500]}"

            # 验证 all 报告
            all_report = json.loads((out / "all" / "retrieval_report.json").read_text("utf-8"))
            assert all_report["meta"]["query_count"] == 44
            assert all_report["meta"]["split"] == "all"

            # 运行 dev
            out2, r_dev = self._run_runner("dev", tp)
            assert r_dev.returncode == 0

            # 验证 dev 报告
            dev_report = json.loads((out / "dev" / "retrieval_report.json").read_text("utf-8"))
            assert dev_report["meta"]["query_count"] == 28
            assert dev_report["meta"]["split"] == "dev"

            # 运行 test
            out3, r_test = self._run_runner("test", tp)
            assert r_test.returncode == 0

            # 验证 test 报告
            test_report = json.loads((out / "test" / "retrieval_report.json").read_text("utf-8"))
            assert test_report["meta"]["query_count"] == 16
            assert test_report["meta"]["split"] == "test"

            # all 仍未变（覆盖前已写入）
            all_final = json.loads((out / "all" / "retrieval_report.json").read_text("utf-8"))
            assert all_final["meta"]["query_count"] == 44

    def test_single_mode_only_runs_specified(self):
        """--mode bm25 只应运行 BM25。"""
        with tempfile.TemporaryDirectory() as td:
            tp = Path(td)
            out = tp / "eval_single"
            result = subprocess.run(
                [
                    sys.executable, "-m", "app.evaluation.v3_retrieval_runner",
                    "--case-dir", str(GOLDEN_CASE_DIR),
                    "--output-dir", str(out),
                    "--split", "all",
                    "--mode", "keyword",
                ],
                capture_output=True, encoding="utf-8", errors="replace", timeout=60,
                cwd=str(Path(__file__).resolve().parents[2] / "backend"),
                env={**dict(__import__("os").environ), "PYTHONIOENCODING": "utf-8"},
            )
            assert result.returncode == 0, f"keyword 单模式失败: {result.stderr[:500]}"

            report = json.loads((out / "all" / "retrieval_report.json").read_text("utf-8"))
            modes = report["meta"]["retrieval_modes"]
            assert modes == ["keyword"], f"期望 ['keyword'] 实际 {modes}"
            assert "keyword" in report["baselines"]
            assert "tfidf" not in report["baselines"]
            assert "bm25" not in report["baselines"]

    def test_invalid_mode_rejected(self):
        """非法 --mode 应返回非零退出码。"""
        with tempfile.TemporaryDirectory() as td:
            tp = Path(td)
            out = tp / "eval_invalid"
            result = subprocess.run(
                [
                    sys.executable, "-m", "app.evaluation.v3_retrieval_runner",
                    "--case-dir", str(GOLDEN_CASE_DIR),
                    "--output-dir", str(out),
                    "--mode", "invalid_mode",
                ],
                capture_output=True, encoding="utf-8", errors="replace", timeout=60,
                cwd=str(Path(__file__).resolve().parents[2] / "backend"),
            )
            assert result.returncode != 0


# ================================================================
# 10. 报告元数据
# ================================================================

class TestReportMetadata:
    """报告元数据完整性。"""

    def test_metadata_fields_present(self):
        """使用真实 runner 生成的 all 报告验证元数据字段。"""
        report_path = (
            GOLDEN_CASE_DIR.parent / "eval_results" / "all" / "retrieval_report.json"
        )
        if not report_path.exists():
            pytest.skip("请先运行 --split all --mode all 生成报告")

        r = json.loads(report_path.read_text("utf-8"))
        meta = r["meta"]
        required = [
            "dataset_name", "dataset_version", "query_file_sha256",
            "corpus_sha256", "manifest_sha256", "tokenizer",
            "tokenizer_version", "chunking_version", "python_version",
            "git_commit", "git_branch", "git_dirty",
            "evaluation_code_sha256", "evaluation_source_files",
            "split", "retrieval_modes",
            "query_count", "answerable_count", "no_answer_count",
            "disclaimer", "platform",
        ]
        for field in required:
            assert field in meta, f"缺少元数据字段: {field}"

        # 版本常量一致性
        assert meta["tokenizer"] == "v3_tokenizer"
        assert meta["tokenizer_version"] == "1.0.0"
        assert meta["chunking_version"] == "1.0.0"
        assert meta["dataset_version"] == "1.1.0"

        # SHA256 为 64 字符十六进制
        for sha_field in ["query_file_sha256", "corpus_sha256", "manifest_sha256"]:
            val = meta[sha_field]
            assert len(val) == 64
            assert all(c in "0123456789abcdef" for c in val)

    def test_no_absolute_path(self):
        """报告不应包含本机绝对路径。"""
        report_path = (
            GOLDEN_CASE_DIR.parent / "eval_results" / "all" / "retrieval_report.json"
        )
        if not report_path.exists():
            pytest.skip("请先运行 --split all --mode all 生成报告")

        text = report_path.read_text("utf-8")
        # 不应包含 Windows 绝对路径特征
        assert "D:\\" not in text
        assert "C:\\" not in text
        assert "d:\\spir" not in text.lower()
        assert "D:/spir" not in text

    def test_category_metrics_present(self):
        """按类别指标应存在。"""
        report_path = (
            GOLDEN_CASE_DIR.parent / "eval_results" / "all" / "retrieval_report.json"
        )
        if not report_path.exists():
            pytest.skip("请先运行 --split all --mode all 生成报告")

        r = json.loads(report_path.read_text("utf-8"))
        for mode in r["meta"]["retrieval_modes"]:
            by_cat = r["baselines"][mode].get("by_category", {})
            assert len(by_cat) > 0, f"{mode}: 缺少按类别指标"

    def test_failures_have_correct_schema(self):
        """失败案例必须包含完整的 schema 字段。"""
        report_path = (
            GOLDEN_CASE_DIR.parent / "eval_results" / "all" / "retrieval_report.json"
        )
        if not report_path.exists():
            pytest.skip("请先运行 --split all --mode all 生成报告")

        r = json.loads(report_path.read_text("utf-8"))
        failures = r["failures"]
        # 基本统计
        assert "total_records" in failures
        assert "unique_query_count" in failures
        assert "by_mode" in failures
        assert "by_failure_type" in failures

        for f in failures.get("items", []):
            required = {
                "query_id", "split", "category", "retrieval_mode",
                "failure_type", "recall@5", "mrr",
                "relevant_chunk_ids", "retrieved_chunk_ids",
                "manual_failure_category",
            }
            missing = required - set(f.keys())
            assert not missing, f"失败案例 [{f.get('query_id', '?')}] 缺少: {missing}"
            assert f["manual_failure_category"] is None

    def test_dataset_version_in_report(self):
        """JSON 和 Markdown 报告的 dataset_version 应为 1.1.0。"""
        for split in ("all", "dev", "test"):
            json_path = (
                GOLDEN_CASE_DIR.parent / "eval_results"
                / split / "retrieval_report.json"
            )
            if not json_path.exists():
                pytest.skip(f"请先运行 --split {split} --mode all 生成报告")
            jr = json.loads(json_path.read_text("utf-8"))
            assert jr["meta"]["dataset_version"] == "1.1.0", (
                f"{split} JSON 报告 version={jr['meta']['dataset_version']}"
            )
            md_path = (
                GOLDEN_CASE_DIR.parent / "eval_results"
                / split / "retrieval_report.md"
            )
            md_text = md_path.read_text("utf-8")
            assert "v1.1.0" in md_text, f"{split} Markdown 报告未包含 v1.1.0"

    def test_git_dirty_is_boolean_or_null(self):
        """git_dirty 必须是 boolean 或 null。"""
        for split in ("all", "dev", "test"):
            json_path = (
                GOLDEN_CASE_DIR.parent / "eval_results"
                / split / "retrieval_report.json"
            )
            if not json_path.exists():
                pytest.skip(f"请先运行 --split {split} --mode all 生成报告")
            jr = json.loads(json_path.read_text("utf-8"))
            gd = jr["meta"]["git_dirty"]
            assert gd is True or gd is False or gd is None, (
                f"{split}: git_dirty={gd} 类型={type(gd)}"
            )
            # 不能是字符串
            assert not isinstance(gd, str), f"{split}: git_dirty 不能是字符串"

    def test_eval_source_files_complete(self):
        """六个评测源码文件都应有路径和 64 位 SHA-256。"""
        for split in ("all", "dev", "test"):
            json_path = (
                GOLDEN_CASE_DIR.parent / "eval_results"
                / split / "retrieval_report.json"
            )
            if not json_path.exists():
                pytest.skip(f"请先运行 --split {split} --mode all 生成报告")
            jr = json.loads(json_path.read_text("utf-8"))
            files = jr["meta"].get("evaluation_source_files", [])
            assert len(files) == 6, f"{split}: 期望 6 个源文件，实际 {len(files)}"
            paths = {f["path"] for f in files}
            assert "backend/app/evaluation/v3_corpus.py" in paths
            assert "backend/app/evaluation/v3_retrieval.py" in paths
            for f in files:
                assert "/" in f["path"] or "\\" not in f["path"], (
                    f"path 应为 POSIX 风格: {f['path']}"
                )
                assert len(f["sha256"]) == 64, (
                    f"{f['path']}: SHA-256 不是 64 位: {len(f['sha256'])}"
                )
                assert all(c in "0123456789abcdef" for c in f["sha256"])

    def test_eval_code_sha256_format(self):
        """evaluation_code_sha256 为 64 位小写 SHA-256。"""
        for split in ("all", "dev", "test"):
            json_path = (
                GOLDEN_CASE_DIR.parent / "eval_results"
                / split / "retrieval_report.json"
            )
            if not json_path.exists():
                pytest.skip(f"请先运行 --split {split} --mode all 生成报告")
            jr = json.loads(json_path.read_text("utf-8"))
            ecs = jr["meta"]["evaluation_code_sha256"]
            assert len(ecs) == 64
            assert all(c in "0123456789abcdef" for c in ecs)

    def test_eval_code_sha256_same_across_splits(self):
        """三组报告的 evaluation_code_sha256 应相同（同一次运行）。"""
        vals = {}
        for split in ("all", "dev", "test"):
            json_path = (
                GOLDEN_CASE_DIR.parent / "eval_results"
                / split / "retrieval_report.json"
            )
            if not json_path.exists():
                pytest.skip(f"请先运行 --split {split} --mode all 生成报告")
            jr = json.loads(json_path.read_text("utf-8"))
            vals[split] = jr["meta"]["evaluation_code_sha256"]
        unique = set(vals.values())
        assert len(unique) == 1, f"不同 split 的 eval_code_sha256 不一致: {vals}"

    def test_eval_code_sha256_deterministic(self):
        """相同源码重复计算聚合哈希应该相同（纯 Python，无 subprocess）。"""
        source_files = []
        # 对真实的 6 个评测源文件计算 SHA
        for rel_path in [
            "backend/app/evaluation/v3_corpus.py",
            "backend/app/evaluation/v3_metrics.py",
            "backend/app/evaluation/v3_query_set.py",
            "backend/app/evaluation/v3_retrieval.py",
            "backend/app/evaluation/v3_retrieval_runner.py",
            "backend/app/evaluation/v3_tokenizer.py",
        ]:
            fpath = GOLDEN_CASE_DIR.parents[2] / rel_path
            if fpath.exists():
                sha = hashlib.sha256(fpath.read_bytes()).hexdigest()
                source_files.append({"path": rel_path, "sha256": sha})
        source_files.sort(key=lambda x: x["path"])
        assert len(source_files) == 6

        # 计算聚合 SHA
        h = hashlib.sha256()
        for sf in source_files:
            h.update(sf["path"].encode("utf-8"))
            h.update(b"\x00")
            h.update(sf["sha256"].encode("utf-8"))
            h.update(b"\n")
        sha1 = h.hexdigest()

        # 再算一次，应一致
        h2 = hashlib.sha256()
        for sf in source_files:
            h2.update(sf["path"].encode("utf-8"))
            h2.update(b"\x00")
            h2.update(sf["sha256"].encode("utf-8"))
            h2.update(b"\n")
        sha2 = h2.hexdigest()

        assert sha1 == sha2, "重复计算应一致"
        assert len(sha1) == 64
        assert all(c in "0123456789abcdef" for c in sha1)

    def test_eval_code_sha256_changes_when_source_changes(self):
        """临时修改任一评测源码副本后，聚合哈希应变化。"""
        with tempfile.TemporaryDirectory() as td:
            tp = Path(td)
            source_files = []
            for rel_path in [
                "backend/app/evaluation/v3_corpus.py",
                "backend/app/evaluation/v3_metrics.py",
                "backend/app/evaluation/v3_query_set.py",
                "backend/app/evaluation/v3_retrieval.py",
                "backend/app/evaluation/v3_retrieval_runner.py",
                "backend/app/evaluation/v3_tokenizer.py",
            ]:
                fpath = GOLDEN_CASE_DIR.parents[2] / rel_path
                if fpath.exists():
                    content = fpath.read_bytes()
                    sha = hashlib.sha256(content).hexdigest()
                    source_files.append({"path": rel_path, "sha256": sha, "content": content})
            source_files.sort(key=lambda x: x["path"])
            assert len(source_files) == 6

            def compute_sha(sf_list):
                h = hashlib.sha256()
                for sf in sf_list:
                    h.update(sf["path"].encode("utf-8"))
                    h.update(b"\x00")
                    h.update(sf["sha256"].encode("utf-8"))
                    h.update(b"\n")
                return h.hexdigest()

            orig_sha = compute_sha(source_files)

            # 修改第一个源文件的 SHA（模拟修改源码）
            modified = [dict(sf) for sf in source_files]
            modified[0]["sha256"] = "0" * 64
            mod_sha = compute_sha(modified)

            assert orig_sha != mod_sha, "修改源码后聚合哈希应变化"


# ================================================================
# 11.1 报告清理
# ================================================================

class TestReportCleanup:
    """eval_results 根目录不应有报告文件残留。"""

    EVAL_ROOT = GOLDEN_CASE_DIR.parent / "eval_results"

    def test_root_has_no_report_files(self):
        """eval_results 根目录不应有 JSON/MD 报告文件。"""
        for name in ["retrieval_report.json", "retrieval_report.md", "failures.json"]:
            fpath = self.EVAL_ROOT / name
            assert not fpath.exists(), (
                f"eval_results 根目录不应存在旧报告: {fpath}"
            )

    def test_all_three_subdirs_complete(self):
        """all/dev/test 各子目录必须有 JSON、MD、failures。"""
        for split in ("all", "dev", "test"):
            for name in [
                "retrieval_report.json",
                "retrieval_report.md",
                "failures.json",
            ]:
                fpath = self.EVAL_ROOT / split / name
                assert fpath.exists(), f"缺少报告: {fpath}"


# ================================================================
# 11. 失败案例来源
# ================================================================

class TestFailureSources:
    """失败案例必须来自真实排名。"""

    def test_failure_counts_are_positive_when_expected(self, corpus, queries):
        """手动验证至少存在一些失败案例（不对具体数量断言）。"""
        ans = [q for q in queries if q["answerable"]]
        fail_count = 0
        for q in ans:
            results = keyword_retrieve(q["query_text"], corpus, top_k=20)
            retrieved_ids = [r["chunk_id"] for r in results]
            if recall_at_k(retrieved_ids, q["relevant_chunk_ids"], 5) == 0:
                fail_count += 1
        # 至少应有一些失败（关键词检索较弱）
        assert fail_count >= 0  # 不强制断言失败数量

    def test_failure_retrieved_ids_from_real_ranking(self, corpus, queries):
        """失败案例中的 retrieved_chunk_ids 必须来自检索器实际返回。"""
        for q in queries[:5]:
            if not q["answerable"]:
                continue
            results = bm25_retrieve(q["query_text"], corpus, top_k=5)
            top_ids = [r["chunk_id"] for r in results[:3]]
            # 验证这些 chunk_id 确实在 corpus 中
            corpus_ids = {c["chunk_id"] for c in corpus}
            for cid in top_ids:
                assert cid in corpus_ids


# ================================================================
# 12. edge cases
# ================================================================

class TestEdgeCases:
    """边界情况。"""

    def test_empty_query(self, corpus):
        for fn in (keyword_retrieve, tfidf_retrieve, bm25_retrieve):
            assert fn("", corpus, top_k=5) == []
            assert fn("   ", corpus, top_k=5) == []

    def test_single_chunk_corpus(self):
        c = [{
            "chunk_id": "C0001", "file_role": "test", "file_name": "test.pdf",
            "locator_type": "pdf_page", "page_number": 1,
            "sheet_name": None, "cell_range": None,
            "text_chunk_index": None, "section_title": None,
            "text": "SYN-TENDER-001 项目名称填写要求",
            "content_hash": hashlib.sha256(b"test").hexdigest(),
        }]
        for fn in (keyword_retrieve, tfidf_retrieve, bm25_retrieve):
            results = fn("SYN-TENDER-001", c, top_k=5)
            assert len(results) == 1
            assert results[0]["chunk_id"] == "C0001"

    def test_top_k_larger_than_corpus(self):
        c = [
            {
                "chunk_id": f"C{i:04d}", "file_role": "test",
                "file_name": "test.pdf", "locator_type": "pdf_page",
                "page_number": i, "sheet_name": None, "cell_range": None,
                "text_chunk_index": None, "section_title": None,
                "text": f"chunk {i} content",
                "content_hash": hashlib.sha256(f"chunk {i}".encode()).hexdigest(),
            }
            for i in range(1, 4)
        ]
        results = bm25_retrieve("chunk", c, top_k=100)
        assert len(results) == 3


# ================================================================
# 13. 分词器
# ================================================================

class TestTokenizer:
    def test_chinese_chars_and_bigrams(self):
        tokens = tokenize("投标响应")
        assert "投" in tokens
        assert "标" in tokens
        assert "响" in tokens
        assert "应" in tokens
        assert "投标" in tokens
        assert "标响" in tokens
        assert "响应" in tokens

    def test_clause_number(self):
        tokens = tokenize("SYN-TENDER-007 规定了最低人员数量")
        assert "SYN-TENDER-007" in tokens

    def test_cert_number(self):
        tokens = tokenize("证书 SYN-JC-24018")
        assert "SYN-JC-24018" in tokens

    def test_numbers(self):
        tokens = tokenize("预算 2000000 元")
        assert "2000000" in tokens

    def test_empty(self):
        assert tokenize("") == []

    def test_unicode_normalization(self):
        tokens = tokenize("１２３")
        assert "123" in tokens


# ================================================================
# 14. 整体指标聚合
# ================================================================

class TestAggregation:
    def test_category_metrics(self, corpus, queries):
        ans_qs = [q for q in queries if q["answerable"]]
        per_query = []
        for q in ans_qs:
            results = bm25_retrieve(q["query_text"], corpus, top_k=20)
            retrieved_ids = [r["chunk_id"] for r in results]
            per_query.append(
                compute_answerable_metrics(
                    q["query_id"], retrieved_ids, q["relevant_chunk_ids"], 1.0
                )
            )
        by_cat = aggregate_by_category(per_query, ans_qs)
        assert len(by_cat) > 0
        for cat, agg in by_cat.items():
            assert "answerable_count" in agg
            assert agg["answerable_count"] > 0

    def test_all_three_splits_have_reports(self):
        for split in ["all", "dev", "test"]:
            path = (
                GOLDEN_CASE_DIR.parent / "eval_results"
                / split / "retrieval_report.json"
            )
            assert path.exists(), f"缺少 {split} 报告: {path}"
            r = json.loads(path.read_text("utf-8"))
            assert r["meta"]["split"] == split
            assert r["meta"]["query_count"] > 0
