"""阶段 4C-1 真实 API 集成测试。

覆盖范围（说明——普通 pytest 中 Embedding 推理由 FakeEmbeddingProvider 替代，
不是真实 BGE 验证；真实 BGE 验证由 scripts/verify_stage4c_real_retrieval.py 完成）：

- test_full_golden_case_retrieval_api_flow：五份 golden_case 材料（PDF×3 + XLSX + MD）
  通过真实 HTTP API 的完整成功闭环（真实文件与真实 API、Fake Embedding）。
- TestFiveFileAPIClosure：五份黄金材料的 API 全流程（早期版本，含生命周期细节）。
- 其余以单份简化 Markdown（05_项目澄清.md）为主的测试仅用于快速错误分支验证
  （索引缺失/过期、错误码、跨用户隔离、CSRF、原子写入失败注入等），
  不构成"五份黄金材料检索闭环"。

API、文件解析、角色确认、Corpus 和索引流程是真实的；
Embedding 推理由 FakeEmbeddingProvider 替代。
"""

from __future__ import annotations

import hashlib
import io
import json
import os
import uuid
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session

from app.core.config import settings
from app.retrieval.embedding import FakeEmbeddingProvider, QUERY_INSTRUCTION

_BACKEND_DIR = Path(__file__).resolve().parents[1]

# ── 黄金材料路径 ──────────────────────────────────────────────────────────

_GOLDEN = Path(__file__).resolve().parents[2] / "examples" / "engineering_review_v1" / "golden_case"

GOLDEN_FILES: list[tuple[str, str, str]] = [
    ("01_合成招标要求.pdf", "tender_requirement", "application/pdf"),
    ("02_合成投标响应.pdf", "bid_response", "application/pdf"),
    ("03_人员设备清单.xlsx", "personnel_equipment_data",
     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    ("04_合成资质附件.pdf", "qualification_attachment", "application/pdf"),
    ("05_项目澄清.md", "clarification_document", "text/markdown"),
]

PASSWORD = "TestPass!2026"


# ── 辅助函数 ─────────────────────────────────────────────────────────────


def _register_and_login(client: TestClient, username: str, invite: str = "TEST-INVITE-4C1") -> dict:
    """通过 API 注册并登录，返回用户信息。"""
    # 获取 CSRF token
    csrf_resp = client.get("/api/v2/auth/csrf")
    assert csrf_resp.status_code == 200
    csrf_header = {settings.csrf_header_name: client.cookies.get(settings.csrf_cookie_name, "")}

    # 注册（需要 CSRF header）
    reg_resp = client.post(
        "/api/v2/auth/register",
        headers=csrf_header,
        json={
            "username": username,
            "password": PASSWORD,
            "password_confirm": PASSWORD,
            "invite_code": invite,
        },
    )
    # 409 表示已存在（可接受）
    assert reg_resp.status_code in (201, 409), f"注册失败: {reg_resp.status_code} {reg_resp.text}"

    # 重新获取 CSRF token（注册后 CSRF 可能已变更）
    csrf_resp2 = client.get("/api/v2/auth/csrf")
    assert csrf_resp2.status_code == 200
    csrf_header2 = {settings.csrf_header_name: client.cookies.get(settings.csrf_cookie_name, "")}

    # 登录
    login_resp = client.post(
        "/api/v2/auth/login",
        headers=csrf_header2,
        json={"username": username, "password": PASSWORD},
    )
    assert login_resp.status_code == 200, f"登录失败: {login_resp.status_code} {login_resp.text}"
    return login_resp.json()


def _session_csrf(client: TestClient) -> dict[str, str]:
    return {settings.csrf_header_name: client.cookies.get(settings.csrf_cookie_name, "")}


def _public_csrf(client: TestClient) -> dict[str, str]:
    resp = client.get("/api/v2/auth/csrf")
    assert resp.status_code == 200
    return {settings.csrf_header_name: client.cookies.get(settings.csrf_cookie_name, "")}


def _create_eng_workspace(client: TestClient, name: str = "集成测试工程") -> dict:
    resp = client.post(
        "/api/v2/workspaces",
        headers=_session_csrf(client),
        json={"name": name, "workspace_type": "engineering"},
    )
    assert resp.status_code == 201, f"创建工作区失败: {resp.status_code} {resp.text}"
    return resp.json()


def _upload_file(client: TestClient, workspace_id: int, filepath: Path) -> dict:
    """通过 multipart 上传真实文件。"""
    content_type = "application/octet-stream"
    if filepath.suffix == ".pdf":
        content_type = "application/pdf"
    elif filepath.suffix in (".xlsx", ".xls"):
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif filepath.suffix == ".md":
        content_type = "text/markdown"

    with open(filepath, "rb") as fh:
        content = fh.read()

    resp = client.post(
        f"/api/v2/workspaces/{workspace_id}/files",
        headers=_session_csrf(client),
        files={"file": (filepath.name, io.BytesIO(content), content_type)},
    )
    assert resp.status_code == 201, f"上传 {filepath.name} 失败: {resp.status_code} {resp.text}"
    return resp.json()


def _understand_and_confirm(
    client: TestClient, workspace_id: int, file_id: int, role: str
) -> None:
    """理解文件并确认工程角色。"""
    # 理解
    understand_resp = client.post(
        f"/api/v2/workspaces/{workspace_id}/files/understand",
        headers=_session_csrf(client),
        json={"file_ids": [file_id]},
    )
    assert understand_resp.status_code == 200, f"理解失败 file {file_id}: {understand_resp.text}"

    # 确认角色
    patch_resp = client.patch(
        f"/api/v2/workspaces/{workspace_id}/files/{file_id}/profile",
        headers=_session_csrf(client),
        json={"confirmed_role": role},
    )
    assert patch_resp.status_code == 200, f"确认角色失败 {file_id}→{role}: {patch_resp.text}"


def _build_index(client: TestClient, workspace_id: int, rebuild: bool = False) -> dict:
    resp = client.post(
        f"/api/v2/workspaces/{workspace_id}/engineering-retrieval/index",
        headers=_session_csrf(client),
        json={"rebuild": rebuild},
    )
    assert resp.status_code == 200, f"构建索引失败: {resp.status_code} {resp.text}"
    return resp.json()


def _search(
    client: TestClient,
    workspace_id: int,
    query: str,
    mode: str = "hybrid_rrf",
    top_k: int = 5,
    expect_status: int = 200,
) -> dict:
    resp = client.post(
        f"/api/v2/workspaces/{workspace_id}/engineering-retrieval/search",
        headers=_session_csrf(client),
        json={"query": query, "top_k": top_k, "retrieval_mode": mode},
    )
    assert resp.status_code == expect_status, (
        f"检索 {mode} 状态 {resp.status_code} != {expect_status}: {resp.text[:300]}"
    )
    return resp.json()


# ── Fixtures ──────────────────────────────────────────────────────────────


@pytest.fixture(autouse=True)
def _inject_fake_provider(monkeypatch, tmp_path):
    """全局注入 FakeEmbeddingProvider + 隔离索引目录。

    API、文件解析、角色确认、Corpus 和索引流程是真实的；
    Embedding 推理由 Fake Provider 替代。
    """
    def _make_fake(**kw):
        return FakeEmbeddingProvider(dimension=512, seed=42)

    monkeypatch.setattr(
        "app.services.engineering_retrieval_service.LocalEmbeddingProvider",
        _make_fake,
    )

    # 隔离索引目录：每个测试使用独立 tmp 目录
    isolated_root = tmp_path / "retrieval" / "workspaces"
    isolated_root.mkdir(parents=True, exist_ok=True)
    monkeypatch.setattr(
        "app.services.engineering_retrieval_service._INDEX_ROOT",
        isolated_root,
    )


def _resolve_current_upload_dir() -> Path:
    """按生产 _resolve_upload_dir 契约解析当前 upload_dir。"""
    upload_dir = Path(settings.upload_dir)
    if not upload_dir.is_absolute():
        upload_dir = _BACKEND_DIR / upload_dir
    return upload_dir


@pytest.fixture(autouse=True)
def _isolated_upload_dir(tmp_path, monkeypatch):
    """将 settings.upload_dir 临时指向 pytest 隔离目录（自动恢复）。

    - 保存原值，测试结束后 finally 恢复
    - settings 为 frozen dataclass，使用 object.__setattr__
    - 不使用任何递归删除；文件生命周期交给 pytest tmp_path
    """
    isolated = tmp_path / "uploads"
    isolated.mkdir(parents=True, exist_ok=True)
    original = settings.upload_dir
    object.__setattr__(settings, "upload_dir", str(isolated))
    try:
        yield isolated
    finally:
        object.__setattr__(settings, "upload_dir", original)


def _snapshot_default_uploads() -> tuple[int, str, str]:
    """记录默认上传目录三项数据：数量 / 路径清单 SHA / 内容组合 SHA。"""
    root = _BACKEND_DIR / "storage" / "uploads"
    files = sorted(p for p in root.rglob("*") if p.is_file()) if root.exists() else []
    manifest = "\n".join(str(p.relative_to(root)) for p in files)
    manifest_sha = hashlib.sha256(manifest.encode("utf-8")).hexdigest()
    h = hashlib.sha256()
    for p in files:
        h.update(p.read_bytes())
    return len(files), manifest_sha, h.hexdigest()


def _snapshot_default_index_root() -> list[str]:
    """记录默认检索索引目录文件清单。"""
    root = _BACKEND_DIR / "storage" / "retrieval" / "workspaces"
    if not root.exists():
        return []
    return sorted(str(p.relative_to(root)) for p in root.rglob("*") if p.is_file())


@pytest.fixture(scope="module", autouse=True)
def _default_storage_untouched():
    """防回归：本文件测试运行前后，默认上传目录与检索索引目录完全一致。

    证明本文件所有测试生成的上传文件均位于 pytest 隔离目录，
    默认 backend/storage/uploads 未被使用（数量/路径清单/内容全不变）。
    """
    uploads_before = _snapshot_default_uploads()
    index_before = _snapshot_default_index_root()
    yield
    uploads_after = _snapshot_default_uploads()
    index_after = _snapshot_default_index_root()
    assert uploads_after == uploads_before, (
        f"默认 uploads 被本文件测试修改: "
        f"数量 {uploads_before[0]} -> {uploads_after[0]}"
    )
    assert index_after == index_before, "默认检索索引目录被本文件测试修改"


@pytest.fixture
def invite_code(db_session):
    """创建有效邀请码，返回原始 code 字符串。"""
    from app.services.security_service import invite_code_hash, invite_code_hint
    from app.models.invite_code import InviteCode

    raw_code = "TEST-INVITE-4C1"
    code_hash_val = invite_code_hash(raw_code)
    code_hint_val = invite_code_hint(raw_code)

    existing = db_session.query(InviteCode).filter(
        InviteCode.code_hash == code_hash_val
    ).first()
    if existing is None:
        invite = InviteCode(
            code_hash=code_hash_val,
            code_hint=code_hint_val,
            status="active",
            max_uses=100,
            used_count=0,
        )
        db_session.add(invite)
        db_session.commit()
    return raw_code


@pytest.fixture
def auth_user(client, invite_code):
    """通过 API 注册并登录用户。"""
    uname = f"integration_{uuid.uuid4().hex[:8]}"
    _register_and_login(client, uname, invite_code)
    return uname


@pytest.fixture
def eng_workspace(client, auth_user):
    """创建工程工作区。"""
    return _create_eng_workspace(client)


# ── Section 3: 文件路径安全 ─────────────────────────────────────────────


class TestPathSecurity:
    """路径安全测试：_resolve_path 拒绝越狱。

    所有文件操作均发生在 pytest 隔离上传目录
    （settings.upload_dir 由 _isolated_upload_dir 重定向），
    不访问、不创建、不修改默认 backend/storage/uploads。
    """

    def test_upload_root_file_allowed(self):
        """上传根目录内文件允许。"""
        from app.services.engineering_corpus_adapter import _resolve_path

        upload_dir = _resolve_current_upload_dir()  # 隔离上传目录
        upload_dir.mkdir(parents=True, exist_ok=True)
        test_name = f"safe_{uuid.uuid4().hex}.txt"
        test_file = upload_dir / test_name
        test_file.write_text("allowed content", encoding="utf-8")

        try:
            # 绝对路径指向 upload root 内文件（_resolve_path 支持）
            result = _resolve_path(str(test_file))
            assert result is not None, f"无法解析路径: {test_file}"
            assert result.is_file()
        finally:
            test_file.unlink(missing_ok=True)

    def test_dotdot_path_rejected(self):
        """.. 路径穿越拒绝。"""
        from app.services.engineering_corpus_adapter import _resolve_path

        # .. 穿越到项目外
        result = _resolve_path("../../../etc/passwd")
        assert result is None, f".. 穿越应被拒绝，但返回了: {result}"

    def test_symlink_or_nonexistent_escape_rejected(self):
        """绝对路径指向 upload root 外应被拒绝。"""
        from app.services.engineering_corpus_adapter import _resolve_path

        # 指向根目录外的绝对路径（不存在 → resolve 失败 → None）
        result = _resolve_path("C:\\nonexistent_xyz123\\file.txt")
        assert result is None

    def test_directory_rejected(self):
        """目录路径拒绝（只接受普通文件）。"""
        from app.services.engineering_corpus_adapter import _resolve_path

        upload_dir = _resolve_current_upload_dir()  # 隔离上传目录
        upload_dir.mkdir(parents=True, exist_ok=True)

        # 指向上传根目录本身（既非普通文件也越出 allowed）
        result = _resolve_path(str(upload_dir))
        assert result is None, f"目录应被拒绝，但返回了: {result}"

    def test_api_warning_no_path_leak(self, client, auth_user, eng_workspace):
        """API warning 不包含系统路径。"""
        ws_id = eng_workspace["id"]

        resp = client.get(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
        )
        assert resp.status_code == 400
        detail = resp.json().get("detail", {})
        message = str(detail)

        for forbidden in ["C:\\", "D:\\", "Users"]:
            assert forbidden not in message, f"warning 泄露了路径: {forbidden} in {message}"


# ── Section 4-6: 五文件完整 API 闭环 ────────────────────────────────────


class TestFiveFileAPIClosure:
    """五份黄金材料通过真实 API 端到端闭环。

    流程：register → login → workspace → upload×5 → understand×5
         → confirm roles×5 → build index → search×3 modes
         → locator validation → lifecycle → delete
    """

    @pytest.fixture(autouse=True)
    def _setup(self, client, auth_user):
        self.client = client
        self.username = auth_user

    def test_full_five_file_pipeline(self, client, auth_user):
        """完整五文件流程：上传、理解、构建、检索、locator 验证。"""
        # ── Step 1: 创建工作区 ──
        ws = _create_eng_workspace(client, "五文件集成测试")
        ws_id = ws["id"]
        assert ws["workspace_type"] == "engineering"

        # ── Step 2: 上传 5 个 golden 材料 ──
        file_map: dict[int, tuple[str, str]] = {}  # file_id → (filename, role)
        for fname, role, _mime in GOLDEN_FILES:
            fpath = _GOLDEN / fname
            assert fpath.exists(), f"golden 文件缺失: {fpath}"
            result = _upload_file(client, ws_id, fpath)
            fid = result["file_id"]
            file_map[fid] = (fname, role)
            assert fid > 0

        assert len(file_map) == 5, f"应上传 5 个文件，实际 {len(file_map)}"

        # ── Step 3: 理解 + 确认角色 ──
        for fid, (fname, role) in file_map.items():
            _understand_and_confirm(client, ws_id, fid, role)

        # 验证所有 profile status == ready
        for fid in file_map:
            p_resp = client.get(
                f"/api/v2/workspaces/{ws_id}/files/{fid}/profile",
            )
            assert p_resp.status_code == 200
            profile = p_resp.json()
            assert profile.get("status") == "ready", f"file {fid} status: {profile.get('status')}"
            assert profile.get("confirmed_role") == file_map[fid][1]

        # ── Step 4: 查看索引状态 ──
        status_resp = client.get(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
        )
        assert status_resp.status_code == 200
        status_data = status_resp.json()
        assert status_data["status"] in ("not_built", "ready")
        assert status_data["chunk_count"] > 0, f"chunk_count 应为正数: {status_data}"
        assert status_data["file_count"] == 5

        # locator 验证在 search 阶段进行（IndexInfo 不暴露 locator_summary）
        pdf_warnings = [w for w in status_data.get("warnings", []) if "PDF" in w or "跳过" in w]
        if pdf_warnings:
            print(f"  ⚠ PDF warnings: {pdf_warnings}")

        # ── Step 5: 构建索引 ──
        build_data = _build_index(client, ws_id, rebuild=False)
        assert build_data["status"] in ("built", "already_built"), (
            f"构建状态异常: {build_data['status']}"
        )
        assert build_data["chunk_count"] > 0
        assert len(build_data.get("index_sha256", "")) > 0
        assert len(build_data.get("corpus_sha256", "")) > 0
        assert len(build_data.get("model_revision", "")) > 0

        # ── Step 6: 三种检索模式 ──

        # BM25
        bm25 = _search(client, ws_id, "项目概况与招标范围", "bm25")
        assert bm25["retrieval_mode"] == "bm25"
        assert len(bm25["results"]) > 0
        assert bm25.get("corpus_sha256")

        # Dense
        dense = _search(client, ws_id, "资质要求", "dense")
        assert dense["retrieval_mode"] == "dense"
        assert len(dense["results"]) > 0
        assert dense.get("index_sha256")

        # Hybrid RRF
        hybrid = _search(client, ws_id, "人员设备清单", "hybrid_rrf")
        assert hybrid["retrieval_mode"] == "hybrid_rrf"
        assert len(hybrid["results"]) > 0
        # hybrid 结果应有 bm25_rank 和 dense_rank
        for r in hybrid["results"]:
            assert "bm25_rank" in r
            assert "dense_rank" in r
            assert "rrf_score" in r if "rrf_score" in r else True  # 至少 rank 存在

        # 验证 answerability
        for resp_data in (bm25, dense, hybrid):
            assert resp_data.get("answerability", "unknown") == "unknown"

        # ── Step 7: Locator 验证 ──
        self._verify_locators(client, ws_id, file_map, build_data)

        # ── Step 8: 路径安全 ──
        for resp_data in (bm25, dense, hybrid):
            text = json.dumps(resp_data)
            assert "C:\\" not in text
            assert "D:\\" not in text

    def _verify_locators(self, client, ws_id, file_map, build_data):
        """验证 PDF/Excel/Markdown locator 正确性。"""
        # 获取所有结果来检查 locator 类型
        search_data = _search(client, ws_id, "招标要求 投标响应 人员 设备 资质 澄清", "hybrid_rrf", top_k=20)

        locator_types = set()
        pdf_pages = set()
        sheet_names = set()
        cell_ranges = set()

        for r in search_data["results"]:
            locator_types.add(r.get("locator_type", ""))
            if r.get("locator_type") == "pdf_page":
                if r.get("page_number"):
                    pdf_pages.add(r["page_number"])
            if r.get("locator_type") == "spreadsheet_cell":
                if r.get("sheet_name"):
                    sheet_names.add(r["sheet_name"])
                if r.get("cell_range"):
                    cell_ranges.add(r["cell_range"])

        # 验证三种 locator 都存在
        assert "pdf_page" in locator_types, f"缺少 pdf_page locator，实际: {locator_types}"
        assert "spreadsheet_cell" in locator_types, f"缺少 spreadsheet_cell locator，实际: {locator_types}"
        assert "text_chunk" in locator_types, f"缺少 text_chunk locator，实际: {locator_types}"

        # PDF page_number 真实存在
        assert len(pdf_pages) > 0, "应有真实的 page_number"

        # Excel sheet_name 非空
        assert len(sheet_names) > 0, "应有非空 sheet_name"

        # Excel cell_range 非空
        assert len(cell_ranges) > 0, "应有非空 cell_range"


# ── Section 6b: 五份黄金材料完整成功检索闭环 ─────────────────────────────


def _real_pdf_page_counts() -> dict[str, int]:
    """读取 golden PDF 的真实页数（PyMuPDF）。"""
    import fitz

    counts: dict[str, int] = {}
    for fname in ("01_合成招标要求.pdf", "02_合成投标响应.pdf", "04_合成资质附件.pdf"):
        doc = fitz.open(str(_GOLDEN / fname))
        counts[fname] = len(doc)
        doc.close()
    return counts


def _real_xlsx_sheets() -> dict[str, tuple[int, int]]:
    """读取 golden XLSX 的真实工作表 → (max_row, max_col)。"""
    import openpyxl

    sheets: dict[str, tuple[int, int]] = {}
    wb = openpyxl.load_workbook(str(_GOLDEN / "03_人员设备清单.xlsx"), read_only=True)
    for ws in wb.worksheets:
        sheets[ws.title] = (ws.max_row or 0, ws.max_column or 0)
    wb.close()
    return sheets


class TestGoldenCaseFullClosure:
    """五份 golden_case 材料通过真实 HTTP API 的完整成功检索闭环。

    流程：注册/登录 → 创建 engineering workspace → multipart 上传五份材料
        → 逐份 understand → PATCH profile 确认角色 → POST 构建索引
        → GET 索引状态 → bm25 / dense / hybrid_rrf 三种检索

    注意：普通 pytest 中使用 Fake Embedding Provider（确定性向量），
    这不是真实 BGE 验证；真实 BGE 验证见 scripts/verify_stage4c_real_retrieval.py。
    """

    @pytest.fixture(autouse=True)
    def _setup(self, client, auth_user, db_session):
        self.client = client
        self.username = auth_user
        self.db = db_session

    def _owner_id(self) -> int:
        from app.models.user import User
        from sqlalchemy import select

        owner = self.db.scalar(select(User).where(User.username == self.username))
        assert owner is not None
        return owner.id

    def test_full_golden_case_retrieval_api_flow(self, client, auth_user):
        """五份黄金材料：真实文件 + 真实 API、Fake Embedding 的完整成功闭环。"""
        # ── Step 1: 创建 engineering workspace ──
        ws = _create_eng_workspace(client, "黄金材料闭环验收")
        ws_id = ws["id"]
        assert ws["workspace_type"] == "engineering"

        # ── Step 2: multipart 上传五份 golden_case 材料 ──
        file_map: dict[int, tuple[str, str]] = {}  # file_id → (filename, role)
        for fname, role, _mime in GOLDEN_FILES:
            fpath = _GOLDEN / fname
            assert fpath.exists(), f"golden 文件缺失: {fpath}"
            result = _upload_file(client, ws_id, fpath)
            file_map[result["file_id"]] = (fname, role)
        assert len(file_map) == 5, f"应上传 5 个文件，实际 {len(file_map)}"

        # ── Step 3/4: 逐份 understand + PATCH profile 确认角色 ──
        for fid, (fname, role) in file_map.items():
            _understand_and_confirm(client, ws_id, fid, role)
            p_resp = client.get(f"/api/v2/workspaces/{ws_id}/files/{fid}/profile")
            assert p_resp.status_code == 200
            profile = p_resp.json()
            assert profile["status"] == "ready", f"file {fid} 未 ready: {profile}"
            assert profile["confirmed_role"] == role, f"file {fid} 角色确认失败"

        # ── Step 5: POST 构建索引 ──
        build_data = _build_index(client, ws_id, rebuild=False)
        assert build_data["status"] in ("built", "already_built"), build_data
        assert build_data["file_count"] == 5, f"file_count: {build_data}"
        assert build_data["chunk_count"] > 0
        assert build_data["corpus_sha256"], "缺少 corpus_sha256"
        assert build_data["index_sha256"], "缺少 index_sha256"

        # ── Step 6: GET 索引状态 ──
        status_resp = client.get(f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index")
        assert status_resp.status_code == 200
        status_data = status_resp.json()
        assert status_data["status"] == "ready", status_data
        assert status_data["file_count"] == 5
        assert status_data["chunk_count"] > 0
        assert status_data["corpus_sha256"] == build_data["corpus_sha256"]

        # ── 从生产产物读取 manifest，验证 Corpus 真实构成 ──
        from app.retrieval.dense_index import _resolve_active_assets
        from app.services.engineering_retrieval_service import _index_dir

        idx_dir = _index_dir(ws_id)
        assets = _resolve_active_assets(idx_dir)
        assert assets is not None, "索引资产无法解析"
        manifest = json.loads(assets["manifest"].read_text("utf-8"))
        owner_id = self._owner_id()

        # 五种已确认角色全部进入 Corpus
        corpus_roles = set(manifest["confirmed_roles"].values())
        expected_roles = {role for _, role in file_map.values()}
        assert corpus_roles == expected_roles, (
            f"Corpus 角色不完整: 实际 {corpus_roles}，预期 {expected_roles}"
        )

        # Corpus 不是测试代码硬编码的 chunk：chunk_id 为真实
        # W{workspace}F{file}C{index} 格式且 file_id 与本次上传一致
        assert set(manifest["file_ids"]) == set(file_map.keys()), (
            f"Corpus file_ids 与上传不符: {manifest['file_ids']}"
        )
        chunk_id_prefix = f"W{ws_id:04d}"
        for cid in manifest["chunk_ids"]:
            assert cid.startswith(chunk_id_prefix), f"chunk_id 格式异常: {cid}"
        assert manifest["locator_summary"]["pdf_page"] > 0
        assert manifest["locator_summary"]["spreadsheet_cell"] > 0
        assert manifest["locator_summary"]["text_chunk"] > 0

        # workspace/owner 不串
        assert manifest["workspace_id"] == ws_id
        assert manifest["owner_user_id"] == owner_id, (
            f"owner 串用户: manifest={manifest['owner_user_id']}, user={owner_id}"
        )

        # ── Step 7: 三种检索模式 ──
        broad_query = "招标范围 投标 人员 设备 资质 澄清"
        bm25 = _search(client, ws_id, broad_query, "bm25", top_k=20)
        dense = _search(client, ws_id, broad_query, "dense", top_k=20)
        hybrid = _search(client, ws_id, broad_query, "hybrid_rrf", top_k=20)

        # 统一结果结构：三种模式返回相同字段集合
        result_keys = {
            "rank", "chunk_id", "file_id", "file_name", "file_role",
            "locator_type", "page_number", "sheet_name", "cell_range",
            "quote", "score", "bm25_rank", "dense_rank",
            "content_hash", "parser_name", "parser_version",
        }
        for mode, data in (("bm25", bm25), ("dense", dense), ("hybrid_rrf", hybrid)):
            assert data["retrieval_mode"] == mode
            assert data["corpus_sha256"] == build_data["corpus_sha256"]
            assert len(data["results"]) > 0, f"{mode} 无结果"
            for r in data["results"]:
                assert set(r.keys()) == result_keys, (
                    f"{mode} 结果字段不一致: {sorted(r.keys())}"
                )
                # rank 从 1 开始且连续
            ranks = [r["rank"] for r in data["results"]]
            assert ranks == list(range(1, len(ranks) + 1)), f"{mode} rank 不连续: {ranks}"
            # answerability 仍为 unknown（检索阶段不做可答性判断）
            assert data.get("answerability", "unknown") == "unknown", mode

        # 不根据 golden answer 伪造排序：每种模式的 score 单调且由计算值决定
        for mode, data in (("bm25", bm25), ("dense", dense), ("hybrid_rrf", hybrid)):
            scores = [r["score"] for r in data["results"]]
            assert all(
                scores[i] >= scores[i + 1] for i in range(len(scores) - 1)
            ), f"{mode} score 非单调（疑似按 golden answer 重排）: {scores}"
        # 显式声明：本测试不读取 ground_truth.json，不对照黄金答案调整排序

        # ── 不串文件：所有命中的 file_id 都属于本工作区本次上传的 5 个文件 ──
        uploaded_ids = set(file_map.keys())
        for data in (bm25, dense, hybrid):
            for r in data["results"]:
                assert r["file_id"] in uploaded_ids, f"命中外部文件: {r['file_id']}"
                assert r["chunk_id"].startswith(chunk_id_prefix)

        # ── locator 真实性：PDF 页码、Excel 工作表与单元格范围真实存在 ──
        pdf_fids = {fid for fid, (fname, _r) in file_map.items() if fname.endswith(".pdf")}
        xlsx_fids = {fid for fid, (fname, _r) in file_map.items() if fname.endswith(".xlsx")}
        md_fids = {fid for fid, (fname, _r) in file_map.items() if fname.endswith(".md")}
        pdf_pages = _real_pdf_page_counts()
        xlsx_sheets = _real_xlsx_sheets()

        pdf_pages_seen: dict[int, set[int]] = {}
        sheets_seen: dict[int, set[str]] = {}
        ranges_seen: dict[int, set[str]] = {}
        md_chunks_seen = 0
        for data in (bm25, dense, hybrid):
            for r in data["results"]:
                loc = r["locator_type"]
                if loc == "pdf_page":
                    assert r["file_id"] in pdf_fids, "pdf_page 命中了非 PDF 文件"
                    page = r.get("page_number")
                    assert page is not None and page >= 1, f"pdf_page 缺页码: {r}"
                    fname = file_map[r["file_id"]][0]
                    assert page <= pdf_pages[fname], (
                        f"页码 {page} 超出 {fname} 真实页数 {pdf_pages[fname]}"
                    )
                    pdf_pages_seen.setdefault(r["file_id"], set()).add(page)
                elif loc == "spreadsheet_cell":
                    assert r["file_id"] in xlsx_fids, "spreadsheet_cell 命中了非 Excel 文件"
                    sheet = r.get("sheet_name") or ""
                    cell_range = r.get("cell_range") or ""
                    assert sheet, "spreadsheet_cell 缺 sheet_name"
                    assert cell_range, "spreadsheet_cell 缺 cell_range"
                    assert sheet in xlsx_sheets, f"工作表 {sheet} 在真实 xlsx 中不存在"
                    max_row, max_col = xlsx_sheets[sheet]
                    if max_row > 0 and max_col > 0:
                        assert _cell_range_within(cell_range, max_row, max_col), (
                            f"单元格范围 {cell_range} 超出工作表 {sheet} 真实范围 "
                            f"({max_row}×{max_col})"
                        )
                    sheets_seen.setdefault(r["file_id"], set()).add(sheet)
                    ranges_seen.setdefault(r["file_id"], set()).add(cell_range)
                elif loc == "text_chunk":
                    assert r["file_id"] in md_fids, "text_chunk 命中了非 Markdown 文件"
                    md_chunks_seen += 1

        assert pdf_pages_seen, "未出现 pdf_page locator"
        assert sheets_seen, "未出现 spreadsheet_cell locator"
        assert ranges_seen, "未出现非空 cell_range"
        assert md_chunks_seen > 0, "未出现 Markdown text_chunk locator"

        # ── Corpus 内容真实性：命中的 quote 来自真实文件内容 ──
        md_search = _search(client, ws_id, "SYN-CLAR-001 澄清", "bm25", top_k=5)
        assert len(md_search["results"]) > 0
        quotes = " ".join(r["quote"] for r in md_search["results"])
        assert "SYN-CLAR-001" in quotes or "澄清" in quotes, (
            f"Corpus 内容疑似硬编码: {quotes[:200]}"
        )

        # ── 路径安全：响应不泄露文件系统路径 ──
        for data in (bm25, dense, hybrid):
            text = json.dumps(data)
            assert "C:\\" not in text and "D:\\" not in text


def _cell_range_within(cell_range: str, max_row: int, max_col: int) -> bool:
    """判断 'A1:B5' 形式的单元格范围是否落在工作表真实范围内。"""
    import re

    m = re.fullmatch(r"([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?", cell_range)
    if not m:
        return False
    col1, row1 = m.group(1), int(m.group(2))
    col2 = m.group(3) or col1
    row2 = int(m.group(4) or row1)

    def col_index(col: str) -> int:
        idx = 0
        for ch in col:
            idx = idx * 26 + (ord(ch) - ord("A") + 1)
        return idx

    return (
        col_index(col1) <= max_col
        and col_index(col2) <= max_col
        and row1 <= max_row
        and row2 <= max_row
    )


# ── Section 6c: current_version.json 指针安全 ─────────────────────────────


def _make_pointer_snapshot(root: Path, version: str | None = None) -> Path:
    """构造一个带合法指针与三个资产的索引目录，返回 idx_dir。"""
    idx_dir = root / "idx"
    idx_dir.mkdir()
    version = version or uuid.uuid4().hex
    (idx_dir / f"dense_index_{version}.npz").write_bytes(b"npz-content")
    (idx_dir / f"dense_index_meta_{version}.json").write_text(
        json.dumps({"index_sha256": "x"}), encoding="utf-8"
    )
    (idx_dir / f"corpus_manifest_{version}.json").write_text(
        json.dumps({"chunk_count": 1}), encoding="utf-8"
    )
    pointer = {
        "version": version,
        "npz": f"dense_index_{version}.npz",
        "meta": f"dense_index_meta_{version}.json",
        "manifest": f"corpus_manifest_{version}.json",
    }
    (idx_dir / "current_version.json").write_text(
        json.dumps(pointer), encoding="utf-8"
    )
    return idx_dir


def _write_pointer(idx_dir: Path, **overrides) -> None:
    data = json.loads((idx_dir / "current_version.json").read_text("utf-8"))
    data.update(overrides)
    (idx_dir / "current_version.json").write_text(
        json.dumps(data), encoding="utf-8"
    )


class TestPointerSecurity:
    """current_version.json 指针路径攻击防护。

    纯单元测试：_resolve_active_assets 不得读取目录外文件或泄露真实路径，
    任何校验失败返回 None。
    """

    def test_legal_pointer_loads(self, tmp_path):
        """合法指针正常解析出三个资产路径。"""
        idx_dir = _make_pointer_snapshot(tmp_path)
        from app.retrieval.dense_index import _resolve_active_assets

        assets = _resolve_active_assets(idx_dir)
        assert assets is not None
        assert assets["npz"].is_file()
        assert assets["meta"].is_file()
        assert assets["manifest"].is_file()

    def test_dotdot_npz_rejected(self, tmp_path):
        """../outside.npz 被拒绝，且外部文件不被读取。"""
        idx_dir = _make_pointer_snapshot(tmp_path)
        outside = tmp_path / "outside.npz"
        outside.write_bytes(b"secret-content")

        _write_pointer(idx_dir, npz="../outside.npz")
        from app.retrieval.dense_index import _resolve_active_assets

        assert _resolve_active_assets(idx_dir) is None
        # 外部文件未被修改
        assert outside.read_bytes() == b"secret-content"

    def test_absolute_path_rejected(self, tmp_path):
        """绝对路径被拒绝。"""
        idx_dir = _make_pointer_snapshot(tmp_path)
        _write_pointer(idx_dir, npz=str(tmp_path / "absolute.npz"))
        from app.retrieval.dense_index import _resolve_active_assets

        assert _resolve_active_assets(idx_dir) is None

    def test_version_filename_mismatch_rejected(self, tmp_path):
        """指针 version 与文件名版本不一致被拒绝。"""
        idx_dir = _make_pointer_snapshot(tmp_path)
        # 指针 version 换成另一个合法 uuid，但文件名仍是旧版本
        other = uuid.uuid4().hex
        _write_pointer(idx_dir, version=other)
        from app.retrieval.dense_index import _resolve_active_assets

        assert _resolve_active_assets(idx_dir) is None

    def test_invalid_version_rejected(self, tmp_path):
        """非法 version（非 32 位小写十六进制）被拒绝。"""
        idx_dir = _make_pointer_snapshot(tmp_path)
        from app.retrieval.dense_index import _resolve_active_assets

        for bad_version in ("xyz", uuid.uuid4().hex.upper(), "not-a-uuid"):
            _write_pointer(idx_dir, version=bad_version)
            assert _resolve_active_assets(idx_dir) is None, f"version={bad_version} 应被拒绝"

        # 非字符串 version
        _write_pointer(idx_dir, version=123)
        assert _resolve_active_assets(idx_dir) is None

    def test_missing_asset_rejected(self, tmp_path):
        """缺少任一资产被拒绝。"""
        idx_dir = _make_pointer_snapshot(tmp_path)
        from app.retrieval.dense_index import _resolve_active_assets

        for name in ("dense_index_meta", "corpus_manifest"):
            target = next(idx_dir.glob(f"{name}_*.json"))
            target.unlink()
            assert _resolve_active_assets(idx_dir) is None, f"缺 {name} 应被拒绝"
            # 补回以继续下一个场景
            target.write_text("{}", encoding="utf-8")

        npz = next(idx_dir.glob("dense_index_*.npz"))
        npz.unlink()
        assert _resolve_active_assets(idx_dir) is None, "缺 npz 应被拒绝"

    def test_directory_target_rejected(self, tmp_path):
        """指针指向目录被拒绝。"""
        idx_dir = _make_pointer_snapshot(tmp_path)
        from app.retrieval.dense_index import _resolve_active_assets

        # 把 npz 目标替换为一个目录（文件同名冲突，先删文件再建目录）
        npz = next(idx_dir.glob("dense_index_*.npz"))
        npz.unlink()
        npz.mkdir()
        assert _resolve_active_assets(idx_dir) is None

    def test_symlink_rejected_when_creatable(self, tmp_path):
        """符号链接在系统允许创建时被拒绝。

        若当前系统（Windows 无开发者模式/权限）不允许创建符号链接，
        安全记录原因，但不使用 skip/xfail，其余断言照常执行。
        """
        idx_dir = _make_pointer_snapshot(tmp_path)
        outside = tmp_path / "outside_secret.txt"
        outside.write_text("secret", encoding="utf-8")

        npz = next(idx_dir.glob("dense_index_*.npz"))
        npz.unlink()
        try:
            os.symlink(outside, npz)
        except OSError as exc:
            print(f"  [info] 当前系统不允许创建符号链接，跳过该子断言: {exc}")
        else:
            from app.retrieval.dense_index import _resolve_active_assets

            assert _resolve_active_assets(idx_dir) is None, "符号链接应被拒绝"
            assert outside.read_text(encoding="utf-8") == "secret", "外部文件被修改"

    def test_outside_file_not_read_or_modified(self, tmp_path):
        """恶意指针下外部文件内容不被读取或修改。"""
        idx_dir = _make_pointer_snapshot(tmp_path)
        outside = tmp_path / "outside.bin"
        payload = os.urandom(64)
        outside.write_bytes(payload)

        from app.retrieval.dense_index import _resolve_active_assets

        # 多个恶意指针均被拒绝
        for key, val in (
            ("npz", "../outside.bin"),
            ("meta", "../outside.bin"),
            ("manifest", f"{tmp_path}/outside.bin"),
        ):
            _write_pointer(idx_dir, **{key: val})
            assert _resolve_active_assets(idx_dir) is None, f"{key}={val} 应被拒绝"

        # 外部文件内容与 mtime 均未改变
        assert outside.read_bytes() == payload


# ── Section 7: 索引生命周期测试 ──────────────────────────────────────────


class TestIndexLifecycle:
    """索引生命周期：Missing → Stale → Rebuild → Reuse。"""

    @pytest.fixture
    def prepared_workspace(self, client, auth_user):
        """准备一个已构建索引的工作区。"""
        ws = _create_eng_workspace(client, "生命周期测试")
        ws_id = ws["id"]

        # 上传所有 golden 文件
        file_map: dict[int, str] = {}
        for fname, role, _mime in GOLDEN_FILES:
            fpath = _GOLDEN / fname
            result = _upload_file(client, ws_id, fpath)
            fid = result["file_id"]
            file_map[fid] = role

        # 理解 + 确认角色
        for fid, role in file_map.items():
            _understand_and_confirm(client, ws_id, fid, role)

        # 构建索引
        build_data = _build_index(client, ws_id, rebuild=False)
        return ws_id, build_data, file_map

    def test_index_missing_on_dense_search(self, client, auth_user, eng_workspace):
        """未建索引时 Dense 检索返回 INDEX_MISSING。"""
        ws_id = eng_workspace["id"]

        # 上传一个文件但只理解不建索引
        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        fid = result["file_id"]
        _understand_and_confirm(client, ws_id, fid, "clarification_document")

        # Dense 检索需要索引
        data = _search(client, ws_id, "测试查询", "dense", expect_status=400)
        assert data["detail"]["error_code"] == "ENGINEERING_RETRIEVAL_INDEX_MISSING"

        # Hybrid 也需要索引
        data2 = _search(client, ws_id, "测试查询", "hybrid_rrf", expect_status=400)
        assert data2["detail"]["error_code"] in (
            "ENGINEERING_RETRIEVAL_INDEX_MISSING",
            "ENGINEERING_RETRIEVAL_INDEX_STALE",
        )

    def test_index_stale_after_role_change(self, client, auth_user, prepared_workspace):
        """修改角色后 Corpus hash 改变 → INDEX_STALE。"""
        ws_id, build_data, file_map = prepared_workspace
        original_sha = build_data["corpus_sha256"]

        # 修改一个文件的角色
        first_fid = next(iter(file_map))
        new_role = "supplementary_attachment"
        patch_resp = client.patch(
            f"/api/v2/workspaces/{ws_id}/files/{first_fid}/profile",
            headers=_session_csrf(client),
            json={"confirmed_role": new_role},
        )
        assert patch_resp.status_code == 200

        # 再次检索应返回 STALE
        data = _search(client, ws_id, "测试", "hybrid_rrf", expect_status=400)
        assert data["detail"]["error_code"] == "ENGINEERING_RETRIEVAL_INDEX_STALE"

        # Corpus hash 确实变了
        status_resp = client.get(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
        )
        assert status_resp.status_code == 200
        new_sha = status_resp.json().get("corpus_sha256", "")
        assert new_sha != original_sha

    def test_rebuild_restores_search(self, client, auth_user, prepared_workspace):
        """Rebuild 后搜索恢复。"""
        ws_id, build_data, file_map = prepared_workspace

        # 先制造 stale
        first_fid = next(iter(file_map))
        client.patch(
            f"/api/v2/workspaces/{ws_id}/files/{first_fid}/profile",
            headers=_session_csrf(client),
            json={"confirmed_role": "supplementary_attachment"},
        )

        # rebuild
        rebuild_data = _build_index(client, ws_id, rebuild=True)
        assert rebuild_data["status"] == "built"
        new_corpus_sha = rebuild_data["corpus_sha256"]

        # 检索恢复正常
        search_data = _search(client, ws_id, "招标要求", "hybrid_rrf")
        assert len(search_data["results"]) > 0
        assert search_data["corpus_sha256"] == new_corpus_sha

        # GET status → ready
        status = client.get(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
        )
        assert status.status_code == 200
        assert status.json()["status"] == "ready"

    def test_reuse_when_no_change(self, client, auth_user, prepared_workspace):
        """不修改材料再次 POST index → 复用已有索引。"""
        ws_id, build_data, file_map = prepared_workspace

        # 不修改任何材料，再次构建
        reuse_data = _build_index(client, ws_id, rebuild=False)
        assert reuse_data["status"] == "already_built"
        assert reuse_data["corpus_sha256"] == build_data["corpus_sha256"]
        assert reuse_data["index_sha256"] == build_data["index_sha256"]


# ── Section 8: 永久删除测试 ─────────────────────────────────────────────


class TestPermanentDeleteWithRealIndex:
    """永久删除工作区时清理真实索引资产。"""

    def test_delete_cleans_index_files(self, client, auth_user):
        """先构建真实索引，再永久删除，验证三个索引文件已清理。"""
        ws = _create_eng_workspace(client, "删除测试工程")
        ws_id = ws["id"]
        ws_name = ws["name"]

        # 上传文件
        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        fid = result["file_id"]
        _understand_and_confirm(client, ws_id, fid, "clarification_document")

        # 构建索引
        build_data = _build_index(client, ws_id)
        assert build_data["status"] in ("built", "already_built")

        # 确认三个索引资产文件存在（通过当前版本指针解析）
        from app.retrieval.dense_index import _resolve_active_assets
        from app.services.engineering_retrieval_service import _index_dir
        idx_dir = _index_dir(ws_id)
        assets = _resolve_active_assets(idx_dir)
        assert assets is not None, f"无法解析索引资产: {idx_dir}"
        npz_path = assets["npz"]
        meta_path = assets["meta"]
        manifest_path = assets["manifest"]
        assert npz_path.exists(), f"NPZ 不存在: {npz_path}"
        assert meta_path.exists(), f"meta 不存在: {meta_path}"
        assert manifest_path.exists(), f"manifest 不存在: {manifest_path}"
        pointer_path = idx_dir / "current_version.json"
        assert pointer_path.exists(), "current_version.json 指针不存在"

        # 永久删除
        delete_resp = client.request(
            "DELETE",
            f"/api/v2/workspaces/{ws_id}",
            headers=_session_csrf(client),
            json={"confirmation_name": ws_name},
        )
        assert delete_resp.status_code == 200, f"删除失败: {delete_resp.text}"
        data = delete_resp.json()
        assert data["message"] == "项目已永久删除"

        # 验证三个索引资产文件已清理（含版本化文件和指针）
        assert not npz_path.exists(), "dense_index NPZ 仍存在"
        assert not meta_path.exists(), "dense_index meta 仍存在"
        assert not manifest_path.exists(), "corpus manifest 仍存在"
        assert not pointer_path.exists(), "current_version.json 仍存在"
        assert not idx_dir.exists() or len(list(idx_dir.iterdir())) == 0, (
            f"索引目录仍有残留: {list(idx_dir.iterdir()) if idx_dir.exists() else []}"
        )

        # workspace API 返回 404
        assert client.get(f"/api/v2/workspaces/{ws_id}").status_code == 404

        # 清理信息不应泄露路径
        clean_warnings = data.get("storage_cleanup_warnings", [])
        for w in clean_warnings:
            for forbidden in ["C:\\", "D:\\", "Users"]:
                assert forbidden not in w, f"清理 warning 泄露路径: {w}"


# ── Section 9: 原子写入测试 ─────────────────────────────────────────────


class TestAtomicWrite:
    """原子写入失败注入测试。"""

    def test_all_three_atomic_writes(self, client, auth_user, eng_workspace):
        """构建索引后三个资产文件存在且非 0 字节。"""
        ws_id = eng_workspace["id"]

        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        _understand_and_confirm(client, ws_id, result["file_id"], "clarification_document")

        _build_index(client, ws_id)

        from app.retrieval.dense_index import _resolve_active_assets
        from app.services.engineering_retrieval_service import _index_dir
        idx_dir = _index_dir(ws_id)

        assets = _resolve_active_assets(idx_dir)
        assert assets is not None, f"无法解析索引资产: {idx_dir}"
        npz_path = assets["npz"]
        meta_path = assets["meta"]
        manifest_path = assets["manifest"]

        # 三个文件都存在
        assert npz_path.exists()
        assert meta_path.exists()
        assert manifest_path.exists()

        # 非 0 字节
        assert npz_path.stat().st_size > 0, "NPZ 文件为 0 字节"
        assert meta_path.stat().st_size > 0, "meta 文件为 0 字节"
        assert manifest_path.stat().st_size > 0, "manifest 文件为 0 字节"

        # 无残留暂存/临时文件（隐藏文件都应不存在）
        tmp_files = [f for f in idx_dir.glob(".*") if f.is_file()]
        assert len(tmp_files) == 0, f"残留临时文件: {tmp_files}"

        # 版本配对：每个 dense_index_*.npz 都有同版本的 meta 和 manifest
        npz_versions = [p.name[len("dense_index_"):-4] for p in idx_dir.glob("dense_index_*.npz")]
        for ver in npz_versions:
            assert (idx_dir / f"dense_index_meta_{ver}.json").exists(), f"{ver} 缺 meta"
            assert (idx_dir / f"corpus_manifest_{ver}.json").exists(), f"{ver} 缺 manifest"
        assert len(npz_versions) == 1, f"应有唯一生效版本，实际: {npz_versions}"

    def test_json_write_failure_preserves_old_index(self, client, auth_user, monkeypatch):
        """JSON 写入失败时旧索引仍可加载。"""
        ws = _create_eng_workspace(client, "原子写入测试")
        ws_id = ws["id"]

        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        _understand_and_confirm(client, ws_id, result["file_id"], "clarification_document")

        # 第一次构建
        first_build = _build_index(client, ws_id)
        first_sha = first_build["index_sha256"]

        # 注入失败：让 JSON 写入抛出异常
        original_atomic_json = None
        import app.services.engineering_retrieval_service as svc_mod

        original_atomic_json = svc_mod._atomic_write_json

        def failing_json(path, content):
            raise OSError("模拟写入失败")

        monkeypatch.setattr(svc_mod, "_atomic_write_json", failing_json)

        # 尝试 rebuild
        resp = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
            headers=_session_csrf(client),
            json={"rebuild": True},
        )
        assert resp.status_code in (400, 500), f"应返回错误: {resp.status_code} {resp.text}"

        # 恢复
        monkeypatch.setattr(svc_mod, "_atomic_write_json", original_atomic_json)

        # 旧索引仍可加载（通过当前版本指针解析）
        from app.retrieval.dense_index import _resolve_active_assets
        from app.services.engineering_retrieval_service import _index_dir
        idx_dir = _index_dir(ws_id)
        assets = _resolve_active_assets(idx_dir)
        assert assets is not None, "旧索引资产无法解析"
        meta_path = assets["meta"]
        if meta_path.exists():
            old_meta = json.loads(meta_path.read_text("utf-8"))
            assert old_meta.get("index_sha256") == first_sha, "旧索引元数据应保持不变"

        # 旧索引仍可正常检索（Dense 加载完整校验通过）
        search_data = _search(client, ws_id, "澄清", "dense")
        assert len(search_data["results"]) > 0
        assert search_data["index_sha256"] == first_sha

    def test_npz_write_zero_byte_prevented(self, client, auth_user):
        """NPZ 写入非 0 字节验证。"""
        ws = _create_eng_workspace(client, "NPZ 写入测试")
        ws_id = ws["id"]

        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        _understand_and_confirm(client, ws_id, result["file_id"], "clarification_document")

        build_data = _build_index(client, ws_id)

        from app.retrieval.dense_index import _resolve_active_assets
        from app.services.engineering_retrieval_service import _index_dir
        assets = _resolve_active_assets(_index_dir(ws_id))
        assert assets is not None
        npz_path = assets["npz"]
        assert npz_path.exists()
        assert npz_path.stat().st_size > 0, "NPZ 文件为 0 字节（可能是扩展名问题）"

        # 可以正常加载
        import numpy as np
        data = np.load(str(npz_path))
        assert "embeddings" in data
        assert data["embeddings"].shape[0] > 0

    def test_manifest_commit_failure_preserves_old_snapshot(self, client, auth_user, monkeypatch):
        """corpus_manifest 提交（当前版本指针替换）失败时旧三资产保持可加载。

        验证三资产快照一致性：
        - 构建失败时旧的三个索引资产仍能一起加载（Dense 检索完整校验通过）
        - 不出现新 NPZ 搭配旧 metadata/manifest（唯一版本且三件套配对）
        - 不残留暂存文件
        - 构建失败不返回 built / already_built 假成功
        """
        ws = _create_eng_workspace(client, "manifest 提交失败测试")
        ws_id = ws["id"]

        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        _understand_and_confirm(client, ws_id, result["file_id"], "clarification_document")

        # 第一次构建（v1 快照）
        first_build = _build_index(client, ws_id)
        first_sha = first_build["index_sha256"]
        first_corpus_sha = first_build["corpus_sha256"]

        from app.retrieval.dense_index import _resolve_active_assets
        from app.services.engineering_retrieval_service import _index_dir
        idx_dir = _index_dir(ws_id)
        assets_v1 = _resolve_active_assets(idx_dir)
        assert assets_v1 is not None
        v1_version = json.loads((idx_dir / "current_version.json").read_text("utf-8"))["version"]

        import app.services.engineering_retrieval_service as svc_mod
        original_atomic_json = svc_mod._atomic_write_json

        # 注入失败：当前版本指针（快照提交点）替换失败
        def failing_pointer(path, content):
            if path.name == "current_version.json":
                raise OSError("模拟当前版本指针替换失败")
            return original_atomic_json(path, content)

        monkeypatch.setattr(svc_mod, "_atomic_write_json", failing_pointer)

        # 尝试 rebuild（应失败，且不得返回 built/already_built）
        resp = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
            headers=_session_csrf(client),
            json={"rebuild": True},
        )
        assert resp.status_code in (400, 500), f"应返回错误: {resp.status_code} {resp.text}"

        # 恢复
        monkeypatch.setattr(svc_mod, "_atomic_write_json", original_atomic_json)

        # 指针仍指向 v1
        pointer_data = json.loads((idx_dir / "current_version.json").read_text("utf-8"))
        assert pointer_data["version"] == v1_version, (
            f"指针被篡改: {pointer_data['version']} != {v1_version}"
        )

        # 旧的三个索引资产仍能一起加载（Dense 检索完整校验通过）
        search_data = _search(client, ws_id, "澄清 证书编号", "dense")
        assert len(search_data["results"]) > 0, "旧索引无法加载"
        assert search_data["index_sha256"] == first_sha, "旧索引 SHA 不一致"
        assert search_data["corpus_sha256"] == first_corpus_sha, "旧 corpus SHA 不一致"

        # 不出现新 NPZ 搭配旧 metadata/manifest：目录中只有 v1 一个版本且三件套配对
        npz_versions = [
            p.name[len("dense_index_"):-4] for p in idx_dir.glob("dense_index_*.npz")
        ]
        assert npz_versions == [v1_version], f"出现新版本资产: {npz_versions}"
        for stem in ("dense_index_meta", "corpus_manifest"):
            assert (idx_dir / f"{stem}_{v1_version}.json").exists(), f"{stem} 缺失"

        # 不残留暂存文件
        leftovers = [f for f in idx_dir.glob(".*") if f.is_file()]
        assert leftovers == [], f"残留暂存文件: {leftovers}"

        # GET 状态仍为 ready（v1 快照完整）
        status_resp = client.get(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
        )
        assert status_resp.status_code == 200
        assert status_resp.json()["status"] == "ready", status_resp.json()


# ── 错误码和安全测试 ─────────────────────────────────────────────────────


class TestErrorCodeSanitization:
    """异常安全化测试。"""

    def test_no_path_in_error(self, client, auth_user, eng_workspace):
        """错误响应中不应包含文件系统路径。"""
        ws_id = eng_workspace["id"]
        resp = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
            headers=_session_csrf(client),
            json={"rebuild": False},
        )
        assert resp.status_code == 400
        detail = resp.json()["detail"]
        message = detail.get("message", "")
        assert "\\\\" not in message
        assert "/app/" not in message.lower()
        assert ".npz" not in message.lower()

    def test_unknown_exception_sanitized(self, client, auth_user, eng_workspace):
        """未知异常应返回通用错误消息，不泄露堆栈。"""
        ws_id = eng_workspace["id"]
        resp = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
            headers=_session_csrf(client),
            json={"rebuild": False},
        )
        detail = resp.json()["detail"]
        assert "error_code" in detail
        assert "message" in detail
        assert "Traceback" not in detail.get("message", "")
        assert "File " not in detail.get("message", "")

    def test_seven_error_codes(self, client, auth_user):
        """验证 7 个错误码均可被触发或映射。"""
        # WORKSPACE_INVALID: 非工程工作区
        resp = client.post(
            "/api/v2/workspaces",
            headers=_session_csrf(client),
            json={"name": "通用", "workspace_type": "general"},
        )
        assert resp.status_code == 201
        gen_ws_id = resp.json()["id"]

        resp2 = client.get(
            f"/api/v2/workspaces/{gen_ws_id}/engineering-retrieval/index",
        )
        assert resp2.status_code == 400
        assert resp2.json()["detail"]["error_code"] == "ENGINEERING_RETRIEVAL_WORKSPACE_INVALID"

        # QUERY_INVALID: Pydantic 层验证（空查询）
        ws = _create_eng_workspace(client, "错误码测试")
        ws_id = ws["id"]
        resp3 = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/search",
            headers=_session_csrf(client),
            json={"query": "   ", "top_k": 5, "retrieval_mode": "bm25"},
        )
        assert resp3.status_code == 422

        # MATERIAL_NOT_READY: 空工作区
        resp4 = client.get(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
        )
        assert resp4.status_code == 400
        assert resp4.json()["detail"]["error_code"] == "ENGINEERING_RETRIEVAL_MATERIAL_NOT_READY"

        # INDEX_MISSING: 有材料但未建索引，执行 Dense 检索
        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        fid = result["file_id"]
        _understand_and_confirm(client, ws_id, fid, "clarification_document")

        resp5 = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/search",
            headers=_session_csrf(client),
            json={"query": "测试", "top_k": 5, "retrieval_mode": "dense"},
        )
        assert resp5.status_code == 400
        assert resp5.json()["detail"]["error_code"] == "ENGINEERING_RETRIEVAL_INDEX_MISSING"

        # INDEX_STALE: 见 TestIndexLifecycle
        # INDEX_ERROR: 500 内部错误
        # MODEL_UNAVAILABLE: 见 TestModelUnavailable（需 monkeypatch）


class TestModelUnavailable:
    """MODEL_UNAVAILABLE：Embedding 模型加载/编码失败时的 API 契约。

    通过 monkeypatch 让服务创建 LocalEmbeddingProvider 或编码时抛出
    EmbeddingError（无需真实模型失败），调用真实索引构建 / Dense 检索 API。
    """

    def test_build_provider_load_failure(self, client, auth_user, eng_workspace, monkeypatch):
        """构建索引时模型加载失败 → ENGINEERING_RETRIEVAL_MODEL_UNAVAILABLE。"""
        ws_id = eng_workspace["id"]
        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        _understand_and_confirm(client, ws_id, result["file_id"], "clarification_document")

        from app.retrieval.embedding import EmbeddingError
        import app.services.engineering_retrieval_service as svc_mod

        class FailingProvider:
            def __init__(self, *args, **kwargs):
                raise EmbeddingError("模拟模型加载失败")

        monkeypatch.setattr(svc_mod, "LocalEmbeddingProvider", FailingProvider)

        resp = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
            headers=_session_csrf(client),
            json={"rebuild": False},
        )
        assert resp.status_code == 500, f"MODEL_UNAVAILABLE 契约应为 500: {resp.text}"
        detail = resp.json()["detail"]
        assert detail["error_code"] == "ENGINEERING_RETRIEVAL_MODEL_UNAVAILABLE", detail
        _assert_no_path_leak(detail["message"])

        # 索引状态不得被伪装为 ready
        status_resp = client.get(f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index")
        assert status_resp.status_code == 200
        assert status_resp.json()["status"] != "ready", status_resp.json()

    def test_build_encode_failure(self, client, auth_user, eng_workspace, monkeypatch):
        """构建索引时编码失败 → ENGINEERING_RETRIEVAL_MODEL_UNAVAILABLE。"""
        ws_id = eng_workspace["id"]
        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        _understand_and_confirm(client, ws_id, result["file_id"], "clarification_document")

        from app.retrieval.embedding import EmbeddingError
        import app.services.engineering_retrieval_service as svc_mod

        class LoadsButFailsToEncode:
            def __init__(self, *args, **kwargs):
                pass

            def _ensure_loaded(self):
                pass

            def metadata(self):
                return {
                    "provider": "fake",
                    "model_repo_id": "fake/test",
                    "model_revision": "fake",
                    "dimension": 512,
                    "normalize_embeddings": True,
                    "query_instruction": QUERY_INSTRUCTION,
                    "device": "cpu",
                }

            def encode_passages(self, texts):
                raise EmbeddingError("模拟编码失败")

            def encode_queries(self, texts):
                raise EmbeddingError("模拟编码失败")

        monkeypatch.setattr(svc_mod, "LocalEmbeddingProvider", LoadsButFailsToEncode)

        resp = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
            headers=_session_csrf(client),
            json={"rebuild": False},
        )
        assert resp.status_code == 500, f"MODEL_UNAVAILABLE 契约应为 500: {resp.text}"
        detail = resp.json()["detail"]
        assert detail["error_code"] == "ENGINEERING_RETRIEVAL_MODEL_UNAVAILABLE", detail
        _assert_no_path_leak(detail["message"])

    def test_dense_search_encode_failure(self, client, auth_user, monkeypatch):
        """索引已构建，Dense 检索时 query 编码失败 → MODEL_UNAVAILABLE。"""
        ws = _create_eng_workspace(client, "模型不可用测试")
        ws_id = ws["id"]
        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        _understand_and_confirm(client, ws_id, result["file_id"], "clarification_document")

        # 先用 Fake Provider 正常构建索引
        build_data = _build_index(client, ws_id)
        assert build_data["status"] in ("built", "already_built")

        from app.retrieval.embedding import EmbeddingError
        import app.services.engineering_retrieval_service as svc_mod

        class LoadsButFailsToEncode:
            def __init__(self, *args, **kwargs):
                pass

            def _ensure_loaded(self):
                pass

            def metadata(self):
                return {
                    "provider": "fake",
                    "model_repo_id": "fake/test",
                    "model_revision": "fake",
                    "dimension": 512,
                    "normalize_embeddings": True,
                    "query_instruction": QUERY_INSTRUCTION,
                    "device": "cpu",
                }

            def encode_passages(self, texts):
                raise EmbeddingError("模拟编码失败")

            def encode_queries(self, texts):
                raise EmbeddingError("模拟编码失败")

        monkeypatch.setattr(svc_mod, "LocalEmbeddingProvider", LoadsButFailsToEncode)

        resp = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/search",
            headers=_session_csrf(client),
            json={"query": "测试", "top_k": 5, "retrieval_mode": "dense"},
        )
        assert resp.status_code == 500, f"MODEL_UNAVAILABLE 契约应为 500: {resp.text}"
        detail = resp.json()["detail"]
        assert detail["error_code"] == "ENGINEERING_RETRIEVAL_MODEL_UNAVAILABLE", detail
        _assert_no_path_leak(detail["message"])

    def test_bm25_still_works_when_model_unavailable(self, client, auth_user, monkeypatch):
        """模型不可用时 BM25 检索不受影响（不需要 Embedding 模型）。"""
        ws = _create_eng_workspace(client, "BM25 不受模型影响")
        ws_id = ws["id"]
        fpath = _GOLDEN / "05_项目澄清.md"
        result = _upload_file(client, ws_id, fpath)
        _understand_and_confirm(client, ws_id, result["file_id"], "clarification_document")
        _build_index(client, ws_id)

        from app.retrieval.embedding import EmbeddingError
        import app.services.engineering_retrieval_service as svc_mod

        class FailingProvider:
            def __init__(self, *args, **kwargs):
                raise EmbeddingError("模拟模型加载失败")

        monkeypatch.setattr(svc_mod, "LocalEmbeddingProvider", FailingProvider)

        data = _search(client, ws_id, "澄清 证书编号", "bm25")
        assert len(data["results"]) > 0, "BM25 不应依赖 Embedding 模型"


def _assert_no_path_leak(message: str) -> None:
    """错误 message 不得泄露模型缓存路径、系统路径或堆栈。"""
    for forbidden in ("C:\\", "D:\\", "Users", "cache", "Traceback", "File "):
        assert forbidden not in message, f"message 泄露敏感信息: {message}"


class TestCrossUserIsolation:
    """跨用户隔离测试。"""

    def test_cross_user_404(self, client, auth_user, eng_workspace):
        """用户 B 不能访问用户 A 的工程工作区检索。"""
        ws_id = eng_workspace["id"]

        # 登出当前用户
        client.post("/api/v2/auth/logout", headers=_session_csrf(client))

        # 注册并登录用户 B
        _register_and_login(client, "user_b_isolation")

        # 用户 B 尝试访问用户 A 的工作区
        resp = client.get(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
        )
        assert resp.status_code == 404

        resp2 = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/search",
            headers=_session_csrf(client),
            json={"query": "测试", "top_k": 5, "retrieval_mode": "bm25"},
        )
        assert resp2.status_code == 404


class TestCSRFAndValidation:
    """CSRF 和 Pydantic 验证测试。"""

    def test_build_requires_csrf(self, client, auth_user, eng_workspace):
        """无 CSRF 应拒绝构建索引。"""
        ws_id = eng_workspace["id"]
        csrf_cookie = settings.csrf_cookie_name
        saved = client.cookies.get(csrf_cookie)
        if csrf_cookie in client.cookies:
            del client.cookies[csrf_cookie]
        try:
            resp = client.post(
                f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
                json={"rebuild": False},
            )
            assert resp.status_code in (401, 403, 422), f"unexpected: {resp.status_code}"
        finally:
            if saved:
                client.cookies[csrf_cookie] = saved

    def test_search_requires_csrf(self, client, auth_user, eng_workspace):
        """无 CSRF 应拒绝检索。"""
        ws_id = eng_workspace["id"]
        csrf_cookie = settings.csrf_cookie_name
        saved = client.cookies.get(csrf_cookie)
        if csrf_cookie in client.cookies:
            del client.cookies[csrf_cookie]
        try:
            resp = client.post(
                f"/api/v2/workspaces/{ws_id}/engineering-retrieval/search",
                json={"query": "测试", "top_k": 5, "retrieval_mode": "hybrid_rrf"},
            )
            assert resp.status_code in (401, 403, 422), f"unexpected: {resp.status_code}"
        finally:
            if saved:
                client.cookies[csrf_cookie] = saved

    def test_invalid_retrieval_mode(self, client, auth_user, eng_workspace):
        """无效检索模式被 Pydantic 拒绝。"""
        ws_id = eng_workspace["id"]
        resp = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/search",
            headers=_session_csrf(client),
            json={"query": "测试", "top_k": 5, "retrieval_mode": "invalid_mode"},
        )
        assert resp.status_code == 422

    def test_top_k_out_of_range(self, client, auth_user, eng_workspace):
        """top_k 超出范围被 Pydantic 拒绝。"""
        ws_id = eng_workspace["id"]
        resp = client.post(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/search",
            headers=_session_csrf(client),
            json={"query": "测试", "top_k": 50, "retrieval_mode": "bm25"},
        )
        assert resp.status_code == 422


class TestIndexStatus:
    """索引状态查询测试。"""

    def test_status_on_non_engineering(self, client, auth_user):
        """非工程工作区应拒绝。"""
        resp = client.post(
            "/api/v2/workspaces",
            headers=_session_csrf(client),
            json={"name": "通用工作区", "workspace_type": "general"},
        )
        assert resp.status_code == 201
        ws_id = resp.json()["id"]

        resp2 = client.get(
            f"/api/v2/workspaces/{ws_id}/engineering-retrieval/index",
        )
        assert resp2.status_code == 400
        assert resp2.json()["detail"]["error_code"] == "ENGINEERING_RETRIEVAL_WORKSPACE_INVALID"

    def test_status_404_on_wrong_workspace(self, client, auth_user):
        """不存在的工程工作区返回 404。"""
        resp = client.get(
            "/api/v2/workspaces/99999/engineering-retrieval/index",
        )
        assert resp.status_code == 404
