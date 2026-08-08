#!/usr/bin/env python3
"""可复现合成材料生成脚本 v1.1.0 — 工程检测服务投标资料审查黄金案例。

- 固定数据，不调用网络，不含真实企业/人员/证书。
- 字节级可复现：相同输入 → 相同 SHA256。
- PyMuPDF (fitz) 生成中文 PDF，openpyxl 生成 Excel。
"""

from __future__ import annotations

import hashlib
import json
import os
import struct
import time
import zipfile
from datetime import date
from io import BytesIO
from pathlib import Path

import fitz

# ── 固定配置 ──────────────────────────────────────────────────────

FIXTURE_GENERATED_DATE = "2026-08-06"
GENERATOR_VERSION = "1.1.0"
DISCLAIMER = "合成演示数据，不作为真实招投标、工程、资质或法律判断依据。"

PROJECT = {
    "id": "SYN-ENG-2026-001",
    "name": "东海新区综合交通枢纽工程第三方检测服务（合成）",
    "name_short": "东海新区综合交通枢纽第三方检测服务",
    "bidder": "海岳工程检测技术有限公司（虚构）",
    "tenderer": "东海新区城市建设事务中心（虚构）",
    "deadline": "2026-09-30",
    "service_end": "2027-12-31",
    "budget": 2000000,
}

OUTPUT_DIR = Path(__file__).resolve().parents[1] / "examples" / "engineering_review_v1" / "golden_case"

# A4 尺寸 (points)
A4_W, A4_H = 595, 842
MARGIN = 50
FONT = "china-s"
FOOTER_Y = 20
BODY_SIZE = 11
TITLE_SIZE = 18
H2_SIZE = 14
LINE_H = 16          # 单行文本高度
PARA_LINE_H = 18     # 段落文本行高


def _sha256_path(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


# ── PDF 辅助 ──────────────────────────────────────────────────────


def _add_footer(page):
    """页脚免责声明（insert_text 保证不丢字）。"""
    page.insert_text((MARGIN, FOOTER_Y + 4), DISCLAIMER, fontsize=7, fontname=FONT, color=(0.4, 0.4, 0.4))


def _add_page_number(page, num, total):
    page.insert_text((A4_W / 2 - 15, FOOTER_Y + 4), f"{num}/{total}", fontsize=7, fontname=FONT, color=(0.4, 0.4, 0.4))


def _write_line(page, x, y, text, fontsize=BODY_SIZE, bold=False):
    """单行文本 — 使用 insert_text（不会因溢出静默丢弃）。"""
    page.insert_text((x, y), text, fontsize=fontsize, fontname=FONT, color=(0, 0, 0))


def _write_para(page, x, y, text, fontsize=BODY_SIZE, max_w=A4_W - 2 * MARGIN):
    """多行段落 — 逐行 insert_text，返回下一行 y 位置。"""
    chars_per_line = int(max_w / (fontsize * 0.55))
    lines = []
    for paragraph in text.split("\n"):
        para = paragraph
        while len(para) > chars_per_line:
            cut = para[:chars_per_line]
            lines.append(cut)
            para = para[chars_per_line:]
        if para:
            lines.append(para)
    for i, line in enumerate(lines):
        page.insert_text((x, y + i * PARA_LINE_H), line, fontsize=fontsize, fontname=FONT, color=(0, 0, 0))
    return y + len(lines) * PARA_LINE_H


def _draw_table(page, x, y, headers, rows, col_widths):
    """绘制表格，返回表格底部 y。"""
    row_h = 24
    # header
    cx = x
    for i, h in enumerate(headers):
        r = fitz.Rect(cx, y, cx + col_widths[i], y + row_h)
        page.draw_rect(r, color=(0, 0, 0), width=0.5)
        page.draw_rect(r, color=(0.35, 0.35, 0.35), fill=(0.35, 0.35, 0.35))
        page.insert_text((cx + 3, y + 16), h, fontsize=9, fontname=FONT, color=(1, 1, 1))
        cx += col_widths[i]
    # rows
    y += row_h
    for row in rows:
        cx = x
        for i, cell in enumerate(row):
            r = fitz.Rect(cx, y, cx + col_widths[i], y + row_h)
            page.draw_rect(r, color=(0, 0, 0), width=0.5)
            page.insert_text((cx + 3, y + 16), str(cell), fontsize=9, fontname=FONT, color=(0, 0, 0))
            cx += col_widths[i]
        y += row_h
    return y


# ── PDF 生成 ──────────────────────────────────────────────────────


def generate_tender_requirement_pdf() -> bytes:
    """01_合成招标要求.pdf — 6 页"""
    doc = fitz.open()
    total = 6
    for _ in range(total):
        doc.new_page(width=A4_W, height=A4_H)

    # 页面 0：封面
    p = doc[0]
    _add_footer(p); _add_page_number(p, 1, total)
    _write_line(p, MARGIN, 180, "工程检测服务", fontsize=24)
    _write_line(p, MARGIN, 220, "投标邀请书", fontsize=26)
    _write_line(p, MARGIN, 300, f"项目编号：{PROJECT['id']}", fontsize=13)
    _write_line(p, MARGIN, 330, f"项目名称：{PROJECT['name']}", fontsize=13)
    _write_line(p, MARGIN, 370, f"招标人：{PROJECT['tenderer']}", fontsize=13)
    _write_line(p, MARGIN, 400, f"截标日期：{PROJECT['deadline']}", fontsize=13)
    _write_line(p, MARGIN, 440, f"服务结束日期：{PROJECT['service_end']}", fontsize=13)
    _write_line(p, MARGIN, 510, DISCLAIMER, fontsize=9)
    _write_line(p, MARGIN, 535, "（封面）", fontsize=9)

    clauses = [
        ("SYN-TENDER-001", "项目名称填写要求",
         f"投标响应文件必须明确填写项目名称「{PROJECT['name']}」。项目名称为必填字段，不得留空或填写简称。"),
        ("SYN-TENDER-003", "项目负责人资质要求",
         "人员设备清单中必须指定项目负责人，负责人姓名不得为空。负责人须持有有效的工程检测相关证书。"),
        ("SYN-TENDER-005", "文件间一致性要求",
         "投标响应中填写的项目名称必须与人员设备清单中的项目名称完全一致。"),
        ("SYN-TENDER-006", "证书编号一致性",
         "投标响应中填写的项目负责人证书编号与人员清单中记录的证书编号必须一致。"),
        ("SYN-TENDER-007", "人员数量最低要求",
         "拟投入本项目的检测人员不少于 5 人（含项目负责人）。"),
        ("SYN-TENDER-008", "预算上限",
         f"本项目预算上限为 {PROJECT['budget']:,} 元（含税），超过预算上限的报价将被视为无效。"),
        ("SYN-TENDER-009", "资质有效期要求",
         "所有资质证书的有效期必须覆盖至项目服务结束日期 2027-12-31 之后。"),
        ("SYN-TENDER-010", "截标时间",
         f"投标响应签署日期不得晚于 {PROJECT['deadline']}。"),
        ("SYN-TENDER-011", "资质文件提交要求",
         "投标方必须提交有效的检测资质证书附件，作为资格审查的必要文件。"),
        ("SYN-TENDER-012", "人员设备清单要求",
         "投标方须提交完整的人员与设备配置清单，包括但不限于人员岗位、证书编号、设备名称和型号。"),
        ("SYN-TENDER-013", "设备数量最低要求",
         "拟投入本项目的主要检测设备不少于 4 台。"),
        ("SYN-TENDER-014", "设备校准有效期",
         f"全部主要检测设备的校准有效期必须覆盖至 {PROJECT['service_end']} 之后。"),
    ]

    per_page = [clauses[0:3], clauses[3:6], clauses[6:9], clauses[9:11], clauses[11:12]]
    for pi, group in enumerate(per_page):
        p = doc[pi + 1]
        _add_footer(p); _add_page_number(p, pi + 2, total)
        y = MARGIN + 30
        if pi == 0:
            _write_line(p, MARGIN, MARGIN, "招标条款", fontsize=TITLE_SIZE)
            y += 40
        for code, title, desc in group:
            _write_line(p, MARGIN, y, f"{code}：{title}", fontsize=H2_SIZE)
            y += 22
            y = _write_para(p, MARGIN + 10, y, desc, fontsize=BODY_SIZE) + 10

    out = doc.tobytes()
    doc.close()
    return out


def generate_bid_response_pdf() -> bytes:
    """02_合成投标响应.pdf — 4 页"""
    doc = fitz.open()
    total = 4
    for _ in range(total):
        doc.new_page(width=A4_W, height=A4_H)

    # 页面 0：关键字段表
    p = doc[0]
    _add_footer(p); _add_page_number(p, 1, total)
    _write_line(p, MARGIN, MARGIN, "投标响应文件", fontsize=TITLE_SIZE)
    _write_line(p, MARGIN, MARGIN + 40, f"投标人：{PROJECT['bidder']}", fontsize=12)
    _write_line(p, MARGIN, MARGIN + 60, "项目名称：[  留空——此处有意未填写  ]", fontsize=12)

    y = _draw_table(p, MARGIN, MARGIN + 85,
                    ["字段", "值", "备注"],
                    [
                        ["项目名称", "（空）", "SYN-REQ-001"],
                        ["项目负责人", "林海", "虚构姓名"],
                        ["证书编号", "SYN-JC-24018", "SYN-EQ-002 对照值"],
                        ["总报价（元）", "2,150,000", "SYN-NUM-002 超出预算"],
                        ["签署日期", "2026-10-02", "SYN-DATE-002 超期"],
                    ],
                    [120, 180, 140])

    # 页面 1-3：正文（不引入与标准答案矛盾的数值）
    content_pages = [
        [
            "海岳工程检测技术有限公司（以下简称「我方」）就东海新区城市建设事务中心发布的"
            "本项目提交以下投标响应。",
            "我方承诺按照招标文件中规定的服务范围、技术要求和质量标准，"
            "在指定时间内完成全部第三方检测服务。",
            "我方拟投入的检测人员配置详见《人员设备清单》。"
            "项目负责人为林海，证书编号 SYN-JC-24018，持有工程质量检测上岗证书。",
        ],
        [
            "项目团队包括林海（项目负责人）、周岚（检测工程师）、"
            "顾远（检测工程师）、沈宁（检测工程师）、赵刚（检测员）、王芳（检测员），"
            "共计 6 名成员（本项目投标响应中的团队构成，实际人员清单数据以《人员设备清单》为准）。",
            "我方提交的总报价为人民币 2,150,000 元（含税），"
            "报价包括全部检测服务、设备租赁和人员费用。",
        ],
        [
            "本投标响应签署日期为 2026-10-02。",
            "我方拟投入的主要检测设备配置详见《人员设备清单》。",
            "我方承诺在合同签订后按照招标要求提交详细的人员和检测方案，"
            "并接受招标方的监督和验收。",
            "本文件为合成演示数据，不包含真实企业信息、公章或签名。",
            "合成电子印章示意——不具法律效力。",
        ],
    ]
    for pi, lines in enumerate(content_pages):
        p = doc[pi + 1]
        _add_footer(p); _add_page_number(p, pi + 2, total)
        y = MARGIN + 20
        for line in lines:
            y = _write_para(p, MARGIN, y, line, fontsize=BODY_SIZE) + 8

    out = doc.tobytes()
    doc.close()
    return out


def generate_qualification_pdf() -> bytes:
    """04_合成资质附件.pdf — 2 页。第1页文本层，第2页光栅扫描图。"""
    from reportlab.lib.pagesizes import A4 as rl_A4
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.colors import HexColor

    # ── 第 1 页：文本层 ──
    doc = fitz.open()
    doc.new_page(width=A4_W, height=A4_H)
    p1 = doc[0]
    _add_footer(p1); _add_page_number(p1, 1, 2)
    _write_line(p1, MARGIN, MARGIN, "资质证明文件", fontsize=TITLE_SIZE)
    _write_line(p1, MARGIN, MARGIN + 40, f"投标人：{PROJECT['bidder']}", fontsize=12)
    _write_line(p1, MARGIN, MARGIN + 70, "附件名称：合成检测能力证明材料", fontsize=H2_SIZE)

    y = _draw_table(p1, MARGIN, MARGIN + 100,
                    ["项目", "内容"],
                    [
                        ["资质类型", "CMA 检验检测机构资质认定（合成）"],
                        ["有效期至", "2027-06-30"],
                        ["证书编号", "见第 2 页扫描示意图（无文本层）"],
                        ["发证机构", "合成资质认定办公室（虚构）"],
                    ],
                    [120, 300])
    _write_para(p1, MARGIN, y + 10, "有效期 2027-06-30 早于项目服务结束日期 2027-12-31。")
    _write_para(p1, MARGIN, y + 35, "证书编号参见第 2 页合成扫描示意。该页为光栅图像，不含机器可提取文本层。")
    _write_para(p1, MARGIN, y + 60, "（翻至第 2 页查看合成证书扫描示意）")

    # ── 第 2 页：fitz 生成 CJK 扫描图 → PNG → 嵌入（无文本层）──
    scan_doc = fitz.open()
    scan_doc.new_page(width=A4_W, height=A4_H)
    sp = scan_doc[0]

    # 免责声明（fitz CJK 字体渲染，再光栅化）
    sp.insert_text((40, 30), DISCLAIMER, fontsize=8, fontname=FONT, color=(0.4, 0.4, 0.4))

    # 外框
    sp.draw_rect(fitz.Rect(40, 45, A4_W - 40, A4_H - 45), color=(0.3, 0.3, 0.3), width=3)
    sp.draw_rect(fitz.Rect(45, 50, A4_W - 45, A4_H - 50), color=(0.3, 0.3, 0.3), width=1)

    # 标题（英文保持可读）
    sp.insert_text((A4_W / 2 - 100, 100), "Certificate of Qualification", fontsize=18, fontname="helv", color=(0, 0, 0))
    sp.insert_text((A4_W / 2 - 80, 130), "(Synthetic Sample — No Legal Authority)", fontsize=10, fontname="helv", color=(0, 0, 0))
    sp.draw_line(fitz.Point(100, 155), fitz.Point(A4_W - 100, 155), color=(0.3, 0.3, 0.3))

    # 证书编号（★肉眼可见，光栅化后无文本层）
    sp.insert_text((A4_W / 2 - 100, 200), "Cert No: SYN-CMA-2026-014", fontsize=16, fontname="helv", color=(0, 0, 0))

    # 字段（英文）
    sp.insert_text((120, 260), "Entity: Haiyue Engineering Testing Co., Ltd. (fictional)", fontsize=12, fontname="helv", color=(0, 0, 0))
    sp.insert_text((120, 300), "Valid until: 2027-06-30", fontsize=12, fontname="helv", color=(0, 0, 0))
    sp.insert_text((120, 340), "Issue date: 2022-07-01", fontsize=12, fontname="helv", color=(0, 0, 0))

    # 中文免责（用 CJK 字体渲染后光栅化，清晰可见）
    sp.insert_text((80, 40), DISCLAIMER, fontsize=9, fontname=FONT, color=(0.3, 0.3, 0.3))

    # 底部中文免责（在 rendered image 中清晰可见）
    sp.insert_text((40, A4_H - 40), DISCLAIMER, fontsize=8, fontname=FONT, color=(0.4, 0.4, 0.4))

    # 虚构印章
    sp.draw_circle(fitz.Point(A4_W - 140, 320), 50, color=(0.8, 0.2, 0.2), width=2)
    sp.insert_text((A4_W - 185, 340), "SYNTHETIC CERTIFICATION", fontsize=7, fontname="helv", color=(0.8, 0.2, 0.2))
    sp.insert_text((A4_W - 175, 325), "(fictional)", fontsize=7, fontname="helv", color=(0.8, 0.2, 0.2))
    sp.insert_text((40, A4_H - 60), "Synthetic certificate scan — No legal authority", fontsize=8, fontname="helv", color=(0.5, 0.5, 0.5))

    # 光栅化 scan_doc 为 PNG → 嵌入目标 PDF
    # 先对中间 PDF 做确定性处理，再渲染为 PNG（确保 PNG 字节级稳定）
    scan_bytes = _deterministic_pdf(scan_doc.tobytes())
    scan_doc.close()
    scan_doc2 = fitz.open("pdf", scan_bytes)
    pix = scan_doc2[0].get_pixmap(dpi=120)
    img_bytes = pix.tobytes("png")
    scan_doc2.close()

    doc.new_page(width=A4_W, height=A4_H)
    p2 = doc[1]
    p2.insert_image(p2.rect, stream=img_bytes)
    # 叠加页码
    p2.insert_text((A4_W / 2 - 12, FOOTER_Y + 4), "2/2", fontsize=7, fontname=FONT, color=(0.4, 0.4, 0.4))

    out = doc.tobytes()
    doc.close()
    return out


# ── Excel（确定性生成）────────────────────────────────────────────


def generate_personnel_excel() -> bytes:
    """03_人员设备清单.xlsx — 字节级可复现。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    disc_font = Font(bold=True, color="FF0000", size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # Sheet 1: 项目概况
    ws1 = wb.active; ws1.title = "项目概况"
    ws1.merge_cells("A1:D1")
    ws1["A1"] = DISCLAIMER; ws1["A1"].font = disc_font; ws1["A1"].alignment = wrap_align
    ws1.row_dimensions[1].height = 24
    rows1 = [
        ("项目编号", PROJECT["id"]),
        ("项目名称", PROJECT["name_short"]),
        ("投标人", PROJECT["bidder"]),
        ("预算上限", PROJECT["budget"]),
        ("截标日期", "2026-09-30"),
        ("服务结束日期", PROJECT["service_end"]),
    ]
    for r, (k, v) in enumerate(rows1, 2):
        ws1.cell(row=r, column=1, value=k).border = border
        c = ws1.cell(row=r, column=2, value=v); c.border = border
        if k == "预算上限": c.number_format = "#,##0"
    ws1.column_dimensions["A"].width = 18; ws1.column_dimensions["B"].width = 52
    ws1.freeze_panes = "A2"

    # Sheet 2: 人员清单
    ws2 = wb.create_sheet("人员清单")
    ws2.merge_cells("A1:F1"); ws2["A1"] = DISCLAIMER; ws2["A1"].font = disc_font; ws2["A1"].alignment = wrap_align
    ws2.row_dimensions[1].height = 24
    for c, h in enumerate(["序号", "姓名", "岗位", "证书编号", "专业", "备注"], 1):
        cell = ws2.cell(row=2, column=c, value=h); cell.font = header_font; cell.fill = header_fill; cell.border = border
    data2 = [
        [1, "", "项目负责人", "SYN-JC-24081", "工程检测", "姓名留空"],
        [2, "周岚", "检测工程师", "SYN-JC-24082", "结构检测", ""],
        [3, "顾远", "检测工程师", "SYN-JC-24083", "材料检测", ""],
        [4, "沈宁", "检测员", "SYN-JC-24084", "现场取样", ""],
    ]
    for r, row in enumerate(data2, 3):
        for c, v in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=v); cell.border = border
    for c in range(1, 7): ws2.cell(row=3, column=c).fill = warn_fill
    ws2.cell(row=8, column=1, value="合计").font = Font(bold=True)
    ws2.cell(row=8, column=2, value=4)
    for c in ("A","B","C","D","E","F"): ws2.column_dimensions[c].width = 16
    ws2.freeze_panes = "A3"

    # Sheet 3: 设备清单（1+1+1=3）
    ws3 = wb.create_sheet("设备清单")
    ws3.merge_cells("A1:G1"); ws3["A1"] = DISCLAIMER; ws3["A1"].font = disc_font; ws3["A1"].alignment = wrap_align
    ws3.row_dimensions[1].height = 24
    for c, h in enumerate(["序号", "设备名称", "型号", "数量", "生产厂家", "校准有效期", "备注"], 1):
        cell = ws3.cell(row=2, column=c, value=h); cell.font = header_font; cell.fill = header_fill; cell.border = border
    from datetime import date as dt_date
    equipment = [
        [1, "回弹仪", "SYN-HT225", 1, "合成仪器厂", dt_date(2028, 6, 30), ""],
        [2, "钢筋扫描仪", "SYN-RS100", 1, "合成仪器厂", dt_date(2026, 12, 31), "校准早于服务结束"],
        [3, "超声波检测仪", "SYN-UT200", 1, "合成仪器厂", dt_date(2028, 3, 15), ""],
    ]
    for r, row in enumerate(equipment, 3):
        for c, v in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=v); cell.border = border
            if c == 6: cell.number_format = "YYYY-MM-DD"
            if c == 4: cell.number_format = "0"
    for c in range(1, 8): ws3.cell(row=4, column=c).fill = warn_fill
    ws3.cell(row=7, column=1, value="合计").font = Font(bold=True)
    ws3.cell(row=7, column=4, value=3)
    for c in ("A","B","C","D","E","F","G"): ws3.column_dimensions[c].width = 18
    ws3.freeze_panes = "A3"

    # Sheet 4: 数据说明
    ws4 = wb.create_sheet("数据说明")
    ws4.merge_cells("A1:A1")
    ws4["A1"] = DISCLAIMER; ws4["A1"].font = disc_font; ws4["A1"].alignment = wrap_align
    ws4.row_dimensions[1].height = 24
    notes = [
        "合成数据说明",
        "1. 人员清单共 4 人，低于最低 5 人要求",
        "2. 项目负责人姓名有意留空",
        "3. 证书编号 SYN-JC-24081（人员）<> SYN-JC-24018（投标响应）",
        "4. 项目名称缺少「工程」二字",
        "5. 主要设备共 3 台，低于最低 4 台要求；数量明细：1+1+1=3",
        "6. 钢筋扫描仪校准 2026-12-31，早于服务结束 2027-12-31",
        "7. 所有编号、型号、厂家为合成虚构值",
    ]
    for i, n in enumerate(notes, 2):
        ws4.cell(row=i, column=1, value=n)
    ws4.column_dimensions["A"].width = 80
    ws4.freeze_panes = "A2"

    # 确定性输出：固定 ZIP 条目时间戳
    buf = BytesIO()
    # 固定 openpyxl 时间戳，避免 core.xml 中 created/modified 变化
    wb.properties.created = "2026-08-06T00:00:00Z"
    wb.properties.modified = "2026-08-06T00:00:00Z"
    wb.save(buf)
    return _deterministic_xlsx(buf.getvalue())


def _deterministic_xlsx(raw: bytes) -> bytes:
    """重写 XLSX ZIP，固定时间戳；然后用 xml 库固化 core.xml 的 created/modified。"""
    import xml.etree.ElementTree as ET
    stamp = (2026, 8, 6, 0, 0, 0)
    in_buf = BytesIO(raw)
    out_buf = BytesIO()
    with zipfile.ZipFile(in_buf, "r") as zin:
        with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_STORED) as zout:
            for name in sorted(zin.namelist()):
                info = zipfile.ZipInfo(name)
                info.date_time = stamp
                info.compress_type = zipfile.ZIP_STORED
                data = zin.read(name)
                if name == "docProps/core.xml":
                    data = _fix_core_xml_timestamps(data)
                zout.writestr(info, data)
    return out_buf.getvalue()


def _fix_core_xml_timestamps(data: bytes) -> bytes:
    """用 xml.etree 解析 core.xml，替换 created/modified 为固定值。"""
    import xml.etree.ElementTree as ET
    FIXED_TS = "2026-08-06T00:00:00Z"
    # 注册命名空间
    ns = {
        "dcterms": "http://purl.org/dc/terms/",
        "xsi": "http://www.w3.org/2001/XMLSchema-instance",
        "dc": "http://purl.org/dc/elements/1.1/",
        "cp": "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    }
    for prefix, uri in ns.items():
        ET.register_namespace(prefix, uri)
    root = ET.fromstring(data)
    for el in root.iter():
        tag_local = el.tag.rsplit("}", 1)[-1] if "}" in el.tag else el.tag
        if tag_local in ("created", "modified") and el.text:
            el.text = FIXED_TS
    result = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    return result


def _deterministic_pdf(raw: bytes) -> bytes:
    """替换 PDF 文档 ID 为合法固定值，保证字节级可复现。"""
    import re
    # PDF /ID[<32hex><32hex>] — 两个合法 32 位十六进制字符串
    fixed_id = b"<0123456789ABCDEF0123456789ABCDEF><FEDCBA9876543210FEDCBA9876543210>"
    pattern = re.compile(rb'/ID\s*\[<[0-9A-Fa-f]{32}>\s*<[0-9A-Fa-f]{32}>\]')
    return pattern.sub(b'/ID[' + fixed_id + b']', raw)


# ── Markdown ──────────────────────────────────────────────────────


def generate_clarification_md() -> str:
    return f"""# 项目澄清文件

> {DISCLAIMER}

## 澄清通知

根据投标文件初审结果，东海新区城市建设事务中心（虚构）就以下事项进行澄清。

## 澄清条款

### SYN-CLAR-001：人员资质证据要求

项目负责人的所有资质字段（姓名、证书编号）必须能够在投标响应文件或人员设备清单中
定位到具体证据，标注来源文件的页码或单元格位置。仅填写数值不足以构成可审计证据。

### SYN-CLAR-002：证书编号可追溯要求

资质证书编号必须提供可追溯证据，能够在资质附件中定位到具体的页码或证书扫描件区域。
若证书扫描件无法通过文本层可靠识别编号，则必须由人工进行核验。

## 确认事项

- 最低人员数量仍为 5 人
- 最低设备数量为 4 台
- 设备校准有效期需覆盖至 {PROJECT['service_end']}
- 预算上限不变（{PROJECT['budget']:,} 元）
- 扫描件无法可靠识别时必须人工复核
- 澄清文件优先于此前冲突的合成描述
- 本澄清文件不构成真实法律意见

---
*生成日期：{FIXTURE_GENERATED_DATE}，使用固定合成数据。生成器版本：{GENERATOR_VERSION}*
"""


# ── manifest ──────────────────────────────────────────────────────


def _build_manifest(file_paths: dict[str, Path]) -> dict:
    manifest = {
        "case_id": "SYN-ENG-2026-001",
        "manifest_version": "1.1.0",
        "generator_version": GENERATOR_VERSION,
        "generated_date": FIXTURE_GENERATED_DATE,
        "rule_pack_id": "engineering_bid_review_v1",
        "rule_pack_version": "1.1.0",
        "disclaimer": DISCLAIMER,
        "files": [],
    }
    for fn, path in file_paths.items():
        entry = {
            "filename": fn,
            "role": _role_for(fn),
            "mime_type": _mime_for(fn),
            "sha256": _sha256_path(path),
            "size_bytes": path.stat().st_size,
        }
        if fn.endswith(".pdf"):
            doc = fitz.open(path)
            entry["page_count"] = doc.page_count
            if fn == "04_合成资质附件.pdf":
                entry["text_layer_mode"] = "page1_text_page2_raster"
                entry["ocr_required"] = True
                entry["ocr_required_pages"] = [2]
            else:
                entry["text_layer_mode"] = "full"
                entry["ocr_required"] = False
            entry["disclaimer_visible"] = True
            doc.close()
        elif fn.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            entry["sheet_names"] = wb.sheetnames
            entry["formula_count"] = 0
            entry["disclaimer_visible"] = True
            wb.close()
        elif fn.endswith(".md"):
            text = path.read_text(encoding="utf-8")
            entry["encoding"] = "utf-8"
            entry["clause_codes"] = ["SYN-CLAR-001", "SYN-CLAR-002"]
            entry["disclaimer_visible"] = True
        manifest["files"].append(entry)
    return manifest


def _role_for(fn: str) -> str:
    return {
        "01_合成招标要求.pdf": "tender_requirement",
        "02_合成投标响应.pdf": "bid_response",
        "03_人员设备清单.xlsx": "personnel_equipment_data",
        "04_合成资质附件.pdf": "qualification_attachment",
        "05_项目澄清.md": "clarification_document",
    }[fn]


def _mime_for(fn: str) -> str:
    if fn.endswith(".pdf"): return "application/pdf"
    if fn.endswith(".xlsx"): return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return "text/markdown"


# ── 主入口 ────────────────────────────────────────────────────────


def generate_case(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    files = {}

    tender_pdf = _deterministic_pdf(generate_tender_requirement_pdf())
    p = output_dir / "01_合成招标要求.pdf"; p.write_bytes(tender_pdf); files[p.name] = p

    bid_pdf = _deterministic_pdf(generate_bid_response_pdf())
    p = output_dir / "02_合成投标响应.pdf"; p.write_bytes(bid_pdf); files[p.name] = p

    qual_pdf = _deterministic_pdf(generate_qualification_pdf())
    p = output_dir / "04_合成资质附件.pdf"; p.write_bytes(qual_pdf); files[p.name] = p

    xlsx = generate_personnel_excel()
    p = output_dir / "03_人员设备清单.xlsx"; p.write_bytes(xlsx); files[p.name] = p

    md = generate_clarification_md()
    p = output_dir / "05_项目澄清.md"; p.write_text(md, encoding="utf-8"); files[p.name] = p

    manifest = _build_manifest(files)
    (output_dir / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def main():
    generate_case(OUTPUT_DIR)
    print(f"材料已生成至 {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
